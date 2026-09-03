#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


ALLOWED_TYPES = {
    "Cable/Adapter",
    "Dongle",
    "Handheld",
    "Key",
    "OBD1",
    "Other",
    "Probe",
    "Tablet",
    "VCI",
}

STOPWORDS = {
    "and",
    "for",
    "with",
    "from",
    "that",
    "this",
    "your",
    "you",
    "all",
    "new",
    "the",
    "tool",
    "tools",
    "scan",
    "scanner",
    "reader",
    "code",
    "car",
    "cars",
    "vehicle",
    "vehicles",
    "obd",
    "obd2",
    "obdii",
}


def _norm_asin(value: object) -> str:
    return str(value).strip().upper()


def _norm_text(value: object) -> str:
    return str(value or "").strip().lower()


def _tokenize(*parts: object) -> set[str]:
    text = " ".join(_norm_text(p) for p in parts if p is not None)
    tokens = re.findall(r"[a-z0-9]+", text)
    out: set[str] = set()
    for t in tokens:
        if len(t) < 3:
            continue
        if t in STOPWORDS:
            continue
        out.add(t)
    return out


def _to_float(value: object) -> float:
    s = str(value or "").strip()
    if not s:
        return 0.0
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except Exception:
        return 0.0


def _is_missing_text(value: object) -> bool:
    if value is None:
        return True
    s = str(value).strip().lower()
    return s == "" or s == "nan"


def _missing_text_mask(series: pd.Series) -> pd.Series:
    return series.isna() | series.astype(str).str.strip().str.lower().isin(["", "nan"])


def _latest_month_obd2_files(amazon_obd2_dir: Path, max_months: int = 24) -> list[Path]:
    if not amazon_obd2_dir.exists():
        return []
    pattern = re.compile(r"amazon_obd2_(\d{6})\.xlsx$")
    files: list[tuple[str, Path]] = []
    for p in amazon_obd2_dir.iterdir():
        if not p.is_file():
            continue
        m = pattern.match(p.name)
        if not m:
            continue
        files.append((m.group(1), p))
    files.sort(key=lambda x: x[0], reverse=True)
    return [p for _, p in files[:max_months]]


def _load_type_mapping(spec_path: Path) -> pd.DataFrame:
    if not spec_path.exists():
        raise SystemExit(f"Type taxonomy file not found: {spec_path}")
    spec = pd.read_excel(spec_path)
    if "ASIN" not in spec.columns or "Type" not in spec.columns:
        raise SystemExit(f"Type taxonomy file must include ASIN and Type columns: {spec_path}")
    spec = spec.copy()
    spec["ASIN"] = spec["ASIN"].map(_norm_asin)
    spec["Type"] = spec["Type"].astype(str).str.strip()
    spec = spec[spec["ASIN"].ne("") & spec["ASIN"].ne("NAN")].copy()
    spec = spec[spec["Type"].isin(ALLOWED_TYPES)].copy()
    return spec


def _load_obd2_catalog(amazon_obd2_dir: Path) -> pd.DataFrame:
    files = _latest_month_obd2_files(amazon_obd2_dir)
    rows: list[pd.DataFrame] = []
    for p in files:
        try:
            df = pd.read_excel(p, usecols=["ASIN", "Type", "Title", "Brand", "URL", "Price"])
        except Exception:
            continue
        df = df.copy()
        df["ASIN"] = df["ASIN"].map(_norm_asin)
        df["Type"] = df["Type"].astype(str).str.strip()
        df["Brand"] = df["Brand"].map(_norm_text)
        df["Title"] = df["Title"].astype(str)
        df["URL"] = df["URL"].astype(str)
        df["Price"] = df["Price"].map(_to_float)
        df = df[df["ASIN"].ne("") & df["ASIN"].ne("NAN")].copy()
        df = df[df["Type"].isin(ALLOWED_TYPES)].copy()
        rows.append(df)
    if not rows:
        return pd.DataFrame(columns=["ASIN", "Type", "Title", "Brand", "URL", "Price"])
    out = pd.concat(rows, ignore_index=True)
    out = out.drop_duplicates(subset=["ASIN"], keep="first")
    return out


def _keyword_rule(title: str, url: str, price: float) -> tuple[str | None, str | None]:
    text = f"{_norm_text(title)} {_norm_text(url)}"

    def has(*keywords: str) -> bool:
        return any(k in text for k in keywords)

    if has("obd1"):
        return "OBD1", "keyword:obd1"
    if has("vci"):
        return "VCI", "keyword:vci"
    if has("tablet", "8in", "10in", "12in", "touchscreen", "android 12", "android tablet"):
        return "Tablet", "keyword:tablet"
    if has("key programmer", "key programming", "immobilizer", "key fob", "all keys lost"):
        return "Key", "keyword:key"
    if has("probe", "oscilloscope probe", "scope probe"):
        return "Probe", "keyword:probe"
    if has("bluetooth", "wireless", "wifi", "wi-fi", "dongle"):
        return "Dongle", "keyword:dongle"
    if has("cable", "adapter cable", "extension cable", "connector", "adapter"):
        return "Cable/Adapter", "keyword:cable_adapter"
    if has("scan tool", "diagnostic tool", "code reader", "check engine"):
        return "Handheld", "keyword:handheld"
    if price >= 380 and has("scanner"):
        return "Tablet", "keyword:price_tablet_hint"
    return None, None


def _build_brand_profiles(typed_catalog: pd.DataFrame) -> tuple[dict[tuple[str, str], set[str]], dict[str, tuple[str, float, int]]]:
    tokens_by_brand_type: dict[tuple[str, str], Counter] = defaultdict(Counter)
    counts_by_brand_type: Counter = Counter()
    counts_by_brand: Counter = Counter()

    for _, row in typed_catalog.iterrows():
        brand = _norm_text(row.get("Brand"))
        typ = str(row.get("Type", "")).strip()
        if not brand or not typ:
            continue
        counts_by_brand_type[(brand, typ)] += 1
        counts_by_brand[brand] += 1
        tokens = _tokenize(row.get("Title"), row.get("URL"))
        for t in tokens:
            tokens_by_brand_type[(brand, typ)][t] += 1

    profile_tokens: dict[tuple[str, str], set[str]] = {}
    for key, cnt in tokens_by_brand_type.items():
        profile_tokens[key] = {tok for tok, n in cnt.items() if n >= 2}

    dominant_brand_type: dict[str, tuple[str, float, int]] = {}
    for brand, total in counts_by_brand.items():
        pairs = [(bt[1], c) for bt, c in counts_by_brand_type.items() if bt[0] == brand]
        if not pairs:
            continue
        pairs.sort(key=lambda x: x[1], reverse=True)
        typ, c = pairs[0]
        dominant_brand_type[brand] = (typ, c / max(total, 1), total)
    return profile_tokens, dominant_brand_type


def _token_profile_match(
    row_tokens: set[str],
    brand: str,
    profile_tokens: dict[tuple[str, str], set[str]],
) -> tuple[str | None, str | None]:
    if not brand or not row_tokens:
        return None, None
    candidates = [(bt[1], toks) for bt, toks in profile_tokens.items() if bt[0] == brand and toks]
    if not candidates:
        return None, None

    best_type = None
    best_score = 0.0
    best_overlap = 0
    for typ, toks in candidates:
        inter = row_tokens & toks
        if not inter:
            continue
        overlap = len(inter)
        score = overlap / max(len(row_tokens), 1)
        if score > best_score:
            best_score = score
            best_overlap = overlap
            best_type = typ

    if best_type and best_score >= 0.50 and best_overlap >= 3:
        return best_type, f"token_model_match(score={best_score:.2f},overlap={best_overlap})"
    return None, None


def _write_excel(path: Path, df: pd.DataFrame, dry_run: bool) -> None:
    if dry_run:
        print(f"[dry-run] Would write: {path} ({len(df):,} rows)")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)


def _write_manual_review_excel(path: Path, df: pd.DataFrame, dry_run: bool) -> None:
    if dry_run:
        print(f"[dry-run] Would write: {path} ({len(df):,} rows)")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Type Review", index=False)
        ws = writer.book["Type Review"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False

        header_fill = PatternFill("solid", fgColor="17365D")
        input_fill = PatternFill("solid", fgColor="FFF2CC")
        missing_fill = PatternFill("solid", fgColor="F4CCCC")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        reviewed_type_col = df.columns.get_loc("Reviewed Type") + 1
        for row in range(2, len(df) + 2):
            ws.cell(row=row, column=reviewed_type_col).fill = input_fill
        reviewed_letter = ws.cell(row=1, column=reviewed_type_col).column_letter
        valid_values = ",".join(sorted(ALLOWED_TYPES))
        validation = DataValidation(
            type="list",
            formula1=f'"{valid_values}"',
            allow_blank=False,
        )
        validation.error = "Choose one of the approved Type values from the dropdown."
        validation.errorTitle = "Invalid Type"
        validation.prompt = "Required: select the final Type for this ASIN."
        validation.promptTitle = "Reviewed Type"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        if len(df):  # openpyxl rejects an inverted range like "B2:B1" when the review set is empty
            validation.add(f"{reviewed_letter}2:{reviewed_letter}{len(df) + 1}")
            ws.conditional_formatting.add(
                f"{reviewed_letter}2:{reviewed_letter}{len(df) + 1}",
                FormulaRule(formula=[f'LEN(TRIM({reviewed_letter}2))=0'], fill=missing_fill),
            )

        widths = {
            "ASIN": 16,
            "Title": 72,
            "Brand": 20,
            "Price": 12,
            "URL": 42,
            "predicted_type": 18,
            "confidence": 12,
            "reason": 44,
            "action": 24,
            "Reviewed Type": 20,
            "Review Notes": 34,
        }
        for index, name in enumerate(df.columns, start=1):
            ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = widths.get(name, 16)
            for row in range(2, len(df) + 2):
                ws.cell(row=row, column=index).alignment = Alignment(
                    vertical="top",
                    wrap_text=name in {"Title", "URL", "reason", "Review Notes"},
                )


def _write_csv(path: Path, df: pd.DataFrame, dry_run: bool) -> None:
    if dry_run:
        print(f"[dry-run] Would write: {path} ({len(df):,} rows)")
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Auto-categorize script/runs/YYYYMM/extra.xlsx with deterministic hybrid rules. "
            "Writes review artifacts and optionally appends high-confidence mappings to amazon_scanner_type.xlsx."
        )
    )
    parser.add_argument("--month", required=True, help="YYYYMM")
    parser.add_argument("--extra-xlsx", type=Path, required=True)
    parser.add_argument("--spec", type=Path, required=True, help="Path to amazon_scanner_type.xlsx")
    parser.add_argument("--amazon-obd2-dir", type=Path, required=True)
    parser.add_argument("--run-dir", type=Path, required=True)
    parser.add_argument("--writeback-threshold", type=float, default=0.90)
    parser.add_argument("--review-threshold", type=float, default=0.75)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if not re.fullmatch(r"\d{6}", str(args.month)):
        raise SystemExit("--month must be YYYYMM")
    if not (0 <= args.review_threshold <= 1 and 0 <= args.writeback_threshold <= 1):
        raise SystemExit("Threshold values must be in [0, 1].")
    if args.review_threshold > args.writeback_threshold:
        raise SystemExit("--review-threshold must be <= --writeback-threshold.")

    extra_xlsx = args.extra_xlsx.resolve()
    spec_path = args.spec.resolve()
    run_dir = args.run_dir.resolve()
    amazon_obd2_dir = args.amazon_obd2_dir.resolve()

    if not extra_xlsx.exists():
        raise SystemExit(f"extra.xlsx not found: {extra_xlsx}")

    spec_df = _load_type_mapping(spec_path)
    typed_catalog = _load_obd2_catalog(amazon_obd2_dir)

    spec_asin_to_type = dict(zip(spec_df["ASIN"], spec_df["Type"], strict=False))
    hist_asin_to_type = dict(zip(typed_catalog["ASIN"], typed_catalog["Type"], strict=False))
    profile_tokens, dominant_brand_type = _build_brand_profiles(typed_catalog)

    extra = pd.read_excel(extra_xlsx).copy()
    for col in ("ASIN", "Title", "Brand", "URL", "Price"):
        if col not in extra.columns:
            extra[col] = ""

    extra["ASIN"] = extra["ASIN"].map(_norm_asin)
    extra["Title"] = extra["Title"].astype(str).str.strip()
    extra["Brand"] = extra["Brand"].astype(str).str.strip()
    extra["BrandNorm"] = extra["Brand"].map(_norm_text)
    extra["URL"] = extra["URL"].astype(str).str.strip()
    extra["Price"] = extra["Price"].map(_to_float)
    extra = extra[extra["ASIN"].ne("") & extra["ASIN"].ne("NAN")].copy()
    extra = extra.drop_duplicates(subset=["ASIN"], keep="first")

    rows: list[dict[str, object]] = []
    for _, row in extra.iterrows():
        asin = str(row["ASIN"])
        brand = str(row["BrandNorm"])
        brand_raw = str(row["Brand"]).strip()
        title = str(row["Title"])
        url = str(row["URL"])
        price = float(row["Price"])
        existing_type = spec_asin_to_type.get(asin)

        predicted_type: str | None = None
        confidence = 0.0
        reason = "unresolved"

        if existing_type:
            predicted_type = existing_type
            confidence = 1.00
            reason = "exact_asin_type_mapping"
            # Even when an ASIN is already mapped, surface strong contradictory evidence for review.
            alt_type: str | None = None
            alt_conf = 0.0
            alt_reason = ""
            if asin in hist_asin_to_type:
                alt_type = hist_asin_to_type[asin]
                alt_conf = 1.00
                alt_reason = "exact_asin_obd2_history"
            else:
                row_tokens = _tokenize(title, url)
                token_type, token_reason = _token_profile_match(row_tokens, brand, profile_tokens)
                if token_type:
                    alt_type = token_type
                    alt_conf = 0.93
                    alt_reason = token_reason or "token_model_match"
                else:
                    kw_type, kw_reason = _keyword_rule(title, url, price)
                    if kw_type:
                        alt_type = kw_type
                        alt_conf = 0.90
                        alt_reason = kw_reason or "keyword_rule"
                    else:
                        dom = dominant_brand_type.get(brand)
                        if dom and dom[1] >= 0.70 and dom[2] >= 10:
                            alt_type = dom[0]
                            alt_conf = 0.78
                            alt_reason = f"brand_dominant_type(share={dom[1]:.2f},n={dom[2]})"

            if (
                alt_type
                and alt_type != existing_type
                and alt_conf >= args.review_threshold
            ):
                predicted_type = alt_type
                confidence = alt_conf
                reason = f"existing_mapping_conflict:{alt_reason}"
        elif asin in hist_asin_to_type:
            predicted_type = hist_asin_to_type[asin]
            confidence = 1.00
            reason = "exact_asin_obd2_history"
        else:
            row_tokens = _tokenize(title, url)
            token_type, token_reason = _token_profile_match(row_tokens, brand, profile_tokens)
            if token_type:
                predicted_type = token_type
                confidence = 0.93
                reason = token_reason or "token_model_match"
            else:
                kw_type, kw_reason = _keyword_rule(title, url, price)
                if kw_type:
                    predicted_type = kw_type
                    confidence = 0.90
                    reason = kw_reason or "keyword_rule"
                else:
                    dom = dominant_brand_type.get(brand)
                    if dom and dom[1] >= 0.70 and dom[2] >= 10:
                        predicted_type = dom[0]
                        confidence = 0.78
                        reason = f"brand_dominant_type(share={dom[1]:.2f},n={dom[2]})"

        action = "unresolved"
        conflict = False
        if predicted_type:
            if predicted_type not in ALLOWED_TYPES:
                predicted_type = None
                confidence = 0.0
                reason = "invalid_predicted_type"
                action = "unresolved"
            elif existing_type:
                if existing_type == predicted_type:
                    action = "already_mapped"
                else:
                    action = "conflict_existing_mapping"
                    conflict = True
            elif confidence >= args.writeback_threshold:
                action = "auto_append_mapping"
            elif confidence >= args.review_threshold:
                action = "review_recommended"
            else:
                action = "review_low_confidence"

        rows.append(
            {
                "ASIN": asin,
                "Title": title,
                "Brand": brand_raw,
                "URL": url,
                "Price": price,
                "predicted_type": predicted_type or "",
                "confidence": confidence,
                "reason": reason,
                "existing_type": existing_type or "",
                "conflict": conflict,
                "action": action,
            }
        )

    result = pd.DataFrame(rows)

    classification_path = run_dir / "extra_auto_classification.xlsx"
    review_path = run_dir / "extra_review_required.xlsx"
    manual_review_path = run_dir / "extra_manual_type_review.xlsx"
    appended_path = run_dir / "type_mapping_auto_appended.csv"
    conflict_path = run_dir / "type_mapping_conflicts.csv"

    review = result[
        result["conflict"]
        | result["predicted_type"].astype(str).str.strip().eq("")
        | result["confidence"].lt(args.review_threshold)
    ].copy()

    manual_review = result[
        ~result["action"].isin(["auto_append_mapping", "already_mapped"])
    ][
        [
            "ASIN",
            "Title",
            "Brand",
            "Price",
            "URL",
            "predicted_type",
            "confidence",
            "reason",
            "action",
        ]
    ].copy()
    manual_review["Reviewed Type"] = ""
    manual_review["Review Notes"] = ""

    to_append = result[
        result["action"].eq("auto_append_mapping")
        & result["predicted_type"].astype(str).str.strip().ne("")
        & ~result["conflict"]
    ][["ASIN", "Title", "Brand", "Price", "URL", "predicted_type", "confidence", "reason"]].rename(columns={"predicted_type": "Type"})

    conflicts = result[result["conflict"]].copy()

    _write_excel(classification_path, result, args.dry_run)
    _write_excel(review_path, review, args.dry_run)
    _write_manual_review_excel(manual_review_path, manual_review, args.dry_run)
    _write_csv(appended_path, to_append, args.dry_run)
    _write_csv(conflict_path, conflicts, args.dry_run)

    appended_count = 0
    metadata_backfilled_rows = 0
    if not args.dry_run:
        before_path = run_dir / "amazon_scanner_type.before_auto.xlsx"
        shutil.copy2(spec_path, before_path)
        print(f"[info] Backed up type taxonomy: {before_path}")

        spec_raw = pd.read_excel(spec_path).copy()
        for col in ("ASIN", "Type", "Title", "Brand", "Price", "URL"):
            if col not in spec_raw.columns:
                spec_raw[col] = ""
        spec_raw["ASIN"] = spec_raw["ASIN"].map(_norm_asin)
        spec_raw["Type"] = spec_raw["Type"].astype(str).str.strip()
        spec_raw["Title"] = spec_raw["Title"].astype(str).str.strip()
        spec_raw["Brand"] = spec_raw["Brand"].astype(str).str.strip()
        spec_raw["URL"] = spec_raw["URL"].astype(str).str.strip()
        spec_raw["Price"] = pd.to_numeric(spec_raw["Price"], errors="coerce")

        append_rows = pd.DataFrame({col: [pd.NA] * len(to_append) for col in spec_raw.columns})
        append_rows["ASIN"] = to_append["ASIN"].values
        append_rows["Type"] = to_append["Type"].values
        append_rows["Title"] = to_append["Title"].values
        append_rows["Brand"] = to_append["Brand"].values
        append_rows["Price"] = to_append["Price"].values
        append_rows["URL"] = to_append["URL"].values

        combined = pd.concat([spec_raw, append_rows], ignore_index=True, sort=False)
        combined = combined[combined["ASIN"].ne("") & combined["ASIN"].ne("NAN")].copy()
        combined = combined.drop_duplicates(subset=["ASIN"], keep="last")

        # Backfill missing metadata for ASINs present in this run's extra classification.
        meta = result[["ASIN", "Title", "Brand", "Price", "URL"]].drop_duplicates(subset=["ASIN"], keep="first")
        meta["ASIN"] = meta["ASIN"].map(_norm_asin)
        meta["Title"] = meta["Title"].astype(str).str.strip()
        meta["Brand"] = meta["Brand"].astype(str).str.strip()
        meta["URL"] = meta["URL"].astype(str).str.strip()
        meta["Price"] = pd.to_numeric(meta["Price"], errors="coerce")
        meta = meta.set_index("ASIN")

        # Also use OBD2 typed catalog metadata so reruns can repair already-appended rows.
        cat_meta = typed_catalog[["ASIN", "Title", "Brand", "Price", "URL"]].drop_duplicates(subset=["ASIN"], keep="first")
        cat_meta["ASIN"] = cat_meta["ASIN"].map(_norm_asin)
        cat_meta["Title"] = cat_meta["Title"].astype(str).str.strip()
        cat_meta["Brand"] = cat_meta["Brand"].astype(str).str.strip()
        cat_meta["URL"] = cat_meta["URL"].astype(str).str.strip()
        cat_meta["Price"] = pd.to_numeric(cat_meta["Price"], errors="coerce")
        cat_meta = cat_meta.set_index("ASIN")

        filled_row_index: set[int] = set()

        def _fill_from_source(source: pd.DataFrame) -> None:
            for col in ("Title", "Brand", "URL"):
                src = source[col]
                src = src.where(~src.map(_is_missing_text), pd.NA)
                src_vals = combined["ASIN"].map(src)
                mask = _missing_text_mask(combined[col]) & src_vals.notna()
                if mask.any():
                    combined.loc[mask, col] = src_vals[mask]
                    filled_row_index.update(combined.index[mask].tolist())

            src_price = pd.to_numeric(source["Price"], errors="coerce")
            src_price = src_price.where(src_price.gt(0), pd.NA)
            src_price_vals = pd.to_numeric(combined["ASIN"].map(src_price), errors="coerce")
            current_price = pd.to_numeric(combined["Price"], errors="coerce")
            mask_price = (current_price.isna() | current_price.le(0)) & src_price_vals.notna()
            if mask_price.any():
                combined.loc[mask_price, "Price"] = src_price_vals[mask_price]
                filled_row_index.update(combined.index[mask_price].tolist())

        _fill_from_source(meta)
        _fill_from_source(cat_meta)

        metadata_backfilled_rows = len(filled_row_index)

        with pd.ExcelWriter(spec_path, engine="openpyxl") as writer:
            combined.to_excel(writer, index=False)
        appended_count = len(to_append)
        print(f"[info] Appended {appended_count:,} high-confidence ASIN Type mappings to: {spec_path}")
        if metadata_backfilled_rows:
            print(f"[info] Backfilled metadata fields (Title/Brand/Price/URL) for {metadata_backfilled_rows:,} ASIN rows in taxonomy.")

    print("=== Auto Categorize Extra Summary ===")
    print(f"month: {args.month}")
    print(f"extra_rows: {len(result):,}")
    print(f"auto_append_candidates: {len(to_append):,}")
    print(f"review_required_rows: {len(review):,}")
    print(f"manual_review_rows: {len(manual_review):,}")
    print(f"conflicts: {len(conflicts):,}")
    if args.dry_run:
        print("writeback: dry-run (no taxonomy file changes)")
    else:
        print(f"writeback_appended_rows: {appended_count:,}")
        print(f"writeback_metadata_backfilled_rows: {metadata_backfilled_rows:,}")
    print(f"classification: {classification_path}")
    print(f"review_required: {review_path}")
    print(f"manual_review: {manual_review_path}")
    print(f"type_mapping_auto_appended: {appended_path}")
    print(f"type_mapping_conflicts: {conflict_path}")
    print("====================================")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
