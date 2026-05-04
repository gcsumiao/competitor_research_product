#!/usr/bin/env python3
"""
Generate the two *formatted* deliverables from the raw script outputs:

- Amazon Competitor Report.xlsx  (raw)
- summary.xlsx                  (raw)

using last month's formatted workbooks as templates:

- <DATE> Amazon Competitor Report <Month> Innova Adjusted.xlsx
- <DATE> Amazon Competitor Analysis <Month>.xlsx

This script does NOT run the notebooks. It only automates the final "copy/paste into
template" step.
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
import tempfile
import zipfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as exc:  # pragma: no cover
        raise SystemExit(
            "Missing dependency: openpyxl\n"
            "Install with: python3 -m pip install openpyxl"
        ) from exc


_require_openpyxl()
from openpyxl import load_workbook  # noqa: E402
from openpyxl.chart import BarChart, PieChart, Reference  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.utils.cell import range_boundaries  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402


MONTH_FULL_RE = re.compile(
    r"\b(January|February|March|April|May|June|July|August|September|October|November|December)\s*'\d{2}\b"
)
MONTH_ABBR_RE = re.compile(r"\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*'\d{2}\b")


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _as_path(p: str | Path) -> Path:
    return p if isinstance(p, Path) else Path(p)


def _yyyymm_to_period(month: str) -> dt.date:
    if not re.fullmatch(r"\d{6}", month):
        raise SystemExit(f"--month must be YYYYMM, got: {month}")
    y = int(month[:4])
    m = int(month[4:])
    if m < 1 or m > 12:
        raise SystemExit(f"Invalid month in --month: {month}")
    return dt.date(y, m, 1)


def _month_labels(month: str) -> tuple[str, str]:
    d = _yyyymm_to_period(month)
    full = f"{d.strftime('%B')} '{d.strftime('%y')}"
    abbr = f"{d.strftime('%b')} '{d.strftime('%y')}"
    return full, abbr


def _rolling_12_month_headers(month: str) -> list[str]:
    end = _yyyymm_to_period(month)
    # 12 months ending at end month
    headers = []
    y, m = end.year, end.month
    for offset in range(11, -1, -1):
        mm = m - offset
        yy = y
        while mm <= 0:
            mm += 12
            yy -= 1
        d = dt.date(yy, mm, 1)
        headers.append(f"{d.strftime('%B')} '{d.strftime('%y')}")
    return headers


def _coerce_number(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s.lower() in {"n/a", "na", "nan"}:
        return None
    s = s.replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(s)
    except Exception:
        return None


def _iter_cells(ws, max_rows: int = 2000, max_cols: int = 50):
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            yield ws.cell(r, c)


def _find_cell_text(ws, text: str, max_rows: int = 2000, max_cols: int = 50):
    needle = _norm(text)
    for cell in _iter_cells(ws, max_rows=max_rows, max_cols=max_cols):
        if cell.value is None:
            continue
        if _norm(cell.value) == needle:
            return cell
    return None


def _find_row_by_values(ws, required: Iterable[str], max_rows: int = 2000, max_cols: int = 50) -> int | None:
    required_norm = {_norm(v) for v in required}
    for r in range(1, min(ws.max_row, max_rows) + 1):
        present = set()
        for c in range(1, min(ws.max_column, max_cols) + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            present.add(_norm(v))
        if required_norm.issubset(present):
            return r
    return None


def _read_row(ws, row: int, start_col: int = 1, max_col: int = 60) -> list[Any]:
    out = []
    for c in range(start_col, max_col + 1):
        out.append(ws.cell(row, c).value)
    while out and (out[-1] is None or str(out[-1]).strip() == ""):
        out.pop()
    return out


def _read_table(ws, header_row: int, start_col: int = 1, max_col: int = 60) -> tuple[list[str], list[dict[str, Any]]]:
    header_vals = _read_row(ws, header_row, start_col=start_col, max_col=max_col)
    headers = [str(h).strip() for h in header_vals if h is not None and str(h).strip() != ""]
    if not headers:
        return [], []
    rows: list[dict[str, Any]] = []
    r = header_row + 1
    while r <= ws.max_row:
        row_vals = []
        empty = True
        for i in range(len(headers)):
            v = ws.cell(r, start_col + i).value
            row_vals.append(v)
            if v is not None and str(v).strip() != "":
                empty = False
        if empty:
            break
        rows.append({headers[i]: row_vals[i] for i in range(len(headers))})
        r += 1
    return headers, rows


def _find_second_header_row(ws, first_cell_value: str, start_row: int = 2, max_rows: int = 5000) -> int | None:
    needle = _norm(first_cell_value)
    for r in range(start_row, min(ws.max_row, max_rows) + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        if _norm(v) == needle:
            return r
    return None


def _clear_block(ws, start_row: int, start_col: int, nrows: int, ncols: int):
    for r in range(start_row, start_row + nrows):
        for c in range(start_col, start_col + ncols):
            ws.cell(r, c).value = None


def _replace_month_labels_in_titles(wb, old_full: str | None, old_abbr: str | None, new_full: str, new_abbr: str):
    for ws in wb.worksheets:
        for cell in _iter_cells(ws, max_rows=200, max_cols=30):
            if not isinstance(cell.value, str):
                continue
            s = cell.value
            if old_full:
                s = s.replace(old_full, new_full)
            if old_abbr:
                s = s.replace(old_abbr, new_abbr)
            cell.value = s


def _rewrite_relationship_targets_inplace(xlsx_path: Path) -> None:
    """
    Excel is picky about relationship Targets.

    Some openpyxl versions write absolute in-package Targets like:
      Target="/xl/tables/table1.xml"
    which triggers Excel's "We found a problem with some content..." repair dialog.

    Fix by rewriting to relative Targets, e.g.:
      workbook rels:   "worksheets/sheet1.xml"
      worksheet rels:  "../tables/table1.xml"
    """
    xlsx_path = Path(xlsx_path)
    rels_re = re.compile(r'Target="([^"]+)"')

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = Path(tmp.name)
        try:
            with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    data = zin.read(name)
                    if not name.endswith(".rels"):
                        zout.writestr(name, data)
                        continue

                    try:
                        text = data.decode("utf-8")
                    except Exception:
                        # Leave binary/non-utf8 rels alone.
                        zout.writestr(name, data)
                        continue

                    is_workbook_rels = name == "xl/_rels/workbook.xml.rels"

                    def repl(m: re.Match) -> str:
                        target = m.group(1)
                        new_target = target
                        if target.startswith("/xl/"):
                            tail = target[len("/xl/") :]
                            new_target = tail if is_workbook_rels else f"../{tail}"
                        elif target.startswith("/"):
                            new_target = target.lstrip("/")
                        return f'Target="{new_target}"'

                    new_text, _nsubs = rels_re.subn(repl, text)
                    # Canonicalize relationship XML to avoid Excel repair prompts triggered by
                    # openpyxl's namespace-prefixed serialization.
                    # Example openpyxl output:
                    #   <ns0:Relationships xmlns:ns0="..."><ns0:Relationship ... /></ns0:Relationships>
                    # Prefer:
                    #   <Relationships xmlns="..."><Relationship .../></Relationships>
                    new_text = re.sub(
                        r"<ns0:Relationships\\s+xmlns:ns0=\"([^\"]+)\"",
                        r'<Relationships xmlns="\\1"',
                        new_text,
                    )
                    new_text = new_text.replace("</ns0:Relationships>", "</Relationships>")
                    new_text = new_text.replace("<ns0:Relationship", "<Relationship").replace("</ns0:Relationship>", "</Relationship>")

                    # Normalize XML declaration (Excel templates typically use UTF-8 + standalone).
                    new_text = re.sub(r"^<\\?xml[^>]*\\?>\\s*", "", new_text)
                    new_text = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n' + new_text

                    if new_text == text:
                        zout.writestr(name, data)
                    else:
                        zout.writestr(name, new_text.encode("utf-8"))

            tmp_path.replace(xlsx_path)
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass


def _strip_visual_parts_inplace(xlsx_path: Path) -> None:
    """
    openpyxl cannot reliably round-trip complex Excel visuals (charts/drawings/images) from templates.

    In this repo, formatted templates can contain many charts/drawings. When openpyxl saves the file
    after we fill values, Excel may show a repair dialog. The most reliable fix is to strip visuals
    from the generated output while keeping the table formatting + values intact.
    """
    xlsx_path = Path(xlsx_path)
    remove_prefixes = ("xl/drawings/", "xl/charts/", "xl/media/")

    def _should_remove_part(name: str) -> bool:
        return any(name.startswith(p) for p in remove_prefixes)

    # Regex helpers (avoid XML re-serialization; keep file mostly byte-identical elsewhere).
    override_re = re.compile(r'<Override\\s+PartName=\"([^\"]+)\"[^>]*?/>')
    # Match both prefixed and non-prefixed relationship tags.
    rel_re = re.compile(r"<(?:\\w+:)?Relationship[^>]*?/>")
    # Remove drawing markers from worksheet XML (these would otherwise dangle after stripping parts).
    sheet_drawing_re = re.compile(r"<(?:\\w+:)?(?:drawing|legacyDrawing|legacyDrawingHF)[^>]*?/>")

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = Path(tmp.name)
        try:
            with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    if _should_remove_part(name):
                        continue

                    data = zin.read(name)

                    if name == "[Content_Types].xml":
                        try:
                            text = data.decode("utf-8")
                        except Exception:
                            zout.writestr(name, data)
                            continue

                        # Remove Overrides that reference parts we are stripping.
                        # Use a direct regex on the Override tag rather than a callback; it's more robust
                        # across the different serializer styles (Excel vs openpyxl).
                        new_text = re.sub(
                            r'<Override\s+PartName="/xl/(?:drawings|charts|media)/[^"]+"[^>]*/>',
                            "",
                            text,
                        )
                        zout.writestr(name, new_text.encode("utf-8"))
                        continue

                    if name.endswith(".rels"):
                        try:
                            text = data.decode("utf-8")
                        except Exception:
                            zout.writestr(name, data)
                            continue

                        def keep_rel(m: re.Match) -> str:
                            s = m.group(0)
                            if "relationships/drawing" in s or "relationships/chart" in s or "relationships/image" in s:
                                return ""
                            if "Target=\"../drawings/" in s or "Target=\"../charts/" in s or "Target=\"../media/" in s:
                                return ""
                            if "Target=\"drawings/" in s or "Target=\"charts/" in s or "Target=\"media/" in s:
                                return ""
                            return s

                        new_text = rel_re.sub(keep_rel, text)
                        zout.writestr(name, new_text.encode("utf-8"))
                        continue

                    if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                        try:
                            text = data.decode("utf-8")
                        except Exception:
                            zout.writestr(name, data)
                            continue
                        new_text = sheet_drawing_re.sub("", text)
                        zout.writestr(name, new_text.encode("utf-8"))
                        continue

                    zout.writestr(name, data)

            tmp_path.replace(xlsx_path)
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass

    # Second pass: ensure [Content_Types].xml doesn't reference stripped parts.
    _remove_visual_overrides_from_content_types_inplace(xlsx_path)


def _remove_visual_overrides_from_content_types_inplace(xlsx_path: Path) -> None:
    """
    Remove stale Overrides for stripped visual parts (drawings/charts/media) from [Content_Types].xml.

    Excel will often show a repair dialog if [Content_Types].xml contains Overrides pointing to
    missing parts.
    """
    xlsx_path = Path(xlsx_path)
    override_pat = re.compile(r'<Override\s+PartName="/xl/(?:drawings|charts|media)/[^"]+"[^>]*/>')

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        names = zin.namelist()
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp_path = Path(tmp.name)
        try:
            with zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for name in names:
                    data = zin.read(name)
                    if name != "[Content_Types].xml":
                        zout.writestr(name, data)
                        continue
                    try:
                        text = data.decode("utf-8")
                    except Exception:
                        zout.writestr(name, data)
                        continue
                    new_text = override_pat.sub("", text)
                    zout.writestr(name, new_text.encode("utf-8"))
            tmp_path.replace(xlsx_path)
        finally:
            if tmp_path.exists():
                try:
                    tmp_path.unlink()
                except Exception:
                    pass


_MARKET_TOTALS_CACHE: dict[str, dict[str, dict[str, float]]] = {}
_AMAZON_OBD2_DIR_OVERRIDE: Path | None = None
_OBD2_ASIN_CACHE: dict[str, dict[str, dict[str, Any]]] = {}


def _get_amazon_obd2_dir() -> Path:
    if _AMAZON_OBD2_DIR_OVERRIDE is not None:
        return _AMAZON_OBD2_DIR_OVERRIDE
    base_dir = Path(__file__).resolve().parent.parent
    return base_dir / "amazon_obd2"


def _rolling_12_months_yyyymm(month: str) -> list[str]:
    end = _yyyymm_to_period(month)
    y, m = end.year, end.month
    out: list[str] = []
    for offset in range(11, -1, -1):
        mm = m - offset
        yy = y
        while mm <= 0:
            mm += 12
            yy -= 1
        out.append(f"{yy}{mm:02d}")
    return out


def _shift_month(month: str, offset: int) -> str:
    period = _yyyymm_to_period(month)
    y = period.year
    m = period.month + offset
    while m <= 0:
        m += 12
        y -= 1
    while m > 12:
        m -= 12
        y += 1
    return f"{y}{m:02d}"


def _compute_market_totals(month: str, key_prefix: str) -> dict[str, float]:
    """
    Compute per-month total market revenue/units for the rolling-12 header months,
    using amazon_obd2_YYYYMM.xlsx files.

    key_prefix:
      - 'Monthly Revenue' -> totals of Monthly Revenue (or Price*Monthly Sales fallback)
      - 'Monthly Sales'   -> totals of Monthly Sales
    """
    # Cache must include amazon_obd2 source directory because sanity checks may override it.
    cache_key = f"{month}|{_get_amazon_obd2_dir().resolve()}"
    cache = _MARKET_TOTALS_CACHE.setdefault(cache_key, {})
    if key_prefix in cache:
        return cache[key_prefix]

    obd2_dir = _get_amazon_obd2_dir()
    months_yyyymm = _rolling_12_months_yyyymm(month)
    labels = _rolling_12_month_headers(month)
    totals: dict[str, float] = {}

    for label, yyyymm in zip(labels, months_yyyymm, strict=True):
        path = obd2_dir / f"amazon_obd2_{yyyymm}.xlsx"
        if not path.exists():
            continue
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header]

        def idx(name: str) -> int | None:
            for i, h in enumerate(headers):
                if _norm(h) == _norm(name):
                    return i
            return None

        sales_i = idx("Monthly Sales")
        revenue_i = idx("Monthly Revenue")
        price_i = idx("Price")

        if sales_i is None:
            wb.close()
            continue

        total = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            sales = _coerce_number(row[sales_i] if sales_i < len(row) else None) if sales_i is not None else None
            if sales is None:
                continue
            if key_prefix == "Monthly Sales":
                total += sales
                continue
            if revenue_i is not None:
                rev = _coerce_number(row[revenue_i] if revenue_i < len(row) else None)
                if rev is not None:
                    total += rev
                    continue
            price = _coerce_number(row[price_i] if price_i is not None and price_i < len(row) else None)
            if price is not None:
                total += price * sales

        wb.close()
        totals[label] = total

    cache[key_prefix] = totals
    return totals


def _load_obd2_asin_index(yyyymm: str) -> dict[str, dict[str, Any]]:
    cache_key = f"{yyyymm}|{_get_amazon_obd2_dir().resolve()}"
    if cache_key in _OBD2_ASIN_CACHE:
        return _OBD2_ASIN_CACHE[cache_key]

    path = _get_amazon_obd2_dir() / f"amazon_obd2_{yyyymm}.xlsx"
    if not path.exists():
        _OBD2_ASIN_CACHE[cache_key] = {}
        return {}

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header]
    header_index = {headers[i]: i for i in range(len(headers))}
    asin_i = header_index.get("ASIN")
    sales_i = header_index.get("Monthly Sales")
    revenue_i = header_index.get("Monthly Revenue")
    if asin_i is None:
        wb.close()
        _OBD2_ASIN_CACHE[cache_key] = {}
        return {}

    out: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        asin = row[asin_i] if asin_i < len(row) else None
        if asin is None or str(asin).strip() == "":
            continue
        out[str(asin).strip()] = {
            "Monthly Sales": row[sales_i] if sales_i is not None and sales_i < len(row) else None,
            "Monthly Revenue": row[revenue_i] if revenue_i is not None and revenue_i < len(row) else None,
        }

    wb.close()
    _OBD2_ASIN_CACHE[cache_key] = out
    return out


@dataclass(frozen=True)
class RawTables:
    revenue: list[dict[str, Any]]
    units: list[dict[str, Any]] | None


@dataclass(frozen=True)
class OverrideBundle:
    report_rows: dict[tuple[str, str], dict[str, Any]]
    analysis_metrics: dict[tuple[str, str], dict[str, Any]]


def _split_two_tables(ws, header_first_cell: str) -> RawTables:
    h1 = header_first_cell
    header_row_1 = _find_second_header_row(ws, h1, start_row=1) or 1
    header_row_2 = _find_second_header_row(ws, h1, start_row=header_row_1 + 1)
    _, rev_rows = _read_table(ws, header_row_1)
    unit_rows = None
    if header_row_2:
        _, unit_rows = _read_table(ws, header_row_2)
    return RawTables(revenue=rev_rows, units=unit_rows)


def _normalize_report_brand_display(raw_name: str, template_names: Iterable[str] | None = None) -> str:
    template_idx = {_norm(name): name for name in (template_names or [])}
    raw_norm = _norm(raw_name)
    if raw_norm in template_idx:
        return template_idx[raw_norm]

    special = {
        "autel": "Autel",
        "blcktec": "BLCKTEC",
        "bluedriver": "Bluedriver",
        "diesel laptops": "Diesel Laptops",
        "fixd": "FIXD",
        "foxwell": "Foxwell",
        "generic": "Generic",
        "icarsoft": "iCarsoft",
        "kingbolen": "KINGBOLEN",
        "launch": "Launch",
        "mucar": "Mucar",
        "obdeleven": "OBDeleven",
        "obdlink": "OBDLink",
        "opus | ivs": "OPUS | IVS",
        "otofix": "OTOFIX",
        "topdon": "TOPDON",
        "thinkcar": "Thinkcar",
        "veepeak": "Veepeak",
        "vgate": "Vgate",
        "xtool": "XTOOL",
        "z automotive": "Z Automotive",
    }
    if raw_norm in special:
        return special[raw_norm]
    return " ".join(part.capitalize() for part in str(raw_name).split())


def _normalize_analysis_brand_display(raw_name: str, template_names: Iterable[str] | None = None) -> str:
    template_idx = {_norm(name): name for name in (template_names or [])}
    raw_norm = _norm(raw_name)
    if raw_norm in template_idx:
        return template_idx[raw_norm]

    special = {
        "autel": "AUTEL",
        "launch": "LAUNCH",
        "topdon": "TOPDON",
        "xtool": "XTOOL",
    }
    if raw_norm in special:
        return special[raw_norm]
    return _normalize_report_brand_display(raw_name, template_names)


def _load_override_bundle(path: Path | None) -> OverrideBundle:
    if path is None or not Path(path).exists():
        return OverrideBundle(report_rows={}, analysis_metrics={})

    wb = load_workbook(path, data_only=True)
    report_rows: dict[tuple[str, str], dict[str, Any]] = {}
    analysis_metrics: dict[tuple[str, str], dict[str, Any]] = {}

    if "report_asin_overrides" in wb.sheetnames:
        headers, rows = _read_table(wb["report_asin_overrides"], header_row=1)
        for row in rows:
            target_sheet = _norm(row.get("target_sheet"))
            asin = _norm(row.get("asin"))
            if not target_sheet or not asin:
                continue
            report_rows[(target_sheet, asin)] = row

    if "analysis_metric_overrides" in wb.sheetnames:
        headers, rows = _read_table(wb["analysis_metric_overrides"], header_row=1)
        for row in rows:
            target_sheet = _norm(row.get("target_sheet"))
            category = _norm(row.get("category"))
            if not target_sheet or not category:
                continue
            analysis_metrics[(target_sheet, category)] = row

    wb.close()
    return OverrideBundle(report_rows=report_rows, analysis_metrics=analysis_metrics)


def _apply_report_row_overrides(
    rows: list[dict[str, Any]],
    *,
    target_sheet: str,
    overrides: OverrideBundle,
) -> list[dict[str, Any]]:
    if not overrides.report_rows:
        return rows

    out: list[dict[str, Any]] = []
    for row in rows:
        asin = _norm(row.get("ASIN"))
        override = overrides.report_rows.get((_norm(target_sheet), asin)) or overrides.report_rows.get(("*", asin))
        if not override:
            out.append(row)
            continue

        updated = dict(row)
        field_map = {
            "product_name": "Title",
            "item_number": "Item #",
            "fulfillment": "Fulfillment",
            "type": "Type",
            "avg_price": "Price",
            "estimated_12mo_revenue": "12mo Revenue",
            "monthly_rev": "Monthly Revenue",
            "estimated_12mo_units": "12mo Units",
            "monthly_unit_sales": "Monthly Sales",
            "review_count": "Review Count",
            "tool_rating": "Reviews Rating",
            "avg_rating": "Ave Rating",
            "url": "URL",
        }
        for override_key, raw_key in field_map.items():
            value = override.get(override_key)
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            updated[raw_key] = value
        out.append(updated)
    return out


def _apply_analysis_metric_override(
    row: dict[str, Any],
    *,
    target_sheet: str,
    category: str,
    overrides: OverrideBundle,
) -> dict[str, Any]:
    override = overrides.analysis_metrics.get((_norm(target_sheet), _norm(category)))
    if not override:
        return row

    updated = dict(row)
    field_map = {
        "avg_price": "Avg Price",
        "avg_price_mom": "Avg Price MoM",
        "avg_price_yoy": "Avg Price YoY",
        "quantity_mo": "Quantity/Mo",
        "qty_by_pct": "Qty by %",
        "qty_mom": "Qty MoM",
        "qty_yoy": "Qty YoY",
        "revenue_mo": "Revenue/Mo",
        "revenue_by_pct": "Revenue by %",
        "revenue_mom": "Revenue MoM",
        "revenue_yoy": "Revenue YoY",
    }
    for override_key, target_key in field_map.items():
        value = override.get(override_key)
        if value is None or (isinstance(value, str) and value.strip() == ""):
            continue
        updated[target_key] = value
    return updated


def _worksheet_brand_sheets(sheetnames: Iterable[str]) -> list[str]:
    special = {"summary", "rolling 12 mo", "top 50", "innova", "innova 1p", "innova 3p"}
    return [name for name in sheetnames if _norm(name) not in special]


def _clone_tables_from_archetype(src_ws, dst_ws, *, name_prefix: str) -> None:
    for idx, table in enumerate(src_ws.tables.values(), start=1):
        display_name = re.sub(r"[^A-Za-z0-9_]", "", f"{name_prefix}_{idx}")[:250] or f"T{idx}"
        if display_name[0].isdigit():
            display_name = f"T{display_name}"
        style = None
        if table.tableStyleInfo is not None:
            style = TableStyleInfo(
                name=table.tableStyleInfo.name,
                showFirstColumn=table.tableStyleInfo.showFirstColumn,
                showLastColumn=table.tableStyleInfo.showLastColumn,
                showRowStripes=table.tableStyleInfo.showRowStripes,
                showColumnStripes=table.tableStyleInfo.showColumnStripes,
            )
        new_table = Table(displayName=display_name, ref=table.ref)
        if style is not None:
            new_table.tableStyleInfo = style
        dst_ws.add_table(new_table)


def _copy_brand_sheet_from_archetype(wb, *, archetype_name: str, new_title: str) -> str:
    archetype_ws = wb[archetype_name]
    copied = wb.copy_worksheet(archetype_ws)
    copied.title = new_title
    copied._charts = []
    copied._images = []
    _clone_tables_from_archetype(archetype_ws, copied, name_prefix=new_title)
    return copied.title


def _sync_analysis_brand_sheets(tpl_wb, desired_sheet_names: list[str]) -> list[str]:
    fixed = {
        "Summary",
        "Top 50",
        "Tablet Total",
        "Tablet $800+",
        "Tablet $400-$800",
        "Tablet $400-",
        "Handheld Total",
        "Handheld $75+",
        "Handheld $75-",
        "Dongle",
        "Other Tools",
    }
    brand_sheets = [name for name in tpl_wb.sheetnames if name not in fixed]
    if not brand_sheets:
        return []

    archetype_name = brand_sheets[0]
    existing = {_norm(name): name for name in brand_sheets}
    desired_norm = {_norm(name) for name in desired_sheet_names}

    for name in list(brand_sheets):
        if _norm(name) not in desired_norm:
            tpl_wb.remove(tpl_wb[name])

    resolved: list[str] = []
    for desired in desired_sheet_names:
        current = existing.get(_norm(desired))
        if current and current in tpl_wb.sheetnames:
            resolved.append(current)
            continue
        resolved.append(_copy_brand_sheet_from_archetype(tpl_wb, archetype_name=archetype_name, new_title=desired))
    return resolved


def _anchor_marker_to_cell(marker) -> str:
    return f"{get_column_letter(marker.col + 1)}{marker.row + 1}"


def _chart_anchor_cells(ws) -> list[str]:
    cells: list[str] = []
    for chart in getattr(ws, "_charts", []):
        anchor = getattr(chart, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is not None:
            cells.append(_anchor_marker_to_cell(marker))
    return cells


def _drop_invalid_tables(wb) -> None:
    for ws in wb.worksheets:
        bad_table_names: list[str] = []
        for table in ws.tables.values():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            except Exception:
                bad_table_names.append(table.displayName)
                continue
            headers = [ws.cell(min_row, c).value for c in range(min_col, max_col + 1)]
            if any(not isinstance(header, str) for header in headers):
                bad_table_names.append(table.displayName)
        for table_name in bad_table_names:
            try:
                del ws.tables[table_name]
            except Exception:
                pass


def _find_section_end_row(ws, *, start_row: int, start_col: int = 1) -> int:
    r = start_row
    while r <= ws.max_row:
        value = ws.cell(r, start_col).value
        if value is None or str(value).strip() == "":
            break
        if _norm(value).startswith("rank by"):
            break
        r += 1
    return r - 1


def _find_header_map(ws, header_row: int, start_col: int = 1, max_col: int = 40) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for c in range(start_col, max_col + 1):
        value = ws.cell(header_row, c).value
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        mapping[_norm(text)] = c
    return mapping


def _rebuild_analysis_charts(ws, *, anchor_cells: list[str] | None = None) -> None:
    header_row = _find_row_by_values(ws, required=["Avg Price", "Revenue/Mo"], max_rows=300)
    if not header_row:
        return

    section_end = _find_section_end_row(ws, start_row=header_row + 1)
    if section_end <= header_row:
        return

    anchors = anchor_cells or _chart_anchor_cells(ws)
    if len(anchors) < 6:
        anchors = ["A15", "A30", "H15", "H30", "E30", "E15"]

    ws._charts = []
    header_map = _find_header_map(ws, header_row, start_col=1)

    qty_mom_col = header_map.get("qty mom")
    qty_yoy_col = header_map.get("qty yoy")
    qty_share_col = header_map.get("qty by %")
    rev_mom_col = header_map.get("revenue mom")
    rev_yoy_col = header_map.get("revenue yoy")
    rev_share_col = header_map.get("revenue by %")
    if not all([qty_mom_col, qty_yoy_col, qty_share_col, rev_mom_col, rev_yoy_col, rev_share_col]):
        return

    categories = Reference(ws, min_col=1, min_row=header_row + 1, max_row=section_end)

    pie_rows: list[int] = []
    for r in range(header_row + 1, section_end):
        label = _norm(ws.cell(r, 1).value)
        if label.startswith("total ") and label != "total":
            pie_rows.append(r)
    if not pie_rows:
        pie_rows = [r for r in range(header_row + 1, section_end) if _norm(ws.cell(r, 1).value) != "total"]
    pie_min = pie_rows[0] if pie_rows else header_row + 1
    pie_max = pie_rows[-1] if pie_rows else section_end - 1
    pie_categories = Reference(ws, min_col=1, min_row=pie_min, max_row=pie_max)

    def _bar_chart(value_col: int, title: str):
        chart = BarChart()
        chart.title = title
        chart.height = 7
        chart.width = 8.5
        chart.add_data(Reference(ws, min_col=value_col, min_row=header_row + 1, max_row=section_end), titles_from_data=False)
        chart.set_categories(categories)
        return chart

    def _pie_chart(value_col: int, title: str):
        chart = PieChart()
        chart.title = title
        chart.height = 7
        chart.width = 7
        chart.add_data(Reference(ws, min_col=value_col, min_row=pie_min, max_row=pie_max), titles_from_data=False)
        chart.set_categories(pie_categories)
        return chart

    charts = [
        (_bar_chart(qty_mom_col, "Qty Month Over Month"), anchors[0]),
        (_bar_chart(rev_mom_col, "Revenue Month Over Month"), anchors[1]),
        (_pie_chart(qty_share_col, "Qty Proportion"), anchors[2]),
        (_pie_chart(rev_share_col, "Revenue Proportion"), anchors[3]),
        (_bar_chart(rev_yoy_col, "Revenue Year Over Year"), anchors[4]),
        (_bar_chart(qty_yoy_col, "Qty Year Over Year"), anchors[5]),
    ]
    for chart, anchor in charts:
        ws.add_chart(chart, anchor)


def _normalize_table_link_formulas_inplace(xlsx_path: Path) -> None:
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        for table in ws.tables.values():
            try:
                from openpyxl.utils.cell import range_boundaries
            except Exception:
                continue
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            header_map = {}
            for col in range(min_col, max_col + 1):
                value = ws.cell(min_row, col).value
                if value is None:
                    continue
                header_map[_norm(value)] = (col, str(value))
            link_info = header_map.get("link")
            if not link_info:
                continue
            url_info = header_map.get("column1") or header_map.get("column2") or header_map.get("url")
            if not url_info:
                continue
            link_col, _link_header = link_info
            url_col, url_header = url_info
            preferred_style = None
            for row in range(min_row + 1, max_row + 1):
                current = ws.cell(row, link_col).value
                if isinstance(current, str) and current.startswith("="):
                    preferred_style = current
                    break
            for row in range(min_row + 1, max_row + 1):
                if ws.cell(row, min_col).value in {None, ""}:
                    continue
                ws.cell(row, link_col).value = _render_link_formula(
                    template_formula=preferred_style,
                    table_name=table.displayName,
                    url_header=url_header,
                    url_col=url_col,
                    out_row=row,
                )
    wb.save(xlsx_path)


TEMPLATE_TO_RAW_KEY = {
    "brand": "Brand",
    "# of listings": "# of Listings",
    "number of listings": "# of Listings",
    "asin": "ASIN",
    "product name": "Title",
    "title": "Title",
    "item #": "Item #",
    "1p/3p": "Fulfillment",
    "fulfillment": "Fulfillment",
    "type": "Type",
    "avg price": "Price",
    "price": "Price",
    "monthly rev": "Monthly Revenue",
    "monthly revenue": "Monthly Revenue",
    "est. monthly retail rev": "Monthly Revenue",
    "est. monthly retail revenue": "Monthly Revenue",
    "monthly units": "Monthly Sales",
    "monthly unit sales": "Monthly Sales",
    "est. monthly units sold": "Monthly Sales",
    "est. monthly unit sold": "Monthly Sales",
    "est. monthly units": "Monthly Sales",
    "estimated 12mo revenue": "12mo Revenue",
    "12mo revenue": "12mo Revenue",
    "estimated 12mo units": "12mo Units",
    "12mo units": "12mo Units",
    "# of reviews": "Review Count",
    "total reviews": "Review Count",
    "reviews": "Review Count",
    # Some sheets use "Ave Rating" (Summary), others use "Reviews Rating" (product tables).
    "ave rating": "Ave Rating",
    "avg rating": "Ave Rating",
    "avg. rating": "Reviews Rating",
    "tool rating": "Reviews Rating",
    "market rev share %": "Market Rev Share %",
    "monthly rev market share %": "Market Rev Share %",
    "market units share %": "Market Units Share %",
    "monthly unit market share %": "Market Units Share %",
    "price per unit": "Price Per Unit",
    "link": "URL",
    "column1": "URL",
    "column2": "URL",
}


def _cell_is_formula(cell) -> bool:
    # openpyxl stores formulas as data_type 'f', and cell.value typically starts with '='
    return getattr(cell, "data_type", None) == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))


def _normalize_item_number(value: Any, *, title: Any = None) -> Any:
    if value is not None and str(value).strip() != "":
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    title_text = str(title or "").strip()
    if not title_text:
        return value
    for pattern in (r"\bSDS?\d{2,4}\b", r"\b\d{4}[A-Z]{0,3}\b"):
        match = re.search(pattern, title_text, re.IGNORECASE)
        if match:
            return match.group(0)
    return value


def _render_link_formula(
    *,
    template_formula: Any,
    table_name: str | None,
    url_header: str | None,
    url_col: int | None,
    out_row: int,
) -> str | None:
    if isinstance(template_formula, str) and template_formula.startswith("="):
        structured_match = re.search(
            r"=HYPERLINK\(([^[]+)\[\[#This Row\],\[([^\]]+)\]\]\)",
            template_formula,
            re.IGNORECASE,
        )
        if structured_match:
            resolved_table = table_name or structured_match.group(1).strip()
            resolved_header = url_header or structured_match.group(2).strip()
            return f'=HYPERLINK({resolved_table}[[#This Row],[{resolved_header}]])'

        ref_match = re.search(r"=HYPERLINK\(([^,)]+)", template_formula, re.IGNORECASE)
        if ref_match:
            ref = ref_match.group(1).strip()
            cell_ref = re.match(r"\$?([A-Z]+)\$?\d+$", ref, re.IGNORECASE)
            if cell_ref:
                col_letters = cell_ref.group(1).upper()
                return f"=HYPERLINK({col_letters}{out_row})"

    if table_name and url_header:
        return f'=HYPERLINK({table_name}[[#This Row],[{url_header}]])'
    if url_col:
        return f"=HYPERLINK({get_column_letter(url_col)}{out_row})"
    return None


def _get_template_header_cells(ws, header_row: int, start_col: int = 1, max_col: int = 60) -> list[tuple[int, str]]:
    out: list[tuple[int, str]] = []
    for c in range(start_col, max_col + 1):
        v = ws.cell(header_row, c).value
        if v is None:
            continue
        s = str(v).strip()
        if s == "":
            continue
        out.append((c, s))
    return out


def _fill_table_by_headers(
    ws,
    header_row: int,
    start_col: int,
    rows: list[dict[str, Any]],
    stop_row: int | None = None,
    add_ranking: bool = False,
):
    template_headers = _get_template_header_cells(ws, header_row, start_col=start_col)
    if not template_headers:
        raise SystemExit(f"Could not read headers at {ws.title}!R{header_row}C{start_col}")
    url_col = None
    link_col = None
    url_header = None
    for c, th in template_headers:
        th_norm = _norm(th)
        if th_norm in {"column1", "column2", "url"}:
            url_col = c
            url_header = th
        elif th_norm == "link":
            link_col = c

    start_data_row = header_row + 1

    def _scan_section_end() -> tuple[int, str]:
        """
        Return (marker_row, marker_kind) where marker_row is the first row that
        terminates the section (blank row, Total row, Rank marker, or stop_row).
        marker_kind is one of: 'stop', 'blank', 'total', 'rank by revenue', 'rank by units', 'max'.
        """
        r = start_data_row
        max_rows_to_scan = 5000
        scanned = 0

        # Treat short blank gaps as available capacity (templates often have pre-formatted empty rows).
        blank_run = 0
        first_blank_row: int | None = None
        blank_run_stop = 50

        while True:
            if stop_row and r >= stop_row:
                return r, "stop"

            first = ws.cell(r, start_col).value
            if first is None or str(first).strip() == "":
                blank_run += 1
                if first_blank_row is None:
                    first_blank_row = r
                # If we hit a long run of blanks, assume the section ends at the first blank.
                if blank_run >= blank_run_stop:
                    return first_blank_row, "blank"
                r += 1
                scanned += 1
                if scanned >= max_rows_to_scan:
                    return first_blank_row or r, "max"
                continue

            blank_run = 0
            first_blank_row = None
            kind = _norm(first)
            if kind == "total":
                return r, "total"
            # Templates often include section titles like:
            # - "Rank By Monthly Units - Dec '25 ONLY"
            # - "Rank By Monthly Revenue - Jan '26 ONLY"
            # Treat any "rank by ..." row as a hard boundary so we don't clear the next section.
            if kind.startswith("rank by"):
                return r, "rank"

            r += 1
            scanned += 1
            if scanned >= max_rows_to_scan:
                return r, "max"

    def _clear_rows(start_row: int, end_row_exclusive: int):
        for r in range(start_row, end_row_exclusive):
            for c, th in template_headers:
                if _norm(th) == "link":
                    continue
                cell = ws.cell(r, c)
                if _cell_is_formula(cell):
                    continue
                try:
                    cell.value = None
                except AttributeError:
                    # openpyxl MergedCell is read-only (not the top-left cell)
                    continue

    def _clone_row_format(src_row: int, dst_row: int):
        # Copy style + (most) formulas from src_row into dst_row for the header columns.
        for c, th in template_headers:
            src = ws.cell(src_row, c)
            dst = ws.cell(dst_row, c)
            if src.has_style:
                dst._style = copy(src._style)
                dst.number_format = src.number_format
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)
            if _cell_is_formula(src):
                dst.value = src.value

        # Ensure Link column formula exists (templates often use '=HYPERLINK(<URL cell>)').
        for c, th in template_headers:
            if _norm(th) != "link":
                continue
            link_cell = ws.cell(dst_row, c)
            # Best-effort: infer the URL column from the template's link formula in src_row.
            src_link = ws.cell(src_row, c).value
            if isinstance(src_link, str) and src_link.startswith("="):
                m = re.search(r"=HYPERLINK\(([^,)]+)", src_link, re.IGNORECASE)
                if m:
                    ref = m.group(1).strip()
                    # ref like M9 -> keep column letters, replace row number
                    m2 = re.match(r"\$?([A-Z]+)\$?\d+", ref, re.IGNORECASE)
                    if m2:
                        col_letters = m2.group(1).upper()
                        link_cell.value = f"=HYPERLINK({col_letters}{dst_row})"
                        continue
            # Fallback: leave as-is (or empty) if template doesn't have a recognizable formula.

    def _fix_total_row_formulas(total_row: int, data_start: int, data_end: int):
        # Update simple SUM() ranges like =SUM(G9:G31) to match the expanded data rows.
        for c, _th in template_headers:
            cell = ws.cell(total_row, c)
            v = cell.value
            if not (isinstance(v, str) and v.startswith("=")):
                continue
            m = re.match(r"=SUM\(\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+\)$", v, re.IGNORECASE)
            if not m:
                continue
            col1 = m.group(1).upper()
            col2 = m.group(2).upper()
            if col1 != col2:
                continue
            cell.value = f"=SUM({col1}{data_start}:{col1}{data_end})"

    marker_row, marker_kind = _scan_section_end()
    _clear_rows(start_data_row, marker_row)

    max_capacity = marker_row - start_data_row
    if len(rows) > max_capacity:
        extra = len(rows) - max_capacity
        # Insert rows right before the marker (blank row, total row, or next rank marker).
        # This pushes the marker and anything below downward, preserving the template structure.
        ws.insert_rows(marker_row, amount=extra)

        # Clone formatting for the newly inserted rows from the last existing data row (best-effort).
        ref_row = max(start_data_row, marker_row - 1)
        for i in range(extra):
            _clone_row_format(ref_row, marker_row + i)

        # If the marker was a 'Total' row, it moved down by `extra` and formulas must be widened.
        if marker_kind == "total":
            new_total_row = marker_row + extra
            last_data_row = new_total_row - 1
            _fix_total_row_formulas(new_total_row, start_data_row, last_data_row)

        # Recompute marker/capacity after expansion.
        marker_row, marker_kind = _scan_section_end()
        max_capacity = marker_row - start_data_row

    def _update_excel_table_ref() -> str | None:
        """
        Keep Excel structured table ranges aligned with the row counts we write.

        The templates use real Excel Tables heavily. If we insert rows or write beyond
        the original table ref, Excel will often show a 'problem with some content'
        recovery dialog. Updating the Table.ref avoids most of those repairs.
        """
        try:
            from openpyxl.utils.cell import get_column_letter, range_boundaries
        except Exception:
            return

        table_list = getattr(ws, "_tables", None)
        if not table_list:
            return None

        # Tables are keyed by name -> Table object in this openpyxl version.
        for _name, t in table_list.items():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(t.ref)
            except Exception:
                continue
            if min_row != header_row or min_col != start_col:
                continue

            # Excel tables should have at least 1 data row.
            data_rows = max(1, len(rows))
            new_max_row = header_row + data_rows
            if new_max_row == max_row:
                return t.displayName
            new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
            try:
                t.ref = new_ref
                if getattr(t, "autoFilter", None) is not None:
                    t.autoFilter.ref = new_ref
            except Exception:
                continue
            return t.displayName
        return None

    table_name = _update_excel_table_ref()

    for i, row in enumerate(rows):
        out_row = start_data_row + i
        for c, th in template_headers:
            key_norm = _norm(th)
            if key_norm == "link":
                continue
            if add_ranking and key_norm in {"ranking", "rank", "ranking #"}:
                cell = ws.cell(out_row, c)
                if _cell_is_formula(cell):
                    continue
                try:
                    cell.value = i + 1
                except AttributeError:
                    continue
                continue
            raw_key = TEMPLATE_TO_RAW_KEY.get(key_norm)
            if raw_key is None:
                continue
            cell = ws.cell(out_row, c)
            if _cell_is_formula(cell):
                continue
            value = row.get(raw_key)
            if raw_key == "Item #":
                value = _normalize_item_number(value, title=row.get("Title"))
            try:
                cell.value = value
            except AttributeError:
                continue
        if url_col and link_col:
            link_cell = ws.cell(out_row, link_col)
            rendered_formula = _render_link_formula(
                template_formula=link_cell.value,
                table_name=table_name,
                url_header=url_header,
                url_col=url_col,
                out_row=out_row,
            )
            if rendered_formula:
                link_cell.value = rendered_formula


def _fill_report_workbook(
    *,
    month: str,
    raw_report_path: Path,
    template_path: Path,
    output_path: Path,
    overrides_path: Path | None = None,
    strip_visuals: bool = True,
):
    full_label, abbr_label = _month_labels(month)

    raw_wb = load_workbook(raw_report_path, data_only=True)
    tpl_wb = load_workbook(template_path)
    overrides = _load_override_bundle(overrides_path)

    # Detect old labels from known title cells to do a global replace (best-effort).
    old_full = None
    old_abbr = None
    ws_sum = tpl_wb["Summary"] if "Summary" in tpl_wb.sheetnames else None
    if ws_sum and isinstance(ws_sum["A1"].value, str):
        m = MONTH_FULL_RE.search(ws_sum["A1"].value)
        if m:
            old_full = m.group(0)
        m2 = MONTH_ABBR_RE.search(ws_sum["A1"].value)
        if m2:
            old_abbr = m2.group(0)
    _replace_month_labels_in_titles(tpl_wb, old_full, old_abbr, full_label, abbr_label)

    # Explicitly set a couple of known title cells (keeps consistent wording).
    if ws_sum:
        ws_sum["A1"].value = f"{full_label} Monthly Summary - Revenue "
    if "Top 50" in tpl_wb.sheetnames:
        ws_top = tpl_wb["Top 50"]
        ws_top["A1"].value = f"Rank By Monthly Revenue - {abbr_label} ONLY"
        # The template also contains a second title for the units table further down.
        # Update it explicitly because the global replace only knows `old_abbr` if it appears in Summary A1.
        for r in range(1, min(400, ws_top.max_row) + 1):
            v = ws_top.cell(r, 1).value
            if isinstance(v, str) and _norm(v).startswith("rank by monthly units"):
                ws_top.cell(r, 1).value = f"Rank By Monthly Units - {abbr_label} ONLY"
                break

    # Build sheet name map for raw workbook.
    raw_by_norm = {_norm(n): n for n in raw_wb.sheetnames}

    # Summary (two tables: revenue + units)
    if "Summary" in tpl_wb.sheetnames:
        tpl_ws = tpl_wb["Summary"]
        raw_ws = raw_wb[raw_by_norm.get("summary", "Summary")]
        tables = _split_two_tables(raw_ws, header_first_cell="Brand")
        rev_rows = tables.revenue
        unit_rows = tables.units or []

        brand_cell = _find_cell_text(tpl_ws, "Brand")
        rev_header_row = brand_cell.row if brand_cell else 2
        _fill_table_by_headers(tpl_ws, header_row=rev_header_row, start_col=1, rows=rev_rows)

        unit_header_row = _find_second_header_row(
            tpl_ws,
            first_cell_value="Brand",
            start_row=rev_header_row + 1,
            max_rows=2000,
        )
        if unit_header_row:
            _fill_table_by_headers(tpl_ws, header_row=unit_header_row, start_col=1, rows=unit_rows)

    # Rolling 12 mo (two tables)
    if "Rolling 12 mo" in tpl_wb.sheetnames and "Rolling 12 mo" in raw_wb.sheetnames:
        tpl_ws = tpl_wb["Rolling 12 mo"]
        raw_ws = raw_wb["Rolling 12 mo"]

        tables = _split_two_tables(raw_ws, header_first_cell="Brand")
        rev_rows = _apply_report_row_overrides(tables.revenue, target_sheet="Rolling 12 mo", overrides=overrides)
        unit_rows = _apply_report_row_overrides(tables.units or [], target_sheet="Rolling 12 mo", overrides=overrides)

        # Fill headers with real month names.
        month_headers = _rolling_12_month_headers(month)

        def _fill_rolling_section(
            *,
            header_row: int,
            rows: list[dict[str, Any]],
            key_prefix: str,
        ):
            # Template header row: Brand + 12 named months + Grand Total
            tpl_ws.cell(header_row, 1).value = "Brand"
            for i, mh in enumerate(month_headers, start=2):
                tpl_ws.cell(header_row, i).value = mh
            tpl_ws.cell(header_row, 14).value = "Grand Total"

            keys = [f"{key_prefix}-{i}" for i in range(11, 0, -1)] + [key_prefix]
            total_src = next((r for r in rows if _norm(r.get("Brand", "")) == "total"), None)
            total_market_src = next((r for r in rows if _norm(r.get("Brand", "")) == "total market"), None)
            grand_total_key = "Grand Total Revenue" if key_prefix == "Monthly Revenue" else "Grand Total Units"
            top_rows = [
                r for r in rows
                if _norm(r.get("Brand", "")) not in {"", "total", "total market"}
            ]
            top_rows = sorted(
                top_rows,
                key=lambda r: _coerce_number(r.get(grand_total_key)) or 0,
                reverse=True,
            )[:25]

            def _write_rolling_row(out_row: int, src: dict[str, Any] | None, *, label: str | None = None):
                for col in range(1, 15):
                    tpl_ws.cell(out_row, col).value = None
                if src is None:
                    return
                tpl_ws.cell(out_row, 1).value = label if label is not None else src.get("Brand")
                for j, k in enumerate(keys, start=2):
                    tpl_ws.cell(out_row, j).value = _coerce_number(src.get(k))
                tpl_ws.cell(out_row, 14).value = f"=SUM(B{out_row}:M{out_row})"

            data_start = header_row + 1
            for offset in range(25):
                src = top_rows[offset] if offset < len(top_rows) else None
                _write_rolling_row(data_start + offset, src)

            total_row = data_start + 25
            total_market_row = data_start + 26
            _write_rolling_row(total_row, total_src, label="Total")
            if total_market_src:
                _write_rolling_row(total_market_row, total_market_src, label="Total Market")
            else:
                total_values = _compute_market_totals(month=month, key_prefix=key_prefix)
                _write_rolling_row(total_market_row, {}, label="Total Market")
                for i, mh in enumerate(month_headers, start=2):
                    tpl_ws.cell(total_market_row, i).value = total_values.get(mh)
                tpl_ws.cell(total_market_row, 14).value = f"=SUM(B{total_market_row}:M{total_market_row})"

        # Revenue table in template: header row is the first row containing 'Brand' (row 2 in template).
        rev_header_row = _find_row_by_values(tpl_ws, required=["Brand"], max_rows=200) or 2
        _fill_rolling_section(header_row=rev_header_row, rows=rev_rows, key_prefix="Monthly Revenue")

        # Units table in template: find 'Monthly Units' title row then header below.
        units_title_cell = _find_cell_text(tpl_ws, "Monthly Units", max_rows=500) or _find_cell_text(
            tpl_ws, "Monthly Unit", max_rows=500
        )
        if units_title_cell:
            unit_header_row = units_title_cell.row + 1
        else:
            # fallback: second occurrence of 'Brand' below revenue table
            unit_header_row = _find_row_by_values(tpl_ws, required=["Brand"], max_rows=2000) or (rev_header_row + 30)

        _fill_rolling_section(header_row=unit_header_row, rows=unit_rows, key_prefix="Monthly Sales")

    # Top 50 (two tables: revenue + units)
    if "Top 50" in tpl_wb.sheetnames and "Top 50" in raw_wb.sheetnames:
        tpl_ws = tpl_wb["Top 50"]
        raw_ws = raw_wb["Top 50"]
        tables = _split_two_tables(raw_ws, header_first_cell="ASIN")
        rev_rows = _apply_report_row_overrides(tables.revenue, target_sheet="Top 50", overrides=overrides)
        unit_rows = _apply_report_row_overrides(tables.units or [], target_sheet="Top 50", overrides=overrides)

        rev_header_row = _find_row_by_values(tpl_ws, required=["Ranking", "ASIN"], max_rows=120) or 2
        _fill_table_by_headers(tpl_ws, header_row=rev_header_row, start_col=1, rows=rev_rows, add_ranking=True)

        units_title_cell = _find_cell_text(tpl_ws, "Rank By Monthly Units", max_rows=300) or _find_cell_text(
            tpl_ws, "Rank By Monthly Unit", max_rows=300
        )
        if units_title_cell:
            unit_header_row = units_title_cell.row + 1
        else:
            unit_header_row = _find_second_header_row(
                tpl_ws, first_cell_value="Ranking", start_row=rev_header_row + 1, max_rows=2000
            )

        if unit_header_row:
            _fill_table_by_headers(tpl_ws, header_row=unit_header_row, start_col=1, rows=unit_rows, add_ranking=True)

    # Brand / Innova tabs
    for sheet_name in tpl_wb.sheetnames:
        if sheet_name in {"Summary", "Rolling 12 mo", "Top 50"}:
            continue
        raw_sheet = raw_by_norm.get(_norm(sheet_name))
        if not raw_sheet:
            continue
        tpl_ws = tpl_wb[sheet_name]
        raw_ws = raw_wb[raw_sheet]

        tables = _split_two_tables(raw_ws, header_first_cell="Title")
        rev_rows = _apply_report_row_overrides(tables.revenue, target_sheet=sheet_name, overrides=overrides)
        unit_rows = _apply_report_row_overrides(tables.units or [], target_sheet=sheet_name, overrides=overrides)

        rev_marker = _find_cell_text(tpl_ws, "Rank by Revenue", max_rows=500, max_cols=20)
        unit_marker = _find_cell_text(tpl_ws, "Rank by Units", max_rows=1000, max_cols=20)
        if not rev_marker or not unit_marker:
            continue

        tpl_ws["A1"].value = f"{sheet_name} Monthly Summary"

        # Fill revenue table. It may insert rows and push the units marker down, so re-find it afterward.
        _fill_table_by_headers(
            tpl_ws,
            header_row=rev_marker.row + 1,
            start_col=rev_marker.column,
            rows=rev_rows,
        )

        unit_marker = _find_cell_text(tpl_ws, "Rank by Units", max_rows=2000, max_cols=20)
        if not unit_marker:
            continue

        _fill_table_by_headers(
            tpl_ws,
            header_row=unit_marker.row + 1,
            start_col=unit_marker.column,
            rows=unit_rows,
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _drop_invalid_tables(tpl_wb)
    tpl_wb.save(output_path)
    _normalize_table_link_formulas_inplace(output_path)
    if strip_visuals:
        _strip_visual_parts_inplace(output_path)
    _rewrite_relationship_targets_inplace(output_path)


def _build_index(rows: list[dict[str, Any]], key_fields: list[str]) -> dict[tuple[str, ...], dict[str, Any]]:
    out = {}
    for r in rows:
        key = tuple(_norm(r.get(f, "")) for f in key_fields)
        out[key] = r
    return out


def _aggregate_top50_rows(sheet1_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # sheet1 rows have columns like Avg Price, Qty MoM, Revenue YoY, etc.
    top50 = [r for r in sheet1_rows if _norm(r.get("Brand", "")) == "top 50"]
    idx = _build_index(top50, ["Category"])

    def get(cat: str) -> dict[str, Any] | None:
        return idx.get((_norm(cat),))

    def implied_prev(cur: float | None, mom: float | None) -> float | None:
        if cur is None or mom is None:
            return None
        if (1.0 + mom) == 0:
            return None
        return cur / (1.0 + mom)

    def agg(cats: list[str]) -> dict[str, Any]:
        # Aggregate current qty/rev and compute MoM/YoY by reconstructing prior values.
        cur_qty = 0.0
        cur_rev = 0.0
        prev_qty_m = 0.0
        prev_rev_m = 0.0
        prev_qty_y = 0.0
        prev_rev_y = 0.0
        has_prev_m = True
        has_prev_y = True
        for c in cats:
            r = get(c)
            if not r:
                continue
            q = _coerce_number(r.get("Quantity/Mo"))
            rev = _coerce_number(r.get("Revenue/Mo"))
            if q is not None:
                cur_qty += q
            if rev is not None:
                cur_rev += rev
            q_mom = _coerce_number(r.get("Qty MoM"))
            rev_mom = _coerce_number(r.get("Revenue MoM"))
            q_yoy = _coerce_number(r.get("Qty YoY"))
            rev_yoy = _coerce_number(r.get("Revenue YoY"))

            pqm = implied_prev(q, q_mom)
            prm = implied_prev(rev, rev_mom)
            pqy = implied_prev(q, q_yoy)
            pry = implied_prev(rev, rev_yoy)
            if pqm is None or prm is None:
                has_prev_m = False
            else:
                prev_qty_m += pqm
                prev_rev_m += prm
            if pqy is None or pry is None:
                has_prev_y = False
            else:
                prev_qty_y += pqy
                prev_rev_y += pry

        avg_price = (cur_rev / cur_qty) if cur_qty else None
        avg_price_mom = None
        avg_price_yoy = None
        qty_mom = None
        rev_mom = None
        qty_yoy = None
        rev_yoy = None
        if has_prev_m and prev_qty_m and prev_rev_m:
            qty_mom = cur_qty / prev_qty_m - 1 if prev_qty_m else None
            rev_mom = cur_rev / prev_rev_m - 1 if prev_rev_m else None
            prev_price = prev_rev_m / prev_qty_m if prev_qty_m else None
            if prev_price and avg_price is not None:
                avg_price_mom = avg_price / prev_price - 1
        if has_prev_y and prev_qty_y and prev_rev_y:
            qty_yoy = cur_qty / prev_qty_y - 1 if prev_qty_y else None
            rev_yoy = cur_rev / prev_rev_y - 1 if prev_rev_y else None
            prev_price = prev_rev_y / prev_qty_y if prev_qty_y else None
            if prev_price and avg_price is not None:
                avg_price_yoy = avg_price / prev_price - 1

        return {
            "Avg Price": avg_price,
            "Avg Price MoM": avg_price_mom,
            "Avg Price YoY": avg_price_yoy,
            "Quantity/Mo": cur_qty,
            "Qty MoM": qty_mom,
            "Qty YoY": qty_yoy,
            "Revenue/Mo": cur_rev,
            "Revenue MoM": rev_mom,
            "Revenue YoY": rev_yoy,
        }

    total = agg(["Total"])
    total_qty = total["Quantity/Mo"] or 0.0
    total_rev = total["Revenue/Mo"] or 0.0

    def with_shares(d: dict[str, Any]) -> dict[str, Any]:
        d = dict(d)
        q = d.get("Quantity/Mo") or 0.0
        r = d.get("Revenue/Mo") or 0.0
        d["Qty by %"] = (q / total_qty) if total_qty else None
        d["Revenue by %"] = (r / total_rev) if total_rev else None
        return d

    tablets = with_shares(agg(["Tablet $800+", "Tablet $400-$800", "Tablet $400-"]))
    handheld = with_shares(agg(["Handheld $75+", "Handheld $75-"]))
    dongle = with_shares(agg(["Total Dongle"]))
    other = with_shares(agg(["Total Other Tools"]))
    total = with_shares(total)

    def row(name: str, d: dict[str, Any]) -> dict[str, Any]:
        return {"Category": name, **d}

    return [
        row("Tablets", tablets),
        row("Handheld", handheld),
        row("Dongle", dongle),
        row("Other Tools", other),
        row("Total", total),
    ]


def _safe_div(num: float | None, den: float | None) -> float | None:
    if num is None or den in {None, 0}:
        return None
    return num / den


def _pct_change(current: float | None, previous: float | None) -> float | None:
    if current is None or previous in {None, 0}:
        return None
    return current / previous - 1


def _metric_header_columns(ws, header_row: int) -> list[tuple[int, str]]:
    return _get_template_header_cells(ws, header_row, start_col=2, max_col=30)


def _write_metric_row(ws, *, row_idx: int, header_row: int, source: dict[str, Any]) -> None:
    for col_idx, header in _metric_header_columns(ws, header_row):
        cell = ws.cell(row_idx, col_idx)
        if _cell_is_formula(cell):
            continue
        if header == "Revenue YoY":
            cell.value = source.get("Revenue YoY")
            continue
        cell.value = source.get(header)


def _aggregate_metric_rows(
    rows: list[dict[str, Any]],
    *,
    qty_share_total: float | None = None,
    revenue_share_total: float | None = None,
) -> dict[str, Any]:
    cur_qty = 0.0
    cur_rev = 0.0
    prev_qty_m = 0.0
    prev_rev_m = 0.0
    prev_qty_y = 0.0
    prev_rev_y = 0.0
    any_prev_m = False
    any_prev_y = False

    for row in rows:
        qty = _coerce_number(row.get("Quantity/Mo"))
        rev = _coerce_number(row.get("Revenue/Mo"))
        if qty is not None:
            cur_qty += qty
        if rev is not None:
            cur_rev += rev

        qty_mom = _coerce_number(row.get("Qty MoM"))
        rev_mom = _coerce_number(row.get("Revenue MoM"))
        qty_yoy = _coerce_number(row.get("Qty YoY"))
        rev_yoy = _coerce_number(row.get("Revenue YoY"))

        if qty is not None and qty_mom is not None and (1.0 + qty_mom) != 0:
            prev_qty_m += qty / (1.0 + qty_mom)
            any_prev_m = True
        if rev is not None and rev_mom is not None and (1.0 + rev_mom) != 0:
            prev_rev_m += rev / (1.0 + rev_mom)
            any_prev_m = True
        if qty is not None and qty_yoy is not None and (1.0 + qty_yoy) != 0:
            prev_qty_y += qty / (1.0 + qty_yoy)
            any_prev_y = True
        if rev is not None and rev_yoy is not None and (1.0 + rev_yoy) != 0:
            prev_rev_y += rev / (1.0 + rev_yoy)
            any_prev_y = True

    avg_price = _safe_div(cur_rev, cur_qty)
    prev_price_m = _safe_div(prev_rev_m, prev_qty_m) if any_prev_m else None
    prev_price_y = _safe_div(prev_rev_y, prev_qty_y) if any_prev_y else None

    return {
        "Avg Price": avg_price,
        "Avg Price MoM": _pct_change(avg_price, prev_price_m),
        "Avg Price YoY": _pct_change(avg_price, prev_price_y),
        "Quantity/Mo": cur_qty,
        "Qty by %": _safe_div(cur_qty, qty_share_total),
        "Qty MoM": _pct_change(cur_qty, prev_qty_m if any_prev_m else None),
        "Qty YoY": _pct_change(cur_qty, prev_qty_y if any_prev_y else None),
        "Revenue/Mo": cur_rev,
        "Revenue by %": _safe_div(cur_rev, revenue_share_total),
        "Revenue MoM": _pct_change(cur_rev, prev_rev_m if any_prev_m else None),
        "Revenue YoY": _pct_change(cur_rev, prev_rev_y if any_prev_y else None),
    }


def _analysis_category_bucket(type_value: Any, price_value: Any) -> str:
    type_norm = _norm(type_value)
    price = _coerce_number(price_value) or 0.0
    if type_norm == "tablet":
        if price >= 800:
            return "Tablet $800+"
        if price >= 400:
            return "Tablet $400-$800"
        return "Tablet $400-"
    if type_norm == "handheld":
        if price >= 75:
            return "Handheld $75+"
        return "Handheld $75-"
    if type_norm == "dongle":
        return "Total Dongle"
    return "Total Other Tools"


def _analysis_bucket_rows(current_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    buckets: dict[str, list[dict[str, Any]]] = {
        "Tablet $800+": [],
        "Tablet $400-$800": [],
        "Tablet $400-": [],
        "Total Tablet": [],
        "Handheld $75+": [],
        "Handheld $75-": [],
        "Total Handheld": [],
        "Total Dongle": [],
        "Total Other Tools": [],
        "Total": [],
    }
    for row in current_rows:
        bucket = _analysis_category_bucket(row.get("Type"), row.get("Price"))
        buckets[bucket].append(row)
        if bucket.startswith("Tablet "):
            buckets["Total Tablet"].append(row)
        elif bucket.startswith("Handheld "):
            buckets["Total Handheld"].append(row)
        buckets["Total"].append(row)
    return buckets


def _summary_metric_index(sheet1_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {
        _norm(row.get("Category")): row
        for row in sheet1_rows
        if _norm(row.get("Brand")) == "total" and row.get("Category") is not None
    }


def _summary_category_metric_index(cat_rows: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    out: dict[tuple[str, str], dict[str, Any]] = {}
    for row in cat_rows:
        brand = row.get("Brand")
        if brand in {None, 0}:
            continue
        out[(_norm(row.get("Category")), _norm(brand))] = row
    return out


def _scan_metric_sections(ws, *, max_rows: int = 400) -> list[tuple[int, str | None]]:
    sections: list[tuple[int, str | None]] = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        header_map = _find_header_map(ws, r, start_col=1, max_col=20)
        if "avg price" not in header_map or "revenue/mo" not in header_map:
            continue
        title = None
        for lookback in range(r - 1, max(r - 4, 0), -1):
            value = ws.cell(lookback, 1).value
            if value is None or str(value).strip() == "":
                continue
            title = str(value).strip()
            break
        sections.append((r, title))
    return sections


def _read_simple_table(ws, header_row: int) -> list[dict[str, Any]]:
    _headers, rows = _read_table(ws, header_row=header_row)
    return rows


def _sorted_rows(rows: list[dict[str, Any]], key_name: str) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: (_coerce_number(row.get(key_name)) or 0.0), reverse=True)


def _collect_report_detail_pools(raw_report_wb) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    revenue_by_asin: dict[str, dict[str, Any]] = {}
    units_by_asin: dict[str, dict[str, Any]] = {}
    for sheet_name in raw_report_wb.sheetnames:
        if _norm(sheet_name) in {"summary", "rolling 12 mo", "top 50"}:
            continue
        ws = raw_report_wb[sheet_name]
        header_row = _find_second_header_row(ws, "Title", start_row=1)
        if not header_row:
            continue
        _headers, rows = _read_table(ws, header_row)
        normalized_brand = _normalize_report_brand_display(sheet_name)
        for row in rows:
            asin = str(row.get("ASIN", "")).strip()
            if not asin:
                continue
            enriched = dict(row)
            enriched.setdefault("Brand", normalized_brand.lower())
            revenue_by_asin.setdefault(asin, enriched)
            units_by_asin.setdefault(asin, enriched)
    return list(revenue_by_asin.values()), list(units_by_asin.values())


def _analysis_sheet_matches_row(sheet_name: str, row: dict[str, Any]) -> bool:
    bucket = _analysis_category_bucket(row.get("Type"), row.get("Price"))
    sheet_norm = _norm(sheet_name)
    bucket_norm = _norm(bucket)
    if sheet_norm == "tablet total":
        return bucket_norm in {_norm("Tablet $800+"), _norm("Tablet $400-$800"), _norm("Tablet $400-")}
    if sheet_norm == "handheld total":
        return bucket_norm in {_norm("Handheld $75+"), _norm("Handheld $75-")}
    if sheet_norm == "dongle":
        return bucket_norm == _norm("Total Dongle")
    if sheet_norm == "other tools":
        return bucket_norm == _norm("Total Other Tools")
    return bucket_norm == sheet_norm


def _build_brand_analysis_rows(
    *,
    raw_rows: list[dict[str, Any]],
    month: str,
    target_sheet: str,
    raw_brand_key: str,
    category_metric_index: dict[tuple[str, str], dict[str, Any]],
    overrides: OverrideBundle,
) -> dict[str, dict[str, Any]]:
    prev_idx = _load_obd2_asin_index(_shift_month(month, -1))
    yoy_idx = _load_obd2_asin_index(_shift_month(month, -12))
    current_buckets = _analysis_bucket_rows(raw_rows)

    history_buckets_prev: dict[str, list[dict[str, Any]]] = {key: [] for key in current_buckets}
    history_buckets_yoy: dict[str, list[dict[str, Any]]] = {key: [] for key in current_buckets}
    for row in raw_rows:
        asin = str(row.get("ASIN", "")).strip()
        if not asin:
            continue
        bucket = _analysis_category_bucket(row.get("Type"), row.get("Price"))
        prev_row = prev_idx.get(asin, {})
        yoy_row = yoy_idx.get(asin, {})
        prev_metric_row = {
            "Quantity/Mo": _coerce_number(prev_row.get("Monthly Sales")) or 0.0,
            "Revenue/Mo": _coerce_number(prev_row.get("Monthly Revenue")) or 0.0,
        }
        yoy_metric_row = {
            "Quantity/Mo": _coerce_number(yoy_row.get("Monthly Sales")) or 0.0,
            "Revenue/Mo": _coerce_number(yoy_row.get("Monthly Revenue")) or 0.0,
        }
        history_buckets_prev[bucket].append(prev_metric_row)
        history_buckets_yoy[bucket].append(yoy_metric_row)
        if bucket.startswith("Tablet "):
            history_buckets_prev["Total Tablet"].append(prev_metric_row)
            history_buckets_yoy["Total Tablet"].append(yoy_metric_row)
        elif bucket.startswith("Handheld "):
            history_buckets_prev["Total Handheld"].append(prev_metric_row)
            history_buckets_yoy["Total Handheld"].append(yoy_metric_row)
        history_buckets_prev["Total"].append(prev_metric_row)
        history_buckets_yoy["Total"].append(yoy_metric_row)

    total_qty = sum((_coerce_number(row.get("Monthly Sales")) or 0.0) for row in raw_rows)
    total_rev = sum((_coerce_number(row.get("Monthly Revenue")) or 0.0) for row in raw_rows)
    categories = [
        "Tablet $800+",
        "Tablet $400-$800",
        "Tablet $400-",
        "Total Tablet",
        "Handheld $75+",
        "Handheld $75-",
        "Total Handheld",
        "Total Dongle",
        "Total Other Tools",
        "Total",
    ]

    out: dict[str, dict[str, Any]] = {}
    for category in categories:
        summary_source = category_metric_index.get((_norm(category), _norm(raw_brand_key)))
        if summary_source:
            metric_row = {
                "Avg Price": summary_source.get("Avg Price"),
                "Avg Price MoM": summary_source.get("Avg Price MoM"),
                "Avg Price YoY": summary_source.get("Avg Price YoY"),
                "Quantity/Mo": summary_source.get("Quantity/Mo"),
                "Qty by %": _safe_div(_coerce_number(summary_source.get("Quantity/Mo")), total_qty),
                "Qty MoM": summary_source.get("Qty MoM"),
                "Qty YoY": summary_source.get("Qty YoY"),
                "Revenue/Mo": summary_source.get("Revenue/Mo"),
                "Revenue by %": _safe_div(_coerce_number(summary_source.get("Revenue/Mo")), total_rev),
                "Revenue MoM": summary_source.get("Revenue MoM"),
                "Revenue YoY": summary_source.get("Revenue YoY"),
            }
        else:
            current_qty = sum((_coerce_number(row.get("Monthly Sales")) or 0.0) for row in current_buckets[category])
            current_rev = sum((_coerce_number(row.get("Monthly Revenue")) or 0.0) for row in current_buckets[category])
            prev_qty = sum((_coerce_number(row.get("Quantity/Mo")) or 0.0) for row in history_buckets_prev[category])
            prev_rev = sum((_coerce_number(row.get("Revenue/Mo")) or 0.0) for row in history_buckets_prev[category])
            yoy_qty = sum((_coerce_number(row.get("Quantity/Mo")) or 0.0) for row in history_buckets_yoy[category])
            yoy_rev = sum((_coerce_number(row.get("Revenue/Mo")) or 0.0) for row in history_buckets_yoy[category])

            avg_price = _safe_div(current_rev, current_qty)
            prev_avg_price = _safe_div(prev_rev, prev_qty)
            yoy_avg_price = _safe_div(yoy_rev, yoy_qty)
            metric_row = {
                "Avg Price": avg_price,
                "Avg Price MoM": _pct_change(avg_price, prev_avg_price),
                "Avg Price YoY": _pct_change(avg_price, yoy_avg_price),
                "Quantity/Mo": current_qty,
                "Qty by %": _safe_div(current_qty, total_qty),
                "Qty MoM": _pct_change(current_qty, prev_qty),
                "Qty YoY": _pct_change(current_qty, yoy_qty),
                "Revenue/Mo": current_rev,
                "Revenue by %": _safe_div(current_rev, total_rev),
                "Revenue MoM": _pct_change(current_rev, prev_rev),
                "Revenue YoY": _pct_change(current_rev, yoy_rev),
            }
        out[category] = _apply_analysis_metric_override(
            metric_row,
            target_sheet=target_sheet,
            category=category,
            overrides=overrides,
        )
    return out


def _fill_analysis_workbook(
    *,
    month: str,
    raw_summary_path: Path,
    raw_report_path: Path,
    template_path: Path,
    output_path: Path,
    overrides_path: Path | None = None,
    strip_visuals: bool = False,
):
    full_label, abbr_label = _month_labels(month)
    raw_summary_wb = load_workbook(raw_summary_path, data_only=True)
    raw_report_wb = load_workbook(raw_report_path, data_only=True)
    tpl_wb = load_workbook(template_path)
    overrides = _load_override_bundle(overrides_path)
    top50_market_tables = _split_two_tables(raw_report_wb["Top 50"], header_first_cell="ASIN")

    # Best-effort global title label replacement.
    old_full = None
    if "Summary" in tpl_wb.sheetnames and isinstance(tpl_wb["Summary"]["A1"].value, str):
        m = MONTH_FULL_RE.search(tpl_wb["Summary"]["A1"].value)
        if m:
            old_full = m.group(0)
    _replace_month_labels_in_titles(tpl_wb, old_full, None, full_label, "")

    if "Summary" in tpl_wb.sheetnames:
        tpl_wb["Summary"]["A1"].value = f"{full_label} Monthly Summary"
    if "Top 50" in tpl_wb.sheetnames:
        tpl_wb["Top 50"]["A1"].value = f"Top 50 Products - {full_label} ONLY"

    if "Sheet1" not in raw_summary_wb.sheetnames or "Category" not in raw_summary_wb.sheetnames:
        raise SystemExit(f"{raw_summary_path} must contain sheets: Sheet1, Category")

    sheet1_rows = _read_simple_table(raw_summary_wb["Sheet1"], header_row=1)
    cat_rows = _read_simple_table(raw_summary_wb["Category"], header_row=1)
    summary_total_index = _summary_metric_index(sheet1_rows)
    category_metric_index = _summary_category_metric_index(cat_rows)
    raw_report_by_norm = {_norm(name): name for name in raw_report_wb.sheetnames}
    summary_brand_rows = _read_simple_table(raw_report_wb[raw_report_by_norm.get("summary", "Summary")], header_row=2)

    def fill_metric_section(ws, *, header_row: int, rows_by_label: dict[str, dict[str, Any]]) -> None:
        populated_cols = {col for col, _header in _metric_header_columns(ws, header_row)}
        row_idx = header_row + 1
        while row_idx <= ws.max_row:
            label = ws.cell(row_idx, 1).value
            if label is None or str(label).strip() == "":
                break
            if _norm(label).startswith("rank by"):
                break
            source = rows_by_label.get(_norm(label))
            if source:
                _write_metric_row(ws, row_idx=row_idx, header_row=header_row, source=source)
            for col in range(2, 16):
                if col in populated_cols:
                    continue
                cell = ws.cell(row_idx, col)
                if _cell_is_formula(cell):
                    continue
                cell.value = None
            row_idx += 1

    if "Summary" in tpl_wb.sheetnames:
        ws = tpl_wb["Summary"]
        anchors = _chart_anchor_cells(ws)
        for header_row, _section_title in _scan_metric_sections(ws, max_rows=40):
            section_rows = {
                _norm(category): _apply_analysis_metric_override(
                    summary_total_index[_norm(category)],
                    target_sheet="Summary",
                    category=category,
                    overrides=overrides,
                )
                for category in [
                    "Tablet $800+",
                    "Tablet $400-$800",
                    "Tablet $400-",
                    "Total Tablet",
                    "Handheld $75+",
                    "Handheld $75-",
                    "Total Handheld",
                    "Total Dongle",
                    "Total Other Tools",
                    "Total",
                ]
                if _norm(category) in summary_total_index
            }
            fill_metric_section(ws, header_row=header_row, rows_by_label=section_rows)

        report_tables = _split_two_tables(raw_report_wb[raw_report_by_norm.get("summary", "Summary")], header_first_cell="Brand")
        revenue_rows = _apply_report_row_overrides(report_tables.revenue, target_sheet="Summary", overrides=overrides)
        unit_rows = _apply_report_row_overrides(report_tables.units or [], target_sheet="Summary", overrides=overrides)
        rev_header_row = _find_row_by_values(ws, required=["Brand", "Monthly Revenue"], max_rows=200) or 46
        _fill_table_by_headers(ws, header_row=rev_header_row, start_col=1, rows=revenue_rows)
        unit_header_row = _find_second_header_row(ws, first_cell_value="Brand", start_row=rev_header_row + 1, max_rows=500)
        if unit_header_row:
            _fill_table_by_headers(ws, header_row=unit_header_row, start_col=1, rows=unit_rows)
        _rebuild_analysis_charts(ws, anchor_cells=anchors)

    if "Top 50" in tpl_wb.sheetnames:
        ws = tpl_wb["Top 50"]
        anchors = _chart_anchor_cells(ws)
        agg_rows = { _norm(row["Category"]): row for row in _aggregate_top50_rows(sheet1_rows) }
        for header_row, _section_title in _scan_metric_sections(ws, max_rows=20):
            fill_metric_section(ws, header_row=header_row, rows_by_label=agg_rows)

        tables = _split_two_tables(raw_report_wb["Top 50"], header_first_cell="ASIN")
        rev_rows = _apply_report_row_overrides(tables.revenue, target_sheet="Top 50", overrides=overrides)
        unit_rows = _apply_report_row_overrides(tables.units or [], target_sheet="Top 50", overrides=overrides)
        rev_header_row = _find_row_by_values(ws, required=["Ranking", "ASIN"], max_rows=120) or 41
        _fill_table_by_headers(ws, header_row=rev_header_row, start_col=1, rows=rev_rows, add_ranking=True)
        unit_header_row = _find_second_header_row(ws, first_cell_value="Ranking", start_row=rev_header_row + 1, max_rows=400)
        if unit_header_row:
            _fill_table_by_headers(ws, header_row=unit_header_row, start_col=1, rows=unit_rows, add_ranking=True)
        _rebuild_analysis_charts(ws, anchor_cells=anchors)

    analysis_to_category = {
        "Tablet Total": "Total Tablet",
        "Tablet $800+": "Tablet $800+",
        "Tablet $400-$800": "Tablet $400-$800",
        "Tablet $400-": "Tablet $400-",
        "Handheld Total": "Total Handheld",
        "Handheld $75+": "Handheld $75+",
        "Handheld $75-": "Handheld $75-",
        "Dongle": "Total Dongle",
        "Other Tools": "Total Other Tools",
    }
    section_aliases = {
        "tablet overall": "Total Tablet",
        "handheld overall": "Total Handheld",
        "dongle overall": "Total Dongle",
        "other tools overall": "Total Other Tools",
    }

    for sheet_name, summary_sheet_name in analysis_to_category.items():
        if sheet_name not in tpl_wb.sheetnames:
            continue
        ws = tpl_wb[sheet_name]
        anchors = _chart_anchor_cells(ws)
        total_metrics = summary_total_index.get(_norm(summary_sheet_name))
        total_qty = _coerce_number(total_metrics.get("Quantity/Mo")) if total_metrics else None
        total_rev = _coerce_number(total_metrics.get("Revenue/Mo")) if total_metrics else None

        for header_row, section_title in _scan_metric_sections(ws, max_rows=140):
            lookup_category = summary_sheet_name if header_row < 40 else section_aliases.get(_norm(section_title), section_title or summary_sheet_name)
            lookup_category = lookup_category or summary_sheet_name
            displayed_labels: list[str] = []
            row_idx = header_row + 1
            while row_idx <= ws.max_row:
                label = ws.cell(row_idx, 1).value
                if label is None or str(label).strip() == "":
                    break
                if _norm(label).startswith("rank by"):
                    break
                displayed_labels.append(str(label))
                row_idx += 1

            rows_by_label: dict[str, dict[str, Any]] = {}
            used_rows: list[dict[str, Any]] = []
            for label in displayed_labels:
                if _norm(label) in {"other", "total"}:
                    continue
                source = category_metric_index.get((_norm(lookup_category), _norm(label)))
                if not source:
                    continue
                rows_by_label[_norm(label)] = source
                used_rows.append(source)

            if any(_norm(label) == "other" for label in displayed_labels):
                remaining = [
                    row
                    for (category_key, brand_key), row in category_metric_index.items()
                    if category_key == _norm(lookup_category)
                    and brand_key not in {_norm(lbl) for lbl in displayed_labels if _norm(lbl) not in {"other", "total"}}
                ]
                rows_by_label["other"] = _aggregate_metric_rows(
                    remaining,
                    qty_share_total=total_qty,
                    revenue_share_total=total_rev,
                )
            if total_metrics is not None:
                rows_by_label["total"] = total_metrics

            fill_metric_section(ws, header_row=header_row, rows_by_label=rows_by_label)

        revenue_candidates = [row for row in top50_market_tables.revenue if _analysis_sheet_matches_row(sheet_name, row)]
        unit_candidates = [row for row in (top50_market_tables.units or []) if _analysis_sheet_matches_row(sheet_name, row)]
        if not revenue_candidates and summary_sheet_name in raw_summary_wb.sheetnames:
            revenue_candidates = _read_simple_table(raw_summary_wb[summary_sheet_name], header_row=1)
        if not unit_candidates and summary_sheet_name in raw_summary_wb.sheetnames:
            unit_candidates = _read_simple_table(raw_summary_wb[summary_sheet_name], header_row=1)
        if revenue_candidates or unit_candidates:
            rev_rows = _apply_report_row_overrides(
                revenue_candidates[:10],
                target_sheet=sheet_name,
                overrides=overrides,
            )
            unit_rows = _apply_report_row_overrides(
                unit_candidates[:10],
                target_sheet=sheet_name,
                overrides=overrides,
            )
            rev_header_row = _find_row_by_values(ws, required=["Ranking", "ASIN"], max_rows=200) or 46
            _fill_table_by_headers(ws, header_row=rev_header_row, start_col=1, rows=rev_rows, add_ranking=True)
            unit_header_row = _find_second_header_row(ws, first_cell_value="Ranking", start_row=rev_header_row + 1, max_rows=400)
            if unit_header_row:
                _fill_table_by_headers(ws, header_row=unit_header_row, start_col=1, rows=unit_rows, add_ranking=True)
        _rebuild_analysis_charts(ws, anchor_cells=anchors)

    desired_brand_sheets = []
    for raw_name in raw_report_wb.sheetnames:
        if _norm(raw_name) in {"summary", "rolling 12 mo", "top 50"}:
            continue
        desired_brand_sheets.append(_normalize_analysis_brand_display(raw_name, tpl_wb.sheetnames))
    resolved_brand_sheets = _sync_analysis_brand_sheets(tpl_wb, desired_brand_sheets)

    for sheet_name in resolved_brand_sheets:
        ws = tpl_wb[sheet_name]
        ws["A1"].value = f"{sheet_name} Monthly Summary"
        anchors = _chart_anchor_cells(ws)
        raw_sheet_name = raw_report_by_norm.get(_norm(sheet_name))
        if not raw_sheet_name:
            continue
        raw_ws = raw_report_wb[raw_sheet_name]
        brand_tables = _split_two_tables(raw_ws, header_first_cell="Title")
        raw_rows = brand_tables.revenue
        metric_rows = _build_brand_analysis_rows(
            raw_rows=raw_rows,
            month=month,
            target_sheet=sheet_name,
            raw_brand_key=raw_sheet_name,
            category_metric_index=category_metric_index,
            overrides=overrides,
        )
        for header_row, _section_title in _scan_metric_sections(ws, max_rows=40):
            fill_metric_section(ws, header_row=header_row, rows_by_label={_norm(k): v for k, v in metric_rows.items()})

        rev_rows = _apply_report_row_overrides(brand_tables.revenue, target_sheet=sheet_name, overrides=overrides)
        unit_rows = _apply_report_row_overrides(brand_tables.units or [], target_sheet=sheet_name, overrides=overrides)
        rev_marker = _find_cell_text(ws, "Rank by Revenue", max_rows=500, max_cols=20)
        unit_marker = _find_cell_text(ws, "Rank by Units", max_rows=2000, max_cols=20)
        if rev_marker:
            _fill_table_by_headers(ws, header_row=rev_marker.row + 1, start_col=rev_marker.column, rows=rev_rows)
            unit_marker = _find_cell_text(ws, "Rank by Units", max_rows=4000, max_cols=20)
        if unit_marker:
            _fill_table_by_headers(ws, header_row=unit_marker.row + 1, start_col=unit_marker.column, rows=unit_rows)
        _rebuild_analysis_charts(ws, anchor_cells=anchors)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    _drop_invalid_tables(tpl_wb)
    tpl_wb.save(output_path)
    _normalize_table_link_formulas_inplace(output_path)
    if strip_visuals:
        _strip_visual_parts_inplace(output_path)
    _rewrite_relationship_targets_inplace(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Automate the final formatted Excel outputs.")
    parser.add_argument("--month", required=True, help="Report month in YYYYMM (e.g. 202601).")
    parser.add_argument(
        "--run-date",
        default=None,
        help="Run date for output filename prefix (YYYY-MM-DD). Default: today.",
    )
    parser.add_argument(
        "--raw-report",
        default="Amazon Competitor Report.xlsx",
        help="Raw report workbook (default: ./Amazon Competitor Report.xlsx).",
    )
    parser.add_argument(
        "--raw-summary",
        default="summary.xlsx",
        help="Raw summary workbook (default: ./summary.xlsx).",
    )
    parser.add_argument("--template-report", required=True, help="Formatted report template xlsx path.")
    parser.add_argument("--template-analysis", required=True, help="Formatted analysis template xlsx path.")
    parser.add_argument(
        "--strip-visuals",
        action="store_true",
        help="Strip charts/drawings/images from outputs. Default: preserve visuals and rebuild missing analysis charts.",
    )
    parser.add_argument(
        "--amazon-obd2-dir",
        default=None,
        help="Folder containing amazon_obd2_YYYYMM.xlsx files (default: ../amazon_obd2).",
    )
    parser.add_argument(
        "--out-dir",
        default=None,
        help="Output folder (default: ../YY-MM-reports based on --month).",
    )
    parser.add_argument("--overrides", default=None, help="Override workbook path (optional).")
    args = parser.parse_args()

    run_date = dt.date.today() if not args.run_date else dt.date.fromisoformat(args.run_date)
    month = args.month
    full_label, _abbr_label = _month_labels(month)
    month_name = _yyyymm_to_period(month).strftime("%B")

    yy = run_date.strftime("%y")
    mm = run_date.strftime("%m")
    dd = run_date.strftime("%d")
    prefix = f"{yy}-{mm}-{dd}"

    # Match the existing folder convention like: 25-12-reports, 26-01-reports
    out_dir = Path(args.out_dir) if args.out_dir else Path("..") / f"{month[2:4]}-{month[4:6]}-reports"
    out_dir = out_dir.resolve()

    report_out = out_dir / f"{prefix} Amazon Competitor Report {month_name} Innova Adjusted.xlsx"
    analysis_out = out_dir / f"{prefix} Amazon Competitor Analysis {month_name}.xlsx"

    raw_report_path = _as_path(args.raw_report).resolve()
    raw_summary_path = _as_path(args.raw_summary).resolve()
    template_report_path = _as_path(args.template_report).resolve()
    template_analysis_path = _as_path(args.template_analysis).resolve()
    overrides_path = _as_path(args.overrides).resolve() if args.overrides else None
    global _AMAZON_OBD2_DIR_OVERRIDE
    if args.amazon_obd2_dir:
        _AMAZON_OBD2_DIR_OVERRIDE = _as_path(args.amazon_obd2_dir).resolve()

    if not raw_report_path.exists():
        raise SystemExit(f"Missing raw report: {raw_report_path}")
    if not raw_summary_path.exists():
        raise SystemExit(f"Missing raw summary: {raw_summary_path}")
    if not template_report_path.exists():
        raise SystemExit(f"Missing report template: {template_report_path}")
    if not template_analysis_path.exists():
        raise SystemExit(f"Missing analysis template: {template_analysis_path}")

    _fill_report_workbook(
        month=month,
        raw_report_path=raw_report_path,
        template_path=template_report_path,
        output_path=report_out,
        overrides_path=overrides_path,
        strip_visuals=args.strip_visuals,
    )
    _fill_analysis_workbook(
        month=month,
        raw_summary_path=raw_summary_path,
        raw_report_path=raw_report_path,
        template_path=template_analysis_path,
        output_path=analysis_out,
        overrides_path=overrides_path,
        strip_visuals=args.strip_visuals,
    )

    print("Wrote:")
    print(f"- {report_out}")
    print(f"- {analysis_out}")
    print(f"Month label: {full_label}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
