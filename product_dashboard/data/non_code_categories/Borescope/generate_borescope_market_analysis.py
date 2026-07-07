#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_TYPE_MAP_FILE = (
    BASE_DIR / "_archive" / "non_dashboard_noise_20260305" / "extended_type_mapping.xlsx"
)
DEFAULT_PRICE_CEILING = 1000.0
TOP_N = 50
PRICE_TIERS = [
    ("$0-40", 0, 40),
    ("$40-60", 40, 60),
    ("$60-90", 60, 90),
    ("$90-100", 90, 100),
    ("$100-150", 100, 150),
    ("$150-200", 150, 200),
    ("$200+", 200, float("inf")),
]

INPUT_DIMENSION_COLUMNS = [
    "Type",
    "2/4-way",
    "Display",
    "Lens diameter",
    "Lens count",
    "Cable length",
]

OUTPUT_DIMENSION_COLUMNS = {
    "Type": "Type",
    "2/4-way": "2/4-way",
    "Display": "Display",
    "Lens diameter": "Lens Diameter",
    "Lens count": "Lens Count",
    "Cable length": "Cable Length",
}

INVALID_DIMENSION_TOKENS = {
    "",
    "nan",
    "none",
    "type",
    "2/4-way",
    "display",
    "lens diameter",
    "lens count",
    "cable length",
}

COMMON_LENS_MM = [1.8, 3.0, 3.9, 5.5, 6.0, 6.2, 6.4, 6.5, 7.0, 8.0, 8.5, 23.0]
COMMON_DISPLAY_IN = [2.4, 4.3, 4.5, 5.0, 6.0, 7.0, 9.0]
COMMON_CABLE_FT = [3.0, 3.3, 3.9, 4.0, 5.0, 5.2, 5.3, 5.5, 6.6, 9.8, 10.0, 16.5, 32.8, 33.0, 50.0, 100.0, 300.0]
ARTICULATION_PATTERN = re.compile(
    r"(articulat|2-?way|two-?way|4-?way|four-?way|joystick|210°|220°|360°|360 degree)",
    re.IGNORECASE,
)
CONNECTED_PATTERN = re.compile(
    r"\b(wireless|wi-?fi|app|iphone|ios|android|smartphone)\b",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Borescope monthly market analysis workbooks.")
    parser.add_argument(
        "--raw-dir",
        type=Path,
        required=True,
        help="Directory containing the monthly raw CSV files (for example raw_data/202603).",
    )
    parser.add_argument(
        "--type-map-file",
        type=Path,
        default=DEFAULT_TYPE_MAP_FILE,
        help="Excel file with authoritative ASIN type/dimension mappings.",
    )
    parser.add_argument(
        "--price-ceiling",
        type=float,
        default=DEFAULT_PRICE_CEILING,
        help="Maximum allowed Price value for included ASINs.",
    )
    parser.add_argument(
        "--output-file",
        type=Path,
        default=None,
        help="Output workbook path. Defaults are derived from the raw month and scope.",
    )
    parser.add_argument(
        "--scope",
        choices=["all", "type_articulation"],
        default="all",
        help="Dataset scope. type_articulation rebuilds the articulating segment including connected articulating products.",
    )
    return parser.parse_args()


def round2(value: float) -> float:
    return round(float(value), 2)


def round4(value: float) -> float:
    return round(float(value), 4)


def infer_output_path(raw_dir: Path, scope: str) -> Path:
    month = infer_snapshot_month(raw_dir)
    if scope == "type_articulation":
        filename = f"Borescope_Articulation_Market_Analysis_{month}.xlsx"
    else:
        filename = f"Borescope_Market_Analysis_{month}.xlsx"
    return raw_dir.parent.parent / "outputs" / filename


def infer_snapshot_month(raw_dir: Path) -> str:
    match = re.search(r"(\d{6})", str(raw_dir))
    if not match:
        raise ValueError(f"Could not infer snapshot month from raw dir: {raw_dir}")
    return match.group(1)


def month_end(snapshot_month: str) -> str:
    year = int(snapshot_month[:4])
    month = int(snapshot_month[4:6])
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day).isoformat()


def latest_run_date(raw_dir: Path, fallback_snapshot_date: str) -> str:
    dates = sorted(
        {
            match.group(1)
            for file_path in raw_dir.glob("*.csv")
            if (match := re.search(r"(\d{4}-\d{2}-\d{2})", file_path.name))
        }
    )
    return dates[-1] if dates else fallback_snapshot_date


def load_all_raw_data(raw_dir: Path) -> pd.DataFrame:
    csv_files = sorted(raw_dir.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {raw_dir}")

    frames: list[pd.DataFrame] = []
    for csv_file in csv_files:
        frame = pd.read_csv(csv_file, encoding="utf-8-sig")
        frame["source_file"] = csv_file.name
        frames.append(frame)

    return pd.concat(frames, ignore_index=True)


def is_missing_dimension_value(value: object) -> bool:
    if pd.isna(value):
        return True
    return str(value).strip().lower() in INVALID_DIMENSION_TOKENS


def nearest_common(value: float, options: list[float], max_delta: float) -> float:
    nearest = min(options, key=lambda option: abs(option - value))
    return nearest if abs(nearest - value) <= max_delta else value


def format_mm(value: float) -> str:
    if abs(value - round(value)) < 1e-6:
        return f"{int(round(value))}mm"
    return f"{value:.1f}mm"


def format_ft(value: float) -> str:
    if abs(value - round(value)) < 1e-6:
        return f"{int(round(value))}ft"
    return f"{value:.1f}ft"


def extract_display(title_lower: str) -> str:
    candidates: list[float] = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:[\"″”]|-?\s*inch\b|in\b)", title_lower):
        value = float(match.group(1))
        if 2.0 <= value <= 12.0:
            candidates.append(value)

    if candidates:
        display = nearest_common(candidates[0], COMMON_DISPLAY_IN, max_delta=0.35)
        return f'{display:.1f}"'

    if re.search(r"\b(android|iphone|ios|smartphone|app|wi-?fi|wireless)\b", title_lower):
        return "App"

    if re.search(r"\b(screen|monitor|ips|lcd)\b", title_lower):
        return '5.0"'

    return "Unknown"


def extract_lens_diameter(title_lower: str) -> str:
    mm_candidates: list[float] = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*mm\b", title_lower):
        value = float(match.group(1))
        if 0.8 <= value <= 30:
            mm_candidates.append(value)

    if mm_candidates:
        return format_mm(nearest_common(mm_candidates[0], COMMON_LENS_MM, max_delta=0.4))

    inch_candidates: list[float] = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:[\"″”]|inches\b|inch\b|in\b)", title_lower):
        value = float(match.group(1))
        if 0.05 <= value <= 1.2:
            inch_candidates.append(value)

    if inch_candidates:
        return format_mm(nearest_common(inch_candidates[0] * 25.4, COMMON_LENS_MM, max_delta=0.45))

    return "Unknown"


def extract_cable_length(title_lower: str) -> str:
    candidates: list[float] = []
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:ft|feet|foot)\b", title_lower):
        value = float(match.group(1))
        if 1 <= value <= 400:
            candidates.append(value)
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*'(?!')", title_lower):
        value = float(match.group(1))
        if 1 <= value <= 400:
            candidates.append(value)
    for match in re.finditer(r"(\d+(?:\.\d+)?)\s*(?:meter|meters|m)\b", title_lower):
        value = float(match.group(1))
        if 0.3 <= value <= 120:
            candidates.append(value * 3.28084)

    if candidates:
        return format_ft(nearest_common(candidates[0], COMMON_CABLE_FT, max_delta=0.35))

    return "Unknown"


def infer_dimensions_from_title(title: object) -> dict[str, str]:
    title_lower = str(title or "").lower()

    if re.search(r"\b(4-?way|four-?way|4 ways|4 way|joystick|360°|360 degree)\b", title_lower):
        articulation = "4-way"
    elif re.search(r"\b(2-?way|two-?way|210°|220°)\b", title_lower):
        articulation = "2-way"
    elif "articulat" in title_lower:
        articulation = "2-way"
    else:
        articulation = "2-way"

    if re.search(r"\b(triple|3[-\s]?lens|three[-\s]?lens)\b", title_lower):
        lens_count = "Triple"
    elif re.search(r"\b(dual|2[-\s]?lens|two[-\s]?lens|dual[-\s]?view)\b", title_lower):
        lens_count = "Dual"
    else:
        lens_count = "Single"

    if "sewer" in title_lower:
        scope_type = "Sewer camera"
    elif re.search(r"\b(wireless|wi-?fi)\b", title_lower):
        scope_type = "Wireless"
    elif re.search(r"\b(usb|type-c|smartphone|iphone|android|ios)\b", title_lower):
        scope_type = "USB"
    else:
        scope_type = "Articulation"

    return {
        "Type": scope_type,
        "2/4-way": articulation,
        "Display": extract_display(title_lower),
        "Lens Diameter": extract_lens_diameter(title_lower),
        "Lens Count": lens_count,
        "Cable Length": extract_cable_length(title_lower),
    }


def normalize_type_map_columns(type_map: pd.DataFrame) -> pd.DataFrame:
    renamed = type_map.rename(columns=OUTPUT_DIMENSION_COLUMNS).copy()
    for source_col, output_col in OUTPUT_DIMENSION_COLUMNS.items():
        if source_col in renamed.columns and output_col not in renamed.columns:
            renamed[output_col] = renamed[source_col]
    return renamed


def add_type_columns(df: pd.DataFrame, type_map_file: Path) -> pd.DataFrame:
    if not type_map_file.exists():
        raise FileNotFoundError(f"Type mapping file not found: {type_map_file}")

    type_map = normalize_type_map_columns(pd.read_excel(type_map_file))
    if "ASIN" not in type_map.columns:
        raise ValueError(f"Type mapping file must contain ASIN column: {type_map_file}")

    type_map["ASIN"] = type_map["ASIN"].astype(str)
    type_map = type_map.drop_duplicates(subset=["ASIN"])

    for output_col in OUTPUT_DIMENSION_COLUMNS.values():
        if output_col in type_map.columns:
            type_map[output_col] = type_map[output_col].mask(type_map[output_col].apply(is_missing_dimension_value), pd.NA)

    merge_cols = ["ASIN"] + [col for col in OUTPUT_DIMENSION_COLUMNS.values() if col in type_map.columns]
    merged = df.merge(type_map[merge_cols], on="ASIN", how="left")
    return merged


def fill_missing_dimensions_from_title(df: pd.DataFrame) -> pd.DataFrame:
    filled = df.copy()
    inferred = filled["Title"].apply(infer_dimensions_from_title).apply(pd.Series)
    for column in OUTPUT_DIMENSION_COLUMNS.values():
        if column not in filled.columns:
            filled[column] = pd.NA
        missing_mask = filled[column].apply(is_missing_dimension_value)
        if missing_mask.any():
            filled.loc[missing_mask, column] = inferred.loc[missing_mask, column]
    return filled


def is_articulating_scope_row(row: pd.Series) -> bool:
    title = str(row.get("Title") or "")
    scope_type = str(row.get("Type") or "")
    if scope_type == "Articulation":
        return True
    if scope_type in {"USB", "Wireless"} and ARTICULATION_PATTERN.search(title):
        return True
    return False


def classify_connection_type(row: pd.Series) -> str:
    title = str(row.get("Title") or "")
    scope_type = str(row.get("Type") or "")
    display = str(row.get("Display") or "")
    if scope_type == "Wireless" or display == "App" or CONNECTED_PATTERN.search(title):
        return "Wireless/Connected"
    return "Screen/Wired"


def add_articulation_segment_columns(df: pd.DataFrame) -> pd.DataFrame:
    enriched = df.copy()
    enriched["is_articulating_scope"] = enriched.apply(is_articulating_scope_row, axis=1)
    enriched["Connection Type"] = enriched.apply(classify_connection_type, axis=1)
    return enriched


def load_processed_records(raw_dir: Path, type_map_file: Path, price_ceiling: float) -> pd.DataFrame:
    df = load_all_raw_data(raw_dir)
    if "ASIN" not in df.columns:
        raise ValueError("Raw data is missing ASIN column.")

    before = len(df)
    df = df.drop_duplicates(subset=["ASIN"]).copy()
    print(f"Deduplicated ASINs: {before} -> {len(df)} rows")

    if "Subcategory" in df.columns:
        df = df[df["Subcategory"].astype(str).eq("Borescopes")].copy()

    numeric_cols = ["Price", "ASIN Revenue", "ASIN Sales", "Review Count", "Reviews Rating"]
    for column in numeric_cols:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")

    df = df[df["Price"].notna() & (df["Price"] <= price_ceiling)].copy()
    df = add_type_columns(df, type_map_file)
    df = fill_missing_dimensions_from_title(df)
    df = add_articulation_segment_columns(df)
    return df


def apply_scope(df: pd.DataFrame, scope: str) -> pd.DataFrame:
    scoped = df.copy()
    if scope == "type_articulation":
        scoped = scoped[scoped["is_articulating_scope"].fillna(False)].copy()
    return scoped.reset_index(drop=True)


def build_brand_summary(df: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        df.groupby("Brand", dropna=False)
        .agg(
            listings=("ASIN", "nunique"),
            revenue=("ASIN Revenue", "sum"),
            units=("ASIN Sales", "sum"),
            reviews=("Review Count", "sum"),
            weighted_rating=("Reviews Rating", lambda series: 0),
        )
        .reset_index()
    )

    # Weighted average rating by review count, fallback to simple mean for zero-review buckets.
    weighted_ratings: list[float] = []
    for brand in grouped["Brand"]:
        brand_rows = df[df["Brand"] == brand]
        total_reviews = float(brand_rows["Review Count"].fillna(0).sum())
        if total_reviews > 0:
            weighted = (
                (brand_rows["Reviews Rating"].fillna(0) * brand_rows["Review Count"].fillna(0)).sum()
                / total_reviews
            )
            weighted_ratings.append(round2(weighted))
        else:
            weighted_ratings.append(round2(float(brand_rows["Reviews Rating"].fillna(0).mean())))

    grouped["Avg Rating"] = weighted_ratings
    grouped["Monthly Rev"] = grouped["revenue"].map(round2)
    grouped["Monthly Units"] = grouped["units"].map(round2)
    total_revenue = float(grouped["revenue"].sum())
    grouped["Monthly Rev Market Share %"] = grouped["revenue"].apply(
        lambda value: round4(value / total_revenue) if total_revenue > 0 else 0
    )
    grouped["Price Per Unit"] = grouped.apply(
        lambda row: round2(row["revenue"] / row["units"]) if row["units"] else 0,
        axis=1,
    )

    summary = grouped[
        ["Brand", "listings", "Monthly Rev", "Monthly Units", "Monthly Rev Market Share %", "Price Per Unit", "Avg Rating"]
    ].rename(columns={"listings": "# of Listings"})
    summary = summary.sort_values("Monthly Rev", ascending=False).reset_index(drop=True)
    total_row = {
        "Brand": "Total",
        "# of Listings": int(summary["# of Listings"].sum()),
        "Monthly Rev": round2(summary["Monthly Rev"].sum()),
        "Monthly Units": round2(summary["Monthly Units"].sum()),
        "Monthly Rev Market Share %": 1 if total_revenue > 0 else 0,
        "Price Per Unit": round2(total_revenue / summary["Monthly Units"].sum()) if summary["Monthly Units"].sum() else 0,
        "Avg Rating": 0,
    }
    return pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)


def build_top_sheet_rows(df: pd.DataFrame, include_connection_type: bool = False) -> pd.DataFrame:
    rows = {
        "ASIN": df["ASIN"],
        "Title": df["Title"],
        "Brand": df["Brand"],
        "Type": df["Type"],
        "Price": df["Price"].map(round2),
        "Monthly Rev": df["ASIN Revenue"].map(round2),
        "Monthly Units": df["ASIN Sales"].map(round2),
        "Avg Rating": df["Reviews Rating"].fillna(0).map(round2),
        "# of Reviews": df["Review Count"].fillna(0).map(round2),
        "Link": df["URL"].fillna(""),
    }
    if include_connection_type and "Connection Type" in df.columns:
        rows["Connection Type"] = df["Connection Type"].fillna("Screen/Wired")
    rows.update(
        {
            "2/4-way": df["2/4-way"].fillna("Unknown"),
            "Display": df["Display"].fillna("Unknown"),
            "Lens Diameter": df["Lens Diameter"].fillna("Unknown"),
            "Lens Count": df["Lens Count"].fillna("Unknown"),
            "Cable Length": df["Cable Length"].fillna("Unknown"),
        }
    )
    return pd.DataFrame(rows)


def build_summary_section(df: pd.DataFrame, label_column: str) -> list[list[object]]:
    total_revenue = float(df["ASIN Revenue"].sum())
    total_units = float(df["ASIN Sales"].sum())

    grouped = (
        df.groupby(label_column, dropna=False)
        .agg(
            avg_price=("Price", "mean"),
            units=("ASIN Sales", "sum"),
            revenue=("ASIN Revenue", "sum"),
        )
        .reset_index()
        .rename(columns={label_column: label_column})
    )

    grouped[label_column] = grouped[label_column].fillna("Unknown")
    grouped["Qty By %"] = grouped["units"].apply(lambda value: round4(value / total_units) if total_units > 0 else 0)
    grouped["Revenue By %"] = grouped["revenue"].apply(
        lambda value: round4(value / total_revenue) if total_revenue > 0 else 0
    )
    grouped["Avg Price"] = grouped["avg_price"].fillna(0).map(round2)
    grouped["Quantity/Mo"] = grouped["units"].fillna(0).map(round2)
    grouped["Revenue/Mo"] = grouped["revenue"].fillna(0).map(round2)
    grouped = grouped.sort_values("Revenue/Mo", ascending=False)

    rows: list[list[object]] = [[label_column, "Avg Price", "Quantity/Mo", "Qty By %", "Revenue/Mo", "Revenue By %"]]
    for _, row in grouped.iterrows():
        rows.append(
            [
                row[label_column],
                row["Avg Price"],
                row["Quantity/Mo"],
                row["Qty By %"],
                row["Revenue/Mo"],
                row["Revenue By %"],
            ]
        )
    return rows


def build_top50_summary_sheet(df: pd.DataFrame, scope: str) -> list[list[object]]:
    sections = [
        ("Connection Type", "Connection Type"),
        ("2/4-way", "2/4-way"),
        ("Display", "Display"),
        ("Lens Diameter", "Lens Diameter"),
        ("Lens Count", "Lens Count"),
        ("Cable Length", "Cable Length"),
    ]
    if scope != "type_articulation":
        sections[0] = ("Type", "Type")
    rows: list[list[object]] = []
    for index, (title, column) in enumerate(sections):
        if index > 0:
            rows.append([])
        rows.extend(build_summary_section(df, title if column == title else column))
    return rows


def average_price(df: pd.DataFrame) -> float:
    if df.empty:
        return 0
    return round2(float(df["Price"].mean()))


def build_price_tiers(df: pd.DataFrame) -> pd.DataFrame:
    total_revenue = float(df["ASIN Revenue"].sum())
    total_units = float(df["ASIN Sales"].sum())
    rows: list[dict[str, object]] = []
    for label, minimum, maximum in PRICE_TIERS:
        matched = df[(df["Price"] >= minimum) & (df["Price"] < maximum)].copy()
        revenue = float(matched["ASIN Revenue"].sum())
        units = float(matched["ASIN Sales"].sum())
        rows.append(
            {
                "Price Tier": label,
                "Total Revenue": round2(revenue),
                "Total Sales": round2(units),
                "Rev Share %": round4(revenue / total_revenue) if total_revenue > 0 else 0,
                "Unit Share %": round4(units / total_units) if total_units > 0 else 0,
                "Avg Price": average_price(matched),
            }
        )
    return pd.DataFrame(rows)


def build_metadata_rows(scope: str, snapshot_month: str, snapshot_date: str, run_date: str, asin_count: int) -> list[list[object]]:
    if scope == "all":
        category_label = "Borescope"
        category_id = "borescope"
        rule_note = "All borescope ASINs after standard filters."
    else:
        category_label = "Borescope Articulation Split"
        category_id = "borescope_type_articulation_split"
        rule_note = (
            "Includes Type=Articulation plus articulating USB/Wireless ASINs; "
            "Connection Type splits Wireless/Connected vs Screen/Wired."
        )
    return [
        ["Category", category_label],
        ["Category ID", category_id],
        ["Snapshot Month", snapshot_month],
        ["Snapshot Date", snapshot_date],
        ["Latest Run Date", run_date],
        ["Generated At", datetime.now().isoformat()],
        ["ASIN Count", asin_count],
        ["Scope Rule", rule_note],
    ]


def autosize_columns(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            length = len(str(value)) if value is not None else 0
            widths[index] = min(max(widths.get(index, 10), length + 2), 80)
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = width


def append_dataframe_sheet(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    ws = workbook.create_sheet(name)
    ws.append(list(frame.columns))
    for row in frame.astype(object).where(pd.notnull(frame), None).itertuples(index=False, name=None):
        ws.append(list(row))
    autosize_columns(ws)


def append_rows_sheet(workbook: Workbook, name: str, rows: list[list[object]]) -> None:
    ws = workbook.create_sheet(name)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)


def build_workbook(df: pd.DataFrame, scope: str, snapshot_month: str, snapshot_date: str, run_date: str) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    include_connection_type = scope == "type_articulation"

    top_revenue_source = df.sort_values("ASIN Revenue", ascending=False).head(TOP_N).copy()
    top_units_source = df.sort_values("ASIN Sales", ascending=False).head(TOP_N).copy()
    all_asins_source = df.sort_values("ASIN Revenue", ascending=False).copy()

    append_dataframe_sheet(workbook, "Summary", build_brand_summary(df))
    append_dataframe_sheet(workbook, "Top 50 Revenue", build_top_sheet_rows(top_revenue_source, include_connection_type))
    append_dataframe_sheet(workbook, "Top 50 Units", build_top_sheet_rows(top_units_source, include_connection_type))
    append_rows_sheet(workbook, "Top 50 Summary", build_top50_summary_sheet(top_revenue_source, scope))
    append_dataframe_sheet(workbook, "Price Tiers", build_price_tiers(df))
    append_dataframe_sheet(workbook, "All ASINs", build_top_sheet_rows(all_asins_source, include_connection_type))
    append_rows_sheet(
        workbook,
        "Metadata",
        build_metadata_rows(scope, snapshot_month, snapshot_date, run_date, len(df)),
    )
    return workbook


def main() -> None:
    args = parse_args()
    snapshot_month = infer_snapshot_month(args.raw_dir)
    snapshot_date = month_end(snapshot_month)
    run_date = latest_run_date(args.raw_dir, snapshot_date)
    output_file = args.output_file or infer_output_path(args.raw_dir, args.scope)

    processed = load_processed_records(args.raw_dir, args.type_map_file, args.price_ceiling)
    scoped = apply_scope(processed, args.scope)

    workbook = build_workbook(scoped, args.scope, snapshot_month, snapshot_date, run_date)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)

    print(f"Wrote workbook: {output_file}")
    print(f"Scope: {args.scope}")
    print(f"Rows: {len(scoped)}")
    print(f"Revenue: {round2(scoped['ASIN Revenue'].sum())}")
    print(f"Units: {round2(scoped['ASIN Sales'].sum())}")
    print(f"Max price: {round2(scoped['Price'].max()) if not scoped.empty else 0}")


if __name__ == "__main__":
    main()
