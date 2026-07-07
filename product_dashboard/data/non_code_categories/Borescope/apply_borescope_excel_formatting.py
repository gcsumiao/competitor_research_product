#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE_FILL = "B4A7D6"
HEADER_FILL = "D9D2E9"
ALT_FILL = "EAE4F5"
TOTAL_FILL = "2E1A47"
TOTAL_FONT_COLOR = "FFFFFF"
LINK_COLOR = "0563C1"
GRID_COLOR = "CCCCCC"
BRAND_PALETTE = [
    "6A4C93",
    "8E7CC3",
    "B4A7D6",
    "4A90E2",
    "50C2C9",
    "7FB069",
    "F4B942",
    "E8833A",
    "D7567A",
    "9B5DE5",
]

METRIC_HEADERS = ("Avg Price", "Quantity/Mo", "Qty By %", "Revenue/Mo", "Revenue By %")
TOP_SHEETS = ("Top 50 Revenue", "Top 50 Units", "All ASINs")
CHARTABLE_SHEETS = ("Summary", "Top 50 Summary", "Price Tiers")
TITLE_SUFFIX = {
    "Summary": "Summary",
    "Top 50 Revenue": "Top 50 Revenue",
    "Top 50 Units": "Top 50 Units",
    "Top 50 Summary": "Segment Summary",
    "Price Tiers": "Price Tier Analysis",
    "All ASINs": "All ASINs",
    "Metadata": "Metadata",
}
WIDTHS = {
    "Summary": {"A": 22, "B": 14, "C": 16, "D": 16, "E": 18, "F": 14, "G": 12},
    "Top 50 Revenue": {
        "A": 14,
        "B": 62,
        "C": 18,
        "D": 16,
        "E": 12,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 14,
        "J": 22,
        "K": 20,
        "L": 12,
        "M": 14,
        "N": 16,
        "O": 12,
        "P": 14,
    },
    "Top 50 Units": {
        "A": 14,
        "B": 62,
        "C": 18,
        "D": 16,
        "E": 12,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 14,
        "J": 22,
        "K": 20,
        "L": 12,
        "M": 14,
        "N": 16,
        "O": 12,
        "P": 14,
    },
    "Top 50 Summary": {"A": 22, "B": 14, "C": 14, "D": 12, "E": 16, "F": 12},
    "Price Tiers": {"A": 16, "B": 16, "C": 14, "D": 12, "E": 12, "F": 12},
    "All ASINs": {
        "A": 14,
        "B": 62,
        "C": 18,
        "D": 16,
        "E": 12,
        "F": 16,
        "G": 16,
        "H": 12,
        "I": 14,
        "J": 22,
        "K": 20,
        "L": 12,
        "M": 14,
        "N": 16,
        "O": 12,
        "P": 14,
    },
    "Metadata": {"A": 24, "B": 28},
}
REVENUE_HEADERS = {
    "monthly rev",
    "revenue/mo",
    "total revenue",
}
PRICE_HEADERS = {
    "avg price",
    "price",
    "price per unit",
}
PCT_HEADERS = {
    "monthly rev market share %",
    "rev share %",
    "unit share %",
    "revenue by %",
    "qty by %",
}
INTEGER_HEADERS = {
    "# of listings",
    "monthly units",
    "quantity/mo",
    "total sales",
    "# of reviews",
    "lens count",
}
DECIMAL_HEADERS = {
    "avg rating",
}


@dataclass
class Section:
    label: str
    header_row: int
    data_start: int
    data_end: int


def title_font(size: int = 20) -> Font:
    return Font(name="Arial", bold=True, size=size, color="000000")


def header_font() -> Font:
    return Font(name="Arial", bold=True, size=11, color="000000")


def data_font(size: int = 10) -> Font:
    return Font(name="Calibri", size=size, color="000000")


def total_font() -> Font:
    return Font(name="Arial", bold=True, size=11, color=TOTAL_FONT_COLOR)


def link_font() -> Font:
    return Font(name="Calibri", size=10, color=LINK_COLOR, underline="single")


def title_fill() -> PatternFill:
    return PatternFill("solid", fgColor=TITLE_FILL)


def header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=HEADER_FILL)


def alt_fill() -> PatternFill:
    return PatternFill("solid", fgColor=ALT_FILL)


def total_fill() -> PatternFill:
    return PatternFill("solid", fgColor=TOTAL_FILL)


def center_alignment(wrap_text: bool = True) -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap_text)


def left_alignment(wrap_text: bool = False) -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap_text)


def thin_border() -> Border:
    side = Side(style="thin", color=GRID_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


def total_border() -> Border:
    thick = Side(style="medium", color=TOTAL_FILL)
    thin = Side(style="thin", color=TOTAL_FILL)
    return Border(left=thin, right=thin, top=thick, bottom=thick)


def reset_existing_title_row(ws, title_text: str) -> None:
    if ws["A1"].value != title_text:
        return
    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges if rng.min_row == 1 and rng.max_row == 1]
    for merged_range in merged_ranges:
        ws.unmerge_cells(merged_range)
    ws.delete_rows(1)


def apply_title_row(ws, title_text: str) -> None:
    reset_existing_title_row(ws, title_text)
    ncols = ws.max_column
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = title_font(20)
    cell.fill = title_fill()
    cell.alignment = center_alignment()
    ws.row_dimensions[1].height = 34


def format_header_row(ws, row_idx: int, ncols: int) -> None:
    ws.row_dimensions[row_idx].height = 30
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = header_font()
        cell.fill = header_fill()
        cell.alignment = center_alignment()
        cell.border = thin_border()


def set_column_widths(ws) -> None:
    widths = WIDTHS.get(ws.title, {})
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width


def sheet_label(workbook_path: Path) -> str:
    stem = workbook_path.stem.replace("_", " ")
    return stem


def title_text_for_sheet(workbook_path: Path, sheet_name: str) -> str:
    return f"{sheet_label(workbook_path)} - {TITLE_SUFFIX.get(sheet_name, sheet_name)}"


def is_blank_row(values: Iterable[object]) -> bool:
    return all(value in (None, "") for value in values)


def section_rows(ws) -> list[Section]:
    sections: list[Section] = []
    current_header = None
    data_start = None
    for row in range(2, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, min(6, ws.max_column) + 1)]
        if len(values) >= 6 and tuple(values[1:6]) == METRIC_HEADERS and values[0]:
            if current_header is not None and data_start is not None and row - 1 >= data_start:
                sections.append(
                    Section(
                        label=str(ws.cell(current_header, 1).value),
                        header_row=current_header,
                        data_start=data_start,
                        data_end=row - 2 if is_blank_row([ws.cell(row - 1, col).value for col in range(1, 7)]) else row - 1,
                    )
                )
            current_header = row
            data_start = row + 1
        elif current_header is not None and data_start is not None and is_blank_row(values):
            if row - 1 >= data_start:
                sections.append(
                    Section(
                        label=str(ws.cell(current_header, 1).value),
                        header_row=current_header,
                        data_start=data_start,
                        data_end=row - 1,
                    )
                )
            current_header = None
            data_start = None
    if current_header is not None and data_start is not None and ws.max_row >= data_start:
        sections.append(
            Section(
                label=str(ws.cell(current_header, 1).value),
                header_row=current_header,
                data_start=data_start,
                data_end=ws.max_row,
            )
        )
    cleaned: list[Section] = []
    for section in sections:
        end = section.data_end
        while end >= section.data_start and is_blank_row(
            [ws.cell(end, col).value for col in range(1, min(6, ws.max_column) + 1)]
        ):
            end -= 1
        if end >= section.data_start:
            cleaned.append(Section(section.label, section.header_row, section.data_start, end))
    return cleaned


def style_data_range(ws, start_row: int, end_row: int, left_columns: set[int], alt: bool) -> None:
    use_alt = alt and (end_row - start_row + 1) <= 500
    for row in range(start_row, end_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = link_font() if is_hyperlink_formula(cell.value) else data_font(10)
            cell.border = thin_border()
            cell.alignment = left_alignment(wrap_text=col in left_columns) if col in left_columns else center_alignment(False)
            if use_alt and (row - start_row) % 2 == 1:
                cell.fill = alt_fill()


def is_hyperlink_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=HYPERLINK(")


def format_numeric_columns(ws, header_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "").strip().lower()
        if not header:
            continue
        number_format = None
        if header in REVENUE_HEADERS:
            number_format = "$#,##0"
        elif header in PRICE_HEADERS:
            number_format = "$#,##0"
        elif header in PCT_HEADERS:
            number_format = "0.0%"
        elif header in INTEGER_HEADERS:
            number_format = "#,##0"
        elif header in DECIMAL_HEADERS:
            number_format = "0.0"
        if number_format is None:
            continue
        for row in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row, col)
            if cell.value is None or isinstance(cell.value, str):
                continue
            cell.number_format = number_format


def convert_links(ws, header_row: int) -> None:
    link_col = None
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "").strip().lower()
        if header in {"link", "url"}:
            link_col = col
            break
    if link_col is None:
        return
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row, link_col)
        raw = cell.value
        if isinstance(raw, str) and raw.startswith("http"):
            label = "Open Link"
            if "/dp/" in raw:
                asin = raw.split("/dp/")[-1].split("/")[0].split("?")[0]
                label = f"Amazon - {asin}"
            cell.value = f'=HYPERLINK("{raw}","{label}")'
            cell.font = link_font()
            cell.alignment = left_alignment()


def clear_charts(ws) -> None:
    if getattr(ws, "_charts", None):
        ws._charts.clear()


def dollar_labels() -> DataLabelList:
    labels = DataLabelList()
    labels.showVal = True
    labels.showPercent = False
    labels.showCatName = False
    labels.showSerName = False
    labels.showLegendKey = False
    labels.numFmt = "$#,##0"
    return labels


def integer_labels() -> DataLabelList:
    labels = DataLabelList()
    labels.showVal = True
    labels.showPercent = False
    labels.showCatName = False
    labels.showSerName = False
    labels.showLegendKey = False
    labels.numFmt = "#,##0"
    return labels


def pie_labels() -> DataLabelList:
    labels = DataLabelList()
    labels.showVal = False
    labels.showPercent = True
    labels.showCatName = True
    labels.showSerName = False
    labels.showLegendKey = False
    return labels


def pct_labels() -> DataLabelList:
    labels = DataLabelList()
    labels.showVal = True
    labels.showPercent = False
    labels.showCatName = False
    labels.showSerName = False
    labels.showLegendKey = False
    labels.numFmt = "0%"
    return labels


def color_bars_per_category(series, n_points: int) -> None:
    series.dPt = []
    for idx in range(n_points):
        color = BRAND_PALETTE[idx % len(BRAND_PALETTE)]
        series.dPt.append(DataPoint(idx=idx, spPr=GraphicalProperties(solidFill=color)))


def annotate_category_axis(chart, axis: str, title: str) -> None:
    axis_obj = chart.y_axis if axis == "y" else chart.x_axis
    axis_obj.delete = False
    axis_obj.majorTickMark = "out"
    axis_obj.tickLblPos = "nextTo"
    axis_obj.title = title


def visible_category_end(ws, start_row: int, label_col: int, metric_col: int, limit: int, stop_row: int | None = None) -> int:
    last_row = start_row - 1
    upper = stop_row if stop_row is not None else ws.max_row
    count = 0
    for row in range(start_row, upper + 1):
        label = ws.cell(row, label_col).value
        value = ws.cell(row, metric_col).value
        if label in (None, "", "Total"):
            continue
        if ws.row_dimensions[row].hidden:
            continue
        if value in (None, ""):
            continue
        count += 1
        last_row = row
        if count >= limit:
            break
    return last_row


def find_col_by_header(ws, header_row: int, needles: set[str]) -> int | None:
    for col in range(1, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "").strip().lower()
        if header in needles:
            return col
    return None


def find_total_rows(ws, label_col: int = 1) -> list[int]:
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row, label_col).value
        if isinstance(value, str) and value.strip().lower() == "total":
            rows.append(row)
    return rows


def hide_zero_revenue_rows(ws, header_row: int, data_start_row: int, data_end_row: int) -> int:
    rev_col = find_col_by_header(ws, header_row, REVENUE_HEADERS | {"monthly revenue", "est. monthly retail rev"})
    if rev_col is None:
        return 0
    hidden = 0
    for row in range(data_start_row, data_end_row + 1):
        value = ws.cell(row, rev_col).value
        try:
            numeric = float(value) if value not in (None, "") else 0.0
        except (TypeError, ValueError):
            numeric = 0.0
        is_total = isinstance(ws.cell(row, 1).value, str) and ws.cell(row, 1).value.strip().lower() == "total"
        ws.row_dimensions[row].hidden = numeric == 0.0 and not is_total
        if ws.row_dimensions[row].hidden:
            hidden += 1
    return hidden


def style_total_row(ws, total_row: int, ncols: int) -> None:
    ws.row_dimensions[total_row].hidden = False
    ws.row_dimensions[total_row].height = 22
    for col in range(1, ncols + 1):
        cell = ws.cell(total_row, col)
        cell.font = total_font()
        cell.fill = total_fill()
        cell.border = total_border()
        cell.alignment = left_alignment() if col == 1 else center_alignment(False)


def add_summary_charts(ws) -> None:
    clear_charts(ws)
    total_rows = find_total_rows(ws)
    stop_row = total_rows[0] - 1 if total_rows else ws.max_row
    max_rows = visible_category_end(ws, start_row=3, label_col=1, metric_col=3, limit=10, stop_row=stop_row)
    if max_rows < 4:
        return
    labels = Reference(ws, min_col=1, min_row=3, max_row=max_rows)
    revenue = Reference(ws, min_col=3, min_row=2, max_row=max_rows)

    revenue_chart = BarChart()
    revenue_chart.type = "bar"
    revenue_chart.style = 10
    revenue_chart.title = "Top Brands - Monthly Revenue"
    revenue_chart.varyColors = True
    revenue_chart.legend = None
    revenue_chart.x_axis.title = "Monthly Revenue ($)"
    revenue_chart.x_axis.numFmt = "$#,##0"
    revenue_chart.add_data(revenue, titles_from_data=True)
    revenue_chart.set_categories(labels)
    color_bars_per_category(revenue_chart.series[0], max_rows - 2)
    annotate_category_axis(revenue_chart, axis="y", title="Brand")
    revenue_chart.dLbls = dollar_labels()
    revenue_chart.width = 32
    revenue_chart.height = 22
    ws.add_chart(revenue_chart, "I3")

    pie = PieChart()
    pie.title = "Brand Revenue Share"
    pie.style = 10
    pie_data = Reference(ws, min_col=3, min_row=3, max_row=max_rows)
    pie.add_data(pie_data)
    pie.set_categories(labels)
    pie.dLbls = pie_labels()
    pie.width = 26
    pie.height = 22
    ws.add_chart(pie, "I28")


def add_price_tier_charts(ws) -> None:
    clear_charts(ws)
    max_rows = ws.max_row
    if max_rows < 4:
        return

    revenue_chart = BarChart()
    revenue_chart.type = "col"
    revenue_chart.style = 10
    revenue_chart.varyColors = True
    revenue_chart.legend = None
    revenue_chart.title = "Revenue by Price Tier"
    revenue_chart.y_axis.title = "Revenue ($)"
    revenue_chart.y_axis.numFmt = "$#,##0"
    revenue_data = Reference(ws, min_col=2, min_row=2, max_row=max_rows)
    labels = Reference(ws, min_col=1, min_row=3, max_row=max_rows)
    revenue_chart.add_data(revenue_data, titles_from_data=True)
    revenue_chart.set_categories(labels)
    color_bars_per_category(revenue_chart.series[0], max_rows - 2)
    annotate_category_axis(revenue_chart, axis="x", title="Price Tier")
    revenue_chart.dLbls = dollar_labels()
    revenue_chart.width = 26
    revenue_chart.height = 18
    ws.add_chart(revenue_chart, "H3")

    pie = PieChart()
    pie.title = "Unit Share by Price Tier"
    pie.style = 10
    pie_data = Reference(ws, min_col=5, min_row=3, max_row=max_rows)
    pie.add_data(pie_data)
    pie.set_categories(labels)
    pie.dLbls = pie_labels()
    pie.width = 26
    pie.height = 22
    ws.add_chart(pie, "H24")

    units_chart = BarChart()
    units_chart.type = "col"
    units_chart.style = 10
    units_chart.varyColors = True
    units_chart.legend = None
    units_chart.title = "Revenue Share by Price Tier"
    units_chart.y_axis.title = "Revenue Share %"
    units_chart.y_axis.numFmt = "0%"
    share_data = Reference(ws, min_col=4, min_row=2, max_row=max_rows)
    units_chart.add_data(share_data, titles_from_data=True)
    units_chart.set_categories(labels)
    color_bars_per_category(units_chart.series[0], max_rows - 2)
    annotate_category_axis(units_chart, axis="x", title="Price Tier")
    units_chart.dLbls = pct_labels()
    units_chart.width = 26
    units_chart.height = 18
    ws.add_chart(units_chart, "H48")


def add_top50_summary_charts(ws) -> None:
    clear_charts(ws)
    sections = [section for section in section_rows(ws) if (section.data_end - section.data_start + 1) >= 2]
    if not sections:
        return
    first = sections[0]
    add_section_bar_chart(
        ws,
        section=first,
        title=f"{first.label} Revenue",
        data_col=5,
        category_col=1,
        anchor="H3",
        color=TITLE_FILL,
        dollar=True,
    )
    second = sections[1] if len(sections) > 1 else first
    add_section_pie_chart(
        ws,
        section=second,
        title=f"{second.label} Revenue Share",
        data_col=5,
        category_col=1,
        anchor="H24",
    )
    third = sections[2] if len(sections) > 2 else second
    add_section_bar_chart(
        ws,
        section=third,
        title=f"{third.label} Units",
        data_col=3,
        category_col=1,
        anchor="H48",
        color=HEADER_FILL,
        dollar=False,
    )


def add_section_bar_chart(
    ws,
    section: Section,
    title: str,
    data_col: int,
    category_col: int,
    anchor: str,
    color: str,
    dollar: bool,
) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.varyColors = True
    chart.legend = None
    chart.title = title
    chart.y_axis.title = "Revenue ($)" if dollar else "Units"
    chart.y_axis.numFmt = "$#,##0" if dollar else "#,##0"
    data = Reference(ws, min_col=data_col, min_row=section.header_row, max_row=section.data_end)
    labels = Reference(ws, min_col=category_col, min_row=section.data_start, max_row=section.data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    color_bars_per_category(chart.series[0], section.data_end - section.data_start + 1)
    annotate_category_axis(chart, axis="x", title=section.label)
    chart.dLbls = dollar_labels() if dollar else integer_labels()
    chart.width = 26
    chart.height = 18
    ws.add_chart(chart, anchor)


def add_section_pie_chart(ws, section: Section, title: str, data_col: int, category_col: int, anchor: str) -> None:
    chart = PieChart()
    chart.title = title
    chart.style = 10
    data = Reference(ws, min_col=data_col, min_row=section.data_start, max_row=section.data_end)
    labels = Reference(ws, min_col=category_col, min_row=section.data_start, max_row=section.data_end)
    chart.add_data(data)
    chart.set_categories(labels)
    chart.dLbls = pie_labels()
    chart.width = 26
    chart.height = 22
    ws.add_chart(chart, anchor)


def format_summary_sheet(ws) -> None:
    format_header_row(ws, 2, ws.max_column)
    style_data_range(ws, 3, ws.max_row, left_columns={1}, alt=True)
    format_numeric_columns(ws, 2)
    total_rows = find_total_rows(ws)
    if total_rows:
        data_end = total_rows[0] - 1
        hide_zero_revenue_rows(ws, header_row=2, data_start_row=3, data_end_row=data_end)
        for total_row in total_rows:
            style_total_row(ws, total_row, ws.max_column)
    add_summary_charts(ws)


def format_flat_sheet(ws, large: bool = False) -> None:
    format_header_row(ws, 2, ws.max_column)
    convert_links(ws, 2)
    style_data_range(ws, 3, ws.max_row, left_columns={1, 2, 3, 4, 10}, alt=not large)
    format_numeric_columns(ws, 2)
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"


def format_stacked_summary(ws) -> None:
    for section in section_rows(ws):
        format_header_row(ws, section.header_row, ws.max_column)
        style_data_range(ws, section.data_start, section.data_end, left_columns={1}, alt=True)
    format_numeric_columns(ws, 2)
    add_top50_summary_charts(ws)


def format_price_tiers(ws) -> None:
    format_header_row(ws, 2, ws.max_column)
    style_data_range(ws, 3, ws.max_row, left_columns={1}, alt=True)
    format_numeric_columns(ws, 2)
    add_price_tier_charts(ws)


def format_metadata(ws) -> None:
    if str(ws.cell(2, 1).value) != "Field" or str(ws.cell(2, 2).value) != "Value":
        ws.insert_rows(2)
        ws.cell(2, 1).value = "Field"
        ws.cell(2, 2).value = "Value"
    format_header_row(ws, 2, ws.max_column)
    style_data_range(ws, 3, ws.max_row, left_columns={1, 2}, alt=True)


def finalize_sheet(ws) -> None:
    set_column_widths(ws)
    ws.freeze_panes = "A3"


def apply_formatting(path: Path) -> None:
    wb = load_workbook(path)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    for ws in wb.worksheets:
        apply_title_row(ws, title_text_for_sheet(path, ws.title))

        if ws.title == "Summary":
            format_summary_sheet(ws)
        elif ws.title in {"Top 50 Revenue", "Top 50 Units"}:
            format_flat_sheet(ws, large=False)
        elif ws.title == "All ASINs":
            format_flat_sheet(ws, large=True)
        elif ws.title == "Top 50 Summary":
            format_stacked_summary(ws)
        elif ws.title == "Price Tiers":
            format_price_tiers(ws)
        elif ws.title == "Metadata":
            format_metadata(ws)
        else:
            format_flat_sheet(ws, large=ws.max_row > 500)

        finalize_sheet(ws)

    wb.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply branded Excel formatting to Borescope market analysis workbooks.")
    parser.add_argument("files", nargs="+", type=Path, help="Workbook files to format in place.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    for path in args.files:
        apply_formatting(path)
        print(f"Formatted {path}")


if __name__ == "__main__":
    main()
