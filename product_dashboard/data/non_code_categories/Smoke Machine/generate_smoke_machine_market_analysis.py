#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


DEFAULT_PRICE_CEILING = 1000.0
TOP_N = 50
PRICE_TIERS = [
    ("$0-40", 0, 40),
    ("$40-60", 40, 60),
    ("$60-90", 60, 90),
    ("$90+", 90, float("inf")),
]

NUMERIC_COLUMNS = ["Price", "ASIN Revenue", "ASIN Sales", "Review Count", "Reviews Rating"]
TOP_COLUMNS = [
    "ASIN",
    "Title",
    "Brand",
    "Type",
    "Price",
    "Monthly Rev",
    "Monthly Units",
    "Avg Rating",
    "# of Reviews",
    "Link",
    "Subcategory",
    "Size Tier",
    "Is Accessory",
    "Is Built-in Pump",
    "Is Includes Smoke Fluid",
    "Is Pressure Gauge",
]

ACCESSORY_PATTERN = re.compile(
    r"\b("
    r"adapter|adaptor|bladder|plug|plugs|cap|caps|cone|"
    r"hose|cable|clamp|fitting|nozzle|port|connector|replacement|bundle|"
    r"boot|skin|cylinder"
    r")\b",
    re.IGNORECASE,
)
STANDALONE_FLUID_ACCESSORY_PATTERN = re.compile(
    r"^(smoke\s+fluid\s+solution|turbo\s+boost\s+leak\s+testers\s+automotive\s+smoke\s+test\s+fluid.*pack\s+of\s+2)",
    re.IGNORECASE,
)
FLUID_PATTERN = re.compile(
    r"\b(smoke\s+fluid|fluid\s+solution|smoke\s+test\s+fluid|smoke\s+machine\s+liquid|uv\s+fluid|best\s+ranked\s+fluid)\b",
    re.IGNORECASE,
)
PUMP_PATTERN = re.compile(
    r"\b(built[-\s]?in|internal|integrated|with)\s+(?:air\s+)?(?:pump|compressor|motor)\b|"
    r"\bself[-\s]?powered\b|\bair\s+pump\b|\bair\s+compressor\b",
    re.IGNORECASE,
)
HIGH_VOLUME_PATTERN = re.compile(r"\bhigh[-\s]?volume\b", re.IGNORECASE)
LEAK_DETECTOR_PATTERN = re.compile(r"\b(leak\s+detector|leak\s+tester|diagnostic|evap|vacuum|fuel\s+leak)\b", re.IGNORECASE)
PRESSURE_GAUGE_PATTERN = re.compile(r"\b(pressure\s+gauge|flow\s*meter|flowmeter|pressure\s+meter)\b", re.IGNORECASE)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate Smoke Machine monthly market analysis workbook.")
    parser.add_argument(
        "--raw-dir",
        type=Path,
        required=True,
        help="Directory containing the monthly raw CSV files, for example raw_data/202604.",
    )
    parser.add_argument(
        "--price-ceiling",
        type=float,
        default=DEFAULT_PRICE_CEILING,
        help="Maximum allowed Price value for included ASINs.",
    )
    parser.add_argument(
        "--output-file",
        type=Path,
        default=None,
        help="Output workbook path. Defaults to outputs/Smoke_Machine_Market_Analysis_<month>.xlsx.",
    )
    return parser.parse_args()


def round2(value: float) -> float:
    return round(float(value), 2)


def round4(value: float) -> float:
    return round(float(value), 4)


def infer_snapshot_month(raw_dir: Path) -> str:
    match = re.search(r"(\d{6})", str(raw_dir))
    if not match:
        raise ValueError(f"Could not infer snapshot month from raw dir: {raw_dir}")
    return match.group(1)


def month_end(snapshot_month: str) -> str:
    year = int(snapshot_month[:4])
    month = int(snapshot_month[4:6])
    return date(year, month, calendar.monthrange(year, month)[1]).isoformat()


def latest_run_date(raw_dir: Path, fallback_snapshot_date: str) -> str:
    dates = sorted(
        {
            match.group(1)
            for file_path in raw_dir.glob("*.csv")
            if (match := re.search(r"(\d{4}-\d{2}-\d{2})", file_path.name))
        }
    )
    return dates[-1] if dates else fallback_snapshot_date


def infer_output_path(raw_dir: Path) -> Path:
    snapshot_month = infer_snapshot_month(raw_dir)
    return raw_dir.parent.parent / "outputs" / f"Smoke_Machine_Market_Analysis_{snapshot_month}.xlsx"


def load_all_raw_data(raw_dir: Path) -> pd.DataFrame:
    csv_files = sorted(raw_dir.glob("*.csv"))
    if not csv_files:
        raise FileNotFoundError(f"No CSV files found in {raw_dir}")

    frames: list[pd.DataFrame] = []
    for csv_file in csv_files:
        frame = pd.read_csv(csv_file, encoding="utf-8-sig")
        frame["source_file"] = csv_file.name
        frames.append(frame)

    return pd.concat(frames, ignore_index=True)


def price_per_unit(source: pd.DataFrame, revenue_column: str = "ASIN Revenue", units_column: str = "ASIN Sales") -> float:
    revenue = float(source[revenue_column].fillna(0).sum())
    units = float(source[units_column].fillna(0).sum())
    if units > 0:
        return round2(revenue / units)
    fallback = source["Price"].dropna()
    return round2(float(fallback.mean())) if not fallback.empty else 0


def classify_type(title: object) -> str:
    title_text = str(title or "")
    if re.search(r"\bprofessional\s+automotive\s+smoke\s+machine\s+kit\b", title_text, re.IGNORECASE):
        return "Leak Detector Kit"
    if re.search(r"\bheavy\s+duty\s+truck\s+diagnostic\s+smoke\s+machine\b", title_text, re.IGNORECASE):
        return "Smoke Machine"
    if STANDALONE_FLUID_ACCESSORY_PATTERN.search(title_text):
        return "Accessory"
    if ACCESSORY_PATTERN.search(title_text):
        return "Accessory"
    if FLUID_PATTERN.search(title_text):
        return "Smoke Fluid"
    if HIGH_VOLUME_PATTERN.search(title_text):
        return "High-volume Smoke Machine"
    if re.search(r"\bsmoke\s+machine\s+automotive\s+tool\b", title_text, re.IGNORECASE):
        return "Smoke Machine"
    if LEAK_DETECTOR_PATTERN.search(title_text):
        return "Leak Detector Kit"
    return "Smoke Machine"


def is_accessory(title: object, product_type: str) -> str:
    return "Yes" if product_type == "Accessory" or ACCESSORY_PATTERN.search(str(title or "")) else "No"


def has_builtin_pump(title: object) -> str:
    return "Yes" if PUMP_PATTERN.search(str(title or "")) else "No"


def includes_smoke_fluid(title: object) -> str:
    return "Yes" if FLUID_PATTERN.search(str(title or "")) else "No"


def has_pressure_gauge(title: object) -> str:
    return "Yes" if PRESSURE_GAUGE_PATTERN.search(str(title or "")) else "No"


def load_processed_records(raw_dir: Path, price_ceiling: float) -> pd.DataFrame:
    df = load_all_raw_data(raw_dir)
    if "ASIN" not in df.columns:
        raise ValueError("Raw data is missing ASIN column.")

    before = len(df)
    df = df.drop_duplicates(subset=["ASIN"]).copy()
    print(f"Deduplicated ASINs: {before} -> {len(df)} rows")

    for column in NUMERIC_COLUMNS:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce")

    df = df[df["Price"].notna() & (df["Price"] > 0) & (df["Price"] <= price_ceiling)].copy()

    if "Shipping Details" in df.columns and "Size Tier" not in df.columns:
        df["Size Tier"] = df["Shipping Details"]

    df["Type"] = df["Title"].apply(classify_type)
    df["Is Accessory"] = df.apply(lambda row: is_accessory(row.get("Title"), row.get("Type")), axis=1)
    df["Is Built-in Pump"] = df["Title"].apply(has_builtin_pump)
    df["Is Includes Smoke Fluid"] = df["Title"].apply(includes_smoke_fluid)
    df["Is Pressure Gauge"] = df["Title"].apply(has_pressure_gauge)
    return df.reset_index(drop=True)


def weighted_rating(source: pd.DataFrame) -> float:
    total_reviews = float(source["Review Count"].fillna(0).sum())
    if total_reviews > 0:
        weighted = (source["Reviews Rating"].fillna(0) * source["Review Count"].fillna(0)).sum() / total_reviews
        return round2(weighted)
    ratings = source["Reviews Rating"].dropna()
    return round2(float(ratings.mean())) if not ratings.empty else 0


def build_brand_summary(df: pd.DataFrame) -> pd.DataFrame:
    grouped = (
        df.groupby("Brand", dropna=False)
        .agg(
            listings=("ASIN", "nunique"),
            revenue=("ASIN Revenue", "sum"),
            units=("ASIN Sales", "sum"),
            reviews=("Review Count", "sum"),
        )
        .reset_index()
    )

    grouped["Avg Rating"] = [
        weighted_rating(df[df["Brand"].eq(brand)])
        if pd.notna(brand)
        else weighted_rating(df[df["Brand"].isna()])
        for brand in grouped["Brand"]
    ]
    total_revenue = float(grouped["revenue"].fillna(0).sum())
    total_units = float(grouped["units"].fillna(0).sum())
    grouped["Monthly Rev"] = grouped["revenue"].fillna(0).map(round2)
    grouped["Monthly Units"] = grouped["units"].fillna(0).map(round2)
    grouped["Monthly Rev Market Share %"] = grouped["revenue"].fillna(0).apply(
        lambda value: round4(value / total_revenue) if total_revenue > 0 else 0
    )
    grouped["Price Per Unit"] = [
        price_per_unit(df[df["Brand"].eq(brand)]) if pd.notna(brand) else price_per_unit(df[df["Brand"].isna()])
        for brand in grouped["Brand"]
    ]

    summary = grouped[
        ["Brand", "listings", "Monthly Rev", "Monthly Units", "Monthly Rev Market Share %", "Price Per Unit", "Avg Rating"]
    ].rename(columns={"listings": "# of Listings"})
    summary = summary.sort_values("Monthly Rev", ascending=False).reset_index(drop=True)
    total_row = {
        "Brand": "Total",
        "# of Listings": int(summary["# of Listings"].sum()),
        "Monthly Rev": round2(total_revenue),
        "Monthly Units": round2(total_units),
        "Monthly Rev Market Share %": 1 if total_revenue > 0 else 0,
        "Price Per Unit": price_per_unit(df),
        "Avg Rating": 0,
    }
    return pd.concat([summary, pd.DataFrame([total_row])], ignore_index=True)


def build_top_sheet_rows(df: pd.DataFrame) -> pd.DataFrame:
    rows = {
        "ASIN": df["ASIN"],
        "Title": df["Title"],
        "Brand": df["Brand"],
        "Type": df["Type"],
        "Price": df["Price"].map(round2),
        "Monthly Rev": df["ASIN Revenue"].fillna(0).map(round2),
        "Monthly Units": df["ASIN Sales"].fillna(0).map(round2),
        "Avg Rating": df["Reviews Rating"].fillna(0).map(round2),
        "# of Reviews": df["Review Count"].fillna(0).map(round2),
        "Link": df["URL"].fillna(""),
        "Subcategory": df["Subcategory"].fillna(""),
        "Size Tier": df["Size Tier"].fillna(""),
        "Is Accessory": df["Is Accessory"],
        "Is Built-in Pump": df["Is Built-in Pump"],
        "Is Includes Smoke Fluid": df["Is Includes Smoke Fluid"],
        "Is Pressure Gauge": df["Is Pressure Gauge"],
    }
    for column in TOP_COLUMNS:
        if column not in rows:
            rows[column] = None
    return pd.DataFrame(rows, columns=TOP_COLUMNS)


def build_summary_section(df: pd.DataFrame, label_column: str) -> list[list[object]]:
    total_revenue = float(df["ASIN Revenue"].fillna(0).sum())
    total_units = float(df["ASIN Sales"].fillna(0).sum())

    grouped = (
        df.groupby(label_column, dropna=False)
        .agg(
            units=("ASIN Sales", "sum"),
            revenue=("ASIN Revenue", "sum"),
        )
        .reset_index()
    )
    grouped[label_column] = grouped[label_column].fillna("Unknown")
    grouped["Qty By %"] = grouped["units"].fillna(0).apply(lambda value: round4(value / total_units) if total_units > 0 else 0)
    grouped["Revenue By %"] = grouped["revenue"].fillna(0).apply(
        lambda value: round4(value / total_revenue) if total_revenue > 0 else 0
    )
    grouped["Avg Price"] = [
        price_per_unit(df[df[label_column].eq(label)])
        for label in grouped[label_column]
    ]
    grouped["Quantity/Mo"] = grouped["units"].fillna(0).map(round2)
    grouped["Revenue/Mo"] = grouped["revenue"].fillna(0).map(round2)
    grouped = grouped.sort_values("Revenue/Mo", ascending=False)

    rows: list[list[object]] = [[label_column, "Avg Price", "Quantity/Mo", "Qty By %", "Revenue/Mo", "Revenue By %"]]
    for _, row in grouped.iterrows():
        rows.append(
            [
                row[label_column],
                row["Avg Price"],
                row["Quantity/Mo"],
                row["Qty By %"],
                row["Revenue/Mo"],
                row["Revenue By %"],
            ]
        )
    return rows


def build_price_tiers(df: pd.DataFrame) -> pd.DataFrame:
    total_revenue = float(df["ASIN Revenue"].fillna(0).sum())
    total_units = float(df["ASIN Sales"].fillna(0).sum())
    rows: list[dict[str, object]] = []
    for label, minimum, maximum in PRICE_TIERS:
        matched = df[(df["Price"] >= minimum) & (df["Price"] < maximum)].copy()
        revenue = float(matched["ASIN Revenue"].fillna(0).sum())
        units = float(matched["ASIN Sales"].fillna(0).sum())
        rows.append(
            {
                "Price Tier": label,
                "Total Revenue": round2(revenue),
                "Total Sales": round2(units),
                "Rev Share %": round4(revenue / total_revenue) if total_revenue > 0 else 0,
                "Unit Share %": round4(units / total_units) if total_units > 0 else 0,
                "Avg Price": price_per_unit(matched),
            }
        )
    return pd.DataFrame(rows)


def build_metadata_rows(snapshot_month: str, snapshot_date: str, run_date: str, asin_count: int) -> list[list[object]]:
    return [
        ["Category", "Smoke Machine"],
        ["Category ID", "smoke_machine"],
        ["Snapshot Month", snapshot_month],
        ["Snapshot Date", snapshot_date],
        ["Latest Run Date", run_date],
        ["Generated At", datetime.now().isoformat()],
        ["ASIN Count", asin_count],
    ]


def autosize_columns(worksheet) -> None:
    widths: dict[int, int] = {}
    for row in worksheet.iter_rows(values_only=True):
        for index, value in enumerate(row, start=1):
            length = len(str(value)) if value is not None else 0
            widths[index] = min(max(widths.get(index, 10), length + 2), 80)
    for index, width in widths.items():
        worksheet.column_dimensions[get_column_letter(index)].width = width


def append_dataframe_sheet(workbook: Workbook, name: str, frame: pd.DataFrame) -> None:
    ws = workbook.create_sheet(name)
    ws.append(list(frame.columns))
    for row in frame.astype(object).where(pd.notnull(frame), None).itertuples(index=False, name=None):
        ws.append(list(row))
    autosize_columns(ws)


def append_rows_sheet(workbook: Workbook, name: str, rows: list[list[object]]) -> None:
    ws = workbook.create_sheet(name)
    for row in rows:
        ws.append(row)
    autosize_columns(ws)


def build_workbook(df: pd.DataFrame, snapshot_month: str, snapshot_date: str, run_date: str) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    top_revenue_source = df.sort_values("ASIN Revenue", ascending=False).head(TOP_N).copy()
    top_units_source = df.sort_values("ASIN Sales", ascending=False).head(TOP_N).copy()
    all_asins_source = df.sort_values("ASIN Revenue", ascending=False).copy()

    append_dataframe_sheet(workbook, "Summary", build_brand_summary(df))
    append_dataframe_sheet(workbook, "Top 50 Revenue", build_top_sheet_rows(top_revenue_source))
    append_dataframe_sheet(workbook, "Top 50 Units", build_top_sheet_rows(top_units_source))
    append_rows_sheet(workbook, "Top 50 Summary", build_summary_section(top_revenue_source, "Type"))
    append_dataframe_sheet(workbook, "Price Tiers", build_price_tiers(df))
    append_dataframe_sheet(workbook, "All ASINs", build_top_sheet_rows(all_asins_source))
    append_rows_sheet(workbook, "Metadata", build_metadata_rows(snapshot_month, snapshot_date, run_date, len(df)))
    return workbook


def main() -> None:
    args = parse_args()
    snapshot_month = infer_snapshot_month(args.raw_dir)
    snapshot_date = month_end(snapshot_month)
    run_date = latest_run_date(args.raw_dir, snapshot_date)
    output_file = args.output_file or infer_output_path(args.raw_dir)

    processed = load_processed_records(args.raw_dir, args.price_ceiling)
    workbook = build_workbook(processed, snapshot_month, snapshot_date, run_date)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_file)

    print(f"Wrote workbook: {output_file}")
    print(f"Rows: {len(processed)}")
    print(f"Revenue: {round2(processed['ASIN Revenue'].fillna(0).sum())}")
    print(f"Units: {round2(processed['ASIN Sales'].fillna(0).sum())}")
    print(f"Price per unit: {price_per_unit(processed)}")
    print(f"Max price: {round2(processed['Price'].max()) if not processed.empty else 0}")


if __name__ == "__main__":
    main()
