#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)
CHANNEL_SORT = {"1P": 0, "3P": 1, "Total": 2}

CHANNEL_UNITS_SOURCE = {
    "1P": "Ordered Units",
    "3P": "Units Ordered",
    "Total": "1P Ordered Units + 3P Units Ordered",
}

CHANNEL_REVENUE_SOURCE = {
    "1P": "Ordered Revenue",
    "3P": "Ordered Product Sales",
    "Total": "1P Ordered Revenue + 3P Ordered Product Sales",
}


def parse_args() -> argparse.Namespace:
    base_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Build a combined INNOVA DMM 1P/3P monthly analysis workbook.")
    parser.add_argument(
        "--input-1p",
        type=Path,
        default=base_dir / "innova1p_data",
        help="Folder containing monthly INNOVA 1P CSV files.",
    )
    parser.add_argument(
        "--input-3p",
        type=Path,
        default=base_dir / "innova3p_data",
        help="Folder containing monthly INNOVA 3P CSV files.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=base_dir / "outputs" / "Innova_DMM_1P_3P_Monthly_Analysis.xlsx",
        help="Output workbook path.",
    )
    return parser.parse_args()


def parse_currency(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text in {"-", "N/A", "nan"}:
        return None
    return float(text.replace("$", "").replace(",", ""))


def parse_number(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text in {"-", "N/A", "nan"}:
        return None
    return float(text.replace(",", ""))


def parse_int(value: object) -> int | None:
    number = parse_number(value)
    return None if number is None else int(round(number))


def parse_percent(value: object) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text in {"-", "N/A", "nan"}:
        return None
    if text.endswith("%"):
        return float(text[:-1].replace(",", "")) / 100
    return float(text)


def month_end_from_start(month_start: datetime) -> datetime:
    year = month_start.year + (1 if month_start.month == 12 else 0)
    month = 1 if month_start.month == 12 else month_start.month + 1
    return datetime(year, month, 1)


def load_1p_records(input_dir: Path) -> list[dict[str, object]]:
    pattern = re.compile(r"Custom_(\d+-\d+-\d+)_(\d+-\d+-\d+)\.csv$")
    records: list[dict[str, object]] = []

    for csv_path in sorted(input_dir.glob("*.csv")):
        match = pattern.search(csv_path.name)
        if not match:
            raise SystemExit(f"Could not parse month range from 1P filename: {csv_path.name}")
        month_start = datetime.strptime(match.group(1), "%m-%d-%Y")
        month_end = datetime.strptime(match.group(2), "%m-%d-%Y")

        df = pd.read_csv(csv_path, skiprows=1, encoding="utf-8-sig")
        for row in df.to_dict(orient="records"):
            ordered_revenue = parse_currency(row.get("Ordered Revenue"))
            ordered_units = parse_int(row.get("Ordered Units"))
            shipped_revenue = parse_currency(row.get("Shipped Revenue"))
            shipped_cogs = parse_currency(row.get("Shipped COGS"))
            shipped_units = parse_int(row.get("Shipped Units"))
            customer_returns = parse_int(row.get("Customer Returns"))
            asin = str(row.get("ASIN", "")).strip()
            records.append(
                {
                    "report_month": month_start.strftime("%Y-%m"),
                    "month_start": month_start.date(),
                    "month_end": month_end.date(),
                    "channel_type": "1P",
                    "units_metric_source": CHANNEL_UNITS_SOURCE["1P"],
                    "revenue_metric_source": CHANNEL_REVENUE_SOURCE["1P"],
                    "asin": asin,
                    "title": row.get("Product Title"),
                    "brand": row.get("Brand"),
                    "parent_asin": None,
                    "child_asin": asin,
                    "sku": None,
                    "channel_ordered_revenue_usd": ordered_revenue,
                    "channel_ordered_units": ordered_units,
                    "average_selling_price": (ordered_revenue / ordered_units) if ordered_revenue and ordered_units else None,
                    "ordered_revenue_usd": ordered_revenue,
                    "ordered_units": ordered_units,
                    "shipped_revenue_usd": shipped_revenue,
                    "shipped_cogs_usd": shipped_cogs,
                    "shipped_units": shipped_units,
                    "customer_returns": customer_returns,
                    "sessions_total": None,
                    "page_views_total": None,
                    "buy_box_pct": None,
                    "total_order_items": None,
                    "source_file": csv_path.name,
                }
            )
    return records


def load_3p_records(input_dir: Path) -> list[dict[str, object]]:
    pattern = re.compile(r"(\d{4})(\d{2})-\d{2}innova3p.*\.csv$", re.IGNORECASE)
    records: list[dict[str, object]] = []

    for csv_path in sorted(input_dir.glob("*.csv")):
        match = pattern.search(csv_path.name)
        if not match:
            raise SystemExit(f"Could not parse month range from 3P filename: {csv_path.name}")
        month_start = datetime(int(match.group(1)), int(match.group(2)), 1)
        month_end = month_end_from_start(month_start)

        df = pd.read_csv(csv_path, encoding="utf-8-sig")
        for row in df.to_dict(orient="records"):
            ordered_sales = parse_currency(row.get("Ordered Product Sales"))
            units_ordered = parse_int(row.get("Units Ordered"))
            total_order_items = parse_int(row.get("Total Order Items"))
            sessions_total = parse_int(row.get("Sessions - Total"))
            page_views_total = parse_int(row.get("Page Views - Total"))
            buy_box_pct = parse_percent(row.get("Featured Offer (Buy Box) Percentage"))
            child_asin = str(row.get("(Child) ASIN", "")).strip()
            records.append(
                {
                    "report_month": month_start.strftime("%Y-%m"),
                    "month_start": month_start.date(),
                    "month_end": month_end.date(),
                    "channel_type": "3P",
                    "units_metric_source": CHANNEL_UNITS_SOURCE["3P"],
                    "revenue_metric_source": CHANNEL_REVENUE_SOURCE["3P"],
                    "asin": child_asin,
                    "title": row.get("Title"),
                    "brand": "Innova",
                    "parent_asin": row.get("(Parent) ASIN"),
                    "child_asin": child_asin,
                    "sku": row.get("SKU"),
                    "channel_ordered_revenue_usd": ordered_sales,
                    "channel_ordered_units": units_ordered,
                    "average_selling_price": (ordered_sales / units_ordered) if ordered_sales and units_ordered else None,
                    "ordered_revenue_usd": ordered_sales,
                    "ordered_units": units_ordered,
                    "shipped_revenue_usd": None,
                    "shipped_cogs_usd": None,
                    "shipped_units": None,
                    "customer_returns": None,
                    "sessions_total": sessions_total,
                    "page_views_total": page_views_total,
                    "buy_box_pct": buy_box_pct,
                    "total_order_items": total_order_items,
                    "source_file": csv_path.name,
                }
            )
    return records


def build_combined_detail(input_1p: Path, input_3p: Path) -> pd.DataFrame:
    records = load_1p_records(input_1p) + load_3p_records(input_3p)
    if not records:
        raise SystemExit("No 1P or 3P records found.")
    detail = pd.DataFrame(records)
    detail["channel_sort"] = detail["channel_type"].map(CHANNEL_SORT)
    detail = detail.sort_values(["month_start", "channel_sort", "asin"]).drop(columns=["channel_sort"]).reset_index(drop=True)
    return detail


def build_monthly_summary(detail: pd.DataFrame) -> pd.DataFrame:
    base = (
        detail.groupby(["report_month", "month_start", "month_end", "channel_type"], as_index=False)
        .agg(
            channel_ordered_revenue_usd=("channel_ordered_revenue_usd", "sum"),
            channel_ordered_units=("channel_ordered_units", "sum"),
            ordered_revenue_usd=("ordered_revenue_usd", "sum"),
            ordered_units=("ordered_units", "sum"),
            shipped_revenue_usd=("shipped_revenue_usd", "sum"),
            shipped_cogs_usd=("shipped_cogs_usd", "sum"),
            shipped_units=("shipped_units", "sum"),
            customer_returns=("customer_returns", "sum"),
            sessions_total=("sessions_total", "sum"),
            page_views_total=("page_views_total", "sum"),
            total_order_items=("total_order_items", "sum"),
            asin_count=("asin", "nunique"),
        )
        .reset_index(drop=True)
    )
    total = (
        detail.groupby(["report_month", "month_start", "month_end"], as_index=False)
        .agg(
            channel_ordered_revenue_usd=("channel_ordered_revenue_usd", "sum"),
            channel_ordered_units=("channel_ordered_units", "sum"),
            ordered_revenue_usd=("ordered_revenue_usd", "sum"),
            ordered_units=("ordered_units", "sum"),
            shipped_revenue_usd=("shipped_revenue_usd", "sum"),
            shipped_cogs_usd=("shipped_cogs_usd", "sum"),
            shipped_units=("shipped_units", "sum"),
            customer_returns=("customer_returns", "sum"),
            sessions_total=("sessions_total", "sum"),
            page_views_total=("page_views_total", "sum"),
            total_order_items=("total_order_items", "sum"),
            asin_count=("asin", "nunique"),
        )
        .assign(channel_type="Total")
    )
    summary = pd.concat([base, total], ignore_index=True)
    summary["units_metric_source"] = summary["channel_type"].map(CHANNEL_UNITS_SOURCE)
    summary["revenue_metric_source"] = summary["channel_type"].map(CHANNEL_REVENUE_SOURCE)
    summary["average_selling_price"] = summary["channel_ordered_revenue_usd"] / summary["channel_ordered_units"]
    summary.loc[summary["channel_ordered_units"].fillna(0) == 0, "average_selling_price"] = None
    summary["channel_sort"] = summary["channel_type"].map(CHANNEL_SORT)
    summary = summary.sort_values(["month_start", "channel_sort"]).drop(columns=["channel_sort"]).reset_index(drop=True)
    return summary


def build_asin_monthly(detail: pd.DataFrame) -> pd.DataFrame:
    asin_monthly = (
        detail.groupby(
            ["report_month", "month_start", "month_end", "channel_type", "asin", "title", "brand"],
            as_index=False,
        )
        .agg(
            channel_ordered_revenue_usd=("channel_ordered_revenue_usd", "sum"),
            channel_ordered_units=("channel_ordered_units", "sum"),
            ordered_revenue_usd=("ordered_revenue_usd", "sum"),
            ordered_units=("ordered_units", "sum"),
            shipped_revenue_usd=("shipped_revenue_usd", "sum"),
            shipped_cogs_usd=("shipped_cogs_usd", "sum"),
            shipped_units=("shipped_units", "sum"),
            customer_returns=("customer_returns", "sum"),
            sessions_total=("sessions_total", "sum"),
            page_views_total=("page_views_total", "sum"),
            total_order_items=("total_order_items", "sum"),
        )
        .reset_index(drop=True)
    )
    asin_monthly["units_metric_source"] = asin_monthly["channel_type"].map(CHANNEL_UNITS_SOURCE)
    asin_monthly["revenue_metric_source"] = asin_monthly["channel_type"].map(CHANNEL_REVENUE_SOURCE)
    asin_monthly["average_selling_price"] = asin_monthly["channel_ordered_revenue_usd"] / asin_monthly["channel_ordered_units"]
    asin_monthly.loc[asin_monthly["channel_ordered_units"].fillna(0) == 0, "average_selling_price"] = None
    asin_monthly["channel_sort"] = asin_monthly["channel_type"].map(CHANNEL_SORT)
    asin_monthly = asin_monthly.sort_values(["month_start", "channel_sort", "asin"]).drop(columns=["channel_sort"]).reset_index(drop=True)
    return asin_monthly


def build_overview_tables(detail: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    coverage = (
        detail.groupby("channel_type", as_index=False)
        .agg(
            months_covered=("report_month", "nunique"),
            first_month=("report_month", "min"),
            last_month=("report_month", "max"),
            unique_asins=("asin", "nunique"),
            total_channel_ordered_revenue_usd=("channel_ordered_revenue_usd", "sum"),
            total_channel_ordered_units=("channel_ordered_units", "sum"),
        )
        .sort_values("channel_type")
        .reset_index(drop=True)
    )
    total_row = pd.DataFrame(
        [
            {
                "channel_type": "Total",
                "months_covered": detail["report_month"].nunique(),
                "first_month": detail["report_month"].min(),
                "last_month": detail["report_month"].max(),
                "unique_asins": detail["asin"].nunique(),
                "total_channel_ordered_revenue_usd": detail["channel_ordered_revenue_usd"].sum(),
                "total_channel_ordered_units": detail["channel_ordered_units"].sum(),
            }
        ]
    )
    coverage = pd.concat([coverage, total_row], ignore_index=True)

    notes = pd.DataFrame(
        [
            {"note": "1P files cover 2025-01 through 2026-03. 3P files currently cover 2026-01 through 2026-03 only."},
            {"note": "channel_type marks whether each row came from INNOVA 1P retail reporting or INNOVA 3P seller business reports."},
            {"note": "For 1P, the comparable units/revenue fields use Ordered Units and Ordered Revenue."},
            {"note": "For 3P, the comparable units/revenue fields use Units Ordered and Ordered Product Sales."},
            {"note": "1P-specific shipped metrics remain blank for 3P rows, and 3P traffic metrics remain blank for 1P rows."},
        ]
    )
    return coverage, notes


def autosize_columns(worksheet) -> None:
    for idx, column_cells in enumerate(worksheet.columns, start=1):
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_len = max((len(value) for value in values), default=0)
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 36)


def style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def apply_number_formats(worksheet, header_map: dict[str, int]) -> None:
    currency_cols = {
        "channel_ordered_revenue_usd",
        "average_selling_price",
        "ordered_revenue_usd",
        "shipped_revenue_usd",
        "shipped_cogs_usd",
        "total_channel_ordered_revenue_usd",
    }
    integer_cols = {
        "channel_ordered_units",
        "ordered_units",
        "shipped_units",
        "customer_returns",
        "sessions_total",
        "page_views_total",
        "total_order_items",
        "asin_count",
        "months_covered",
        "unique_asins",
        "total_channel_ordered_units",
    }
    percent_cols = {"buy_box_pct"}
    date_cols = {"month_start", "month_end"}

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            header = next((name for name, idx in header_map.items() if idx == cell.column), None)
            if not header or cell.value is None:
                continue
            if header in currency_cols:
                cell.number_format = '$#,##0.00'
            elif header in integer_cols:
                cell.number_format = '#,##0'
            elif header in percent_cols:
                cell.number_format = '0.00%'
            elif header in date_cols:
                cell.number_format = 'yyyy-mm-dd'


def write_dataframe_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(list(frame.columns))
    style_header(worksheet)
    for row in frame.astype(object).where(pd.notnull(frame), None).itertuples(index=False, name=None):
        worksheet.append(list(row))
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    header_map = {cell.value: cell.column for cell in worksheet[1]}
    apply_number_formats(worksheet, header_map)
    autosize_columns(worksheet)


def write_overview_sheet(workbook: Workbook, coverage: pd.DataFrame, notes: pd.DataFrame) -> None:
    worksheet = workbook.create_sheet("overview")
    worksheet["A1"] = "INNOVA DMM 1P + 3P Combined Monthly Analysis"
    worksheet["A1"].font = Font(bold=True, size=14)
    worksheet["A3"] = "Coverage Summary"
    worksheet["A3"].font = SECTION_FONT

    start_row = 4
    worksheet.append([])
    for idx, row in enumerate([list(coverage.columns)] + coverage.astype(object).where(pd.notnull(coverage), None).values.tolist(), start=start_row):
        for col_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=idx, column=col_idx, value=value)
            if idx == start_row:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    notes_start = start_row + len(coverage) + 3
    worksheet.cell(row=notes_start, column=1, value="Notes").font = SECTION_FONT
    for row_offset, note in enumerate(notes["note"].tolist(), start=1):
        worksheet.cell(row=notes_start + row_offset, column=1, value=note)

    header_map = {cell.value: cell.column for cell in worksheet[start_row]}
    for row in worksheet.iter_rows(min_row=start_row + 1, max_row=start_row + len(coverage)):
        for cell in row:
            header = next((name for name, idx in header_map.items() if idx == cell.column), None)
            if header in {"total_channel_ordered_revenue_usd"} and cell.value is not None:
                cell.number_format = '$#,##0.00'
            elif header in {"months_covered", "unique_asins", "total_channel_ordered_units"} and cell.value is not None:
                cell.number_format = '#,##0'

    autosize_columns(worksheet)


def build_workbook(output_path: Path, detail: pd.DataFrame) -> None:
    monthly_summary = build_monthly_summary(detail)
    asin_monthly = build_asin_monthly(detail)
    coverage, notes = build_overview_tables(detail)

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_overview_sheet(workbook, coverage, notes)
    write_dataframe_sheet(workbook, "monthly_summary", monthly_summary)
    write_dataframe_sheet(workbook, "asin_monthly", asin_monthly)
    write_dataframe_sheet(workbook, "combined_detail", detail)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    detail = build_combined_detail(args.input_1p, args.input_3p)
    build_workbook(args.output, detail)
    print(f"Wrote workbook: {args.output}")
    print(f"Combined rows: {len(detail)}")
    print(f"Months covered: {detail['report_month'].nunique()}")
    print(f"Channels: {sorted(detail['channel_type'].unique().tolist())}")


if __name__ == "__main__":
    main()
