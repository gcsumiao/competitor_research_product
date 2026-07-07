#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font


NUMERIC_COLUMNS = [
    "Price",
    "BSR",
    "Subcategory BSR",
    "Parent Level Sales",
    "ASIN Sales",
    "Parent Level Revenue",
    "ASIN Revenue",
    "Review Count",
    "Reviews Rating",
    "Price Trend (90 days) (%)",
    "Sales Trend (90 days) (%)",
    "Last Year Sales",
    "Sales Year Over Year (%)",
    "Length",
    "Width",
    "Height",
    "Weight",
    "Storage Fee (Jan - Sep)",
    "Storage Fee (Oct - Dec)",
    "Listing Age (Months)",
    "Age (Month)",
    "Number of Images",
    "Variation Count",
    "Sales to Reviews",
]

AUTO_CORE_COLUMNS = [
    "URL",
    "Image URL",
    "ASIN",
    "Title",
    "Brand",
    "Fulfillment",
    "Category",
    "BSR",
    "UPC",
    "GTIN",
    "EAN",
    "ISBN",
    "Subcategory",
    "Subcategory BSR",
    "Price",
    "Price Trend (90 days) (%)",
    "Parent Level Sales",
    "ASIN Sales",
    "Sales Trend (90 days) (%)",
    "Parent Level Revenue",
    "ASIN Revenue",
    "Review Count",
    "Reviews Rating",
    "Seller",
    "Seller Country/Region",
    "Number of Active Sellers",
    "Last Year Sales",
    "Sales Year Over Year (%)",
    "Size Tier",
    "Length",
    "Width",
    "Height",
    "Weight",
    "Storage Fee (Jan - Sep)",
    "Storage Fee (Oct - Dec)",
    "Best Sales Period",
    "Age (Month)",
    "Number of Images",
    "Variation Count",
    "Sales to Reviews",
    "source_file",
    "is_multimeter",
    "is_analyzer",
    "is_large_screen_like",
    "is_rechargeable",
    "is_automotive_targeted",
    "product_type",
    "brand_clean",
    "is_innova",
    "price_tier",
]

TOP_ASIN_COLUMNS = [
    "ASIN",
    "Title",
    "brand_clean",
    "product_type",
    "Price",
    "ASIN Sales",
    "ASIN Revenue",
    "Review Count",
    "Reviews Rating",
    "is_large_screen_like",
    "is_rechargeable",
    "is_automotive_targeted",
]

FEATURE_COLUMNS = [
    "is_large_screen_like",
    "total_revenue",
    "total_sales",
    "asin_count",
    "avg_price",
    "avg_rating",
    "total_reviews",
    "feature_flag",
    "is_rechargeable",
    "is_automotive_targeted",
]

PRICE_TIER_LABELS = ["<$30", "$30–59", "$60–99", "$100–199", "$200+"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate the DMM market research workbook from raw CSV files.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path(__file__).resolve().parent / "raw_data" / "202603",
        help="Directory containing the monthly raw CSV exports.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parent / "outputs" / "DMM_market_research_summary_202603.xlsx",
        help="Output workbook path.",
    )
    return parser.parse_args()


def load_raw_csvs(input_dir: Path) -> pd.DataFrame:
    files = sorted(input_dir.glob("*.csv"))
    if not files:
        raise SystemExit(f"No CSV files found in {input_dir}")

    frames: list[pd.DataFrame] = []
    for file_path in files:
        df = pd.read_csv(file_path, encoding="utf-8-sig")
        df["source_file"] = str(file_path)
        frames.append(df)

    raw = pd.concat(frames, ignore_index=True)
    raw = raw.drop_duplicates(subset=["ASIN"], keep="first").copy()

    if "Shipping Details" in raw.columns and "Size Tier" not in raw.columns:
        raw["Size Tier"] = raw["Shipping Details"]
    if "Listing Age (Months)" in raw.columns and "Age (Month)" not in raw.columns:
        raw["Age (Month)"] = raw["Listing Age (Months)"]

    for column in NUMERIC_COLUMNS:
        if column in raw.columns:
            raw[column] = pd.to_numeric(raw[column], errors="coerce")

    return raw


def classify_core_market(raw: pd.DataFrame) -> pd.DataFrame:
    automotive = raw[raw["Category"].astype(str).str.contains("Automotive", case=False, na=False)].copy()
    automotive = automotive[
        automotive["Subcategory"].astype(str).str.contains("Multimeter|Analyzer", case=False, na=False)
    ].copy()

    title_lower = automotive["Title"].fillna("").str.lower()
    automotive["is_multimeter"] = title_lower.str.contains("multimeter")
    automotive["is_analyzer"] = title_lower.str.contains("analyzer|analyser")

    screen_pattern = re.compile(
        r"large screen|large display|big screen|color screen|hd screen|hd display|lcd display|full screen|smart display",
        re.IGNORECASE,
    )
    automotive["is_large_screen_like"] = title_lower.str.contains(screen_pattern)
    automotive["is_rechargeable"] = title_lower.str.contains(r"rechargeable|type-c|usb c|usb-c")
    automotive["is_automotive_targeted"] = title_lower.str.contains(r"car|automotive|vehicle|battery tester")

    automotive["product_type"] = np.select(
        [
            automotive["is_multimeter"] & automotive["is_analyzer"],
            automotive["is_multimeter"],
            automotive["is_analyzer"],
        ],
        ["Multimeter + Analyzer", "Multimeter", "Analyzer"],
        default="Other",
    )
    automotive["brand_clean"] = automotive["Brand"].fillna("UNKNOWN").astype(str).str.strip().str.upper()
    automotive["is_innova"] = automotive["brand_clean"].str.contains("INNOVA", case=False, na=False)
    automotive["price_tier"] = pd.cut(
        automotive["Price"],
        bins=[0, 30, 60, 100, 200, np.inf],
        labels=PRICE_TIER_LABELS,
        right=False,
    )

    core = automotive.copy()
    core = core[core["ASIN Sales"].notna()]
    core = core[core["ASIN Revenue"].notna()]
    core = core[core["Price"] > 0]
    core = core[core["ASIN Sales"] > 0]
    core = core[core["ASIN Revenue"] > 0]

    for column in AUTO_CORE_COLUMNS:
        if column not in core.columns:
            core[column] = None

    return core[AUTO_CORE_COLUMNS].copy()


def with_share_columns(df: pd.DataFrame, total_revenue: float, total_sales: float) -> pd.DataFrame:
    result = df.copy()
    result["rev_share_%"] = np.where(total_revenue > 0, result["total_revenue"] / total_revenue * 100, np.nan)
    result["unit_share_%"] = np.where(total_sales > 0, result["total_sales"] / total_sales * 100, np.nan)
    if "total_revenue" in result.columns:
        result["total_revenue"] = result["total_revenue"].round(2)
    return result


def build_brand_summary(core: pd.DataFrame) -> pd.DataFrame:
    total_revenue = core["ASIN Revenue"].sum()
    total_sales = core["ASIN Sales"].sum()
    summary = (
        core.groupby("brand_clean", as_index=False)
        .agg(
            total_revenue=("ASIN Revenue", "sum"),
            total_sales=("ASIN Sales", "sum"),
            asin_count=("ASIN", "nunique"),
            avg_price=("Price", "mean"),
            median_price=("Price", "median"),
            avg_rating=("Reviews Rating", "mean"),
            total_reviews=("Review Count", "sum"),
        )
        .sort_values("total_revenue", ascending=False)
        .reset_index(drop=True)
    )
    return with_share_columns(summary, total_revenue, total_sales)


def build_product_type_summary(core: pd.DataFrame) -> pd.DataFrame:
    total_revenue = core["ASIN Revenue"].sum()
    total_sales = core["ASIN Sales"].sum()
    summary = (
        core.groupby("product_type", as_index=False)
        .agg(
            total_revenue=("ASIN Revenue", "sum"),
            total_sales=("ASIN Sales", "sum"),
            asin_count=("ASIN", "nunique"),
            avg_price=("Price", "mean"),
            avg_rating=("Reviews Rating", "mean"),
            total_reviews=("Review Count", "sum"),
        )
        .sort_values("total_revenue", ascending=False)
        .reset_index(drop=True)
    )
    return with_share_columns(summary, total_revenue, total_sales)


def build_price_tier_summary(core: pd.DataFrame) -> pd.DataFrame:
    total_revenue = core["ASIN Revenue"].sum()
    total_sales = core["ASIN Sales"].sum()
    tiers = core.dropna(subset=["price_tier"]).copy()
    summary = (
        tiers.groupby("price_tier", as_index=False, observed=False)
        .agg(
            total_revenue=("ASIN Revenue", "sum"),
            total_sales=("ASIN Sales", "sum"),
            asin_count=("ASIN", "nunique"),
            avg_price=("Price", "mean"),
            avg_rating=("Reviews Rating", "mean"),
            total_reviews=("Review Count", "sum"),
        )
        .sort_values("total_revenue", ascending=False)
        .reset_index(drop=True)
    )
    summary["price_tier"] = summary["price_tier"].astype(str)
    return with_share_columns(summary, total_revenue, total_sales)


def build_top_asins(core: pd.DataFrame) -> pd.DataFrame:
    return core.sort_values("ASIN Revenue", ascending=False).loc[:, TOP_ASIN_COLUMNS].head(50).reset_index(drop=True)


def build_innova_asins(core: pd.DataFrame) -> pd.DataFrame:
    return (
        core[core["is_innova"]]
        .sort_values("ASIN Revenue", ascending=False)
        .loc[:, TOP_ASIN_COLUMNS]
        .reset_index(drop=True)
    )


def feature_group_rows(source: pd.DataFrame, feature_name: str) -> list[dict[str, object]]:
    grouped = (
        source.groupby(feature_name, as_index=False)
        .agg(
            total_revenue=("ASIN Revenue", "sum"),
            total_sales=("ASIN Sales", "sum"),
            asin_count=("ASIN", "nunique"),
            avg_price=("Price", "mean"),
            avg_rating=("Reviews Rating", "mean"),
            total_reviews=("Review Count", "sum"),
        )
        .sort_values(feature_name)
        .reset_index(drop=True)
    )

    rows: list[dict[str, object]] = []
    for record in grouped.to_dict(orient="records"):
        row = {
            "is_large_screen_like": record[feature_name] if feature_name == "is_large_screen_like" else None,
            "total_revenue": round(float(record["total_revenue"]), 2),
            "total_sales": record["total_sales"],
            "asin_count": record["asin_count"],
            "avg_price": record["avg_price"],
            "avg_rating": record["avg_rating"],
            "total_reviews": record["total_reviews"],
            "feature_flag": feature_name,
            "is_rechargeable": record[feature_name] if feature_name == "is_rechargeable" else None,
            "is_automotive_targeted": record[feature_name] if feature_name == "is_automotive_targeted" else None,
        }
        rows.append(row)
    return rows


def build_features_all(core: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    rows.extend(feature_group_rows(core, "is_large_screen_like"))
    rows.extend(feature_group_rows(core, "is_rechargeable"))
    rows.extend(feature_group_rows(core, "is_automotive_targeted"))
    return pd.DataFrame(rows, columns=FEATURE_COLUMNS)


def build_features_auto_only(core: pd.DataFrame) -> pd.DataFrame:
    auto_only = core[core["is_automotive_targeted"]].copy()
    rows: list[dict[str, object]] = []
    rows.extend(feature_group_rows(auto_only, "is_large_screen_like"))
    rows.extend(feature_group_rows(auto_only, "is_rechargeable"))
    rows.extend(feature_group_rows(auto_only, "is_automotive_targeted"))
    return pd.DataFrame(rows, columns=FEATURE_COLUMNS)


def build_kpi_overview(core: pd.DataFrame) -> pd.DataFrame:
    overall = {
        "metric": "Overall automotive DMM market (last 30 days)",
        "total_revenue": round(float(core["ASIN Revenue"].sum()), 2),
        "total_sales": core["ASIN Sales"].sum(),
        "asin_count": core["ASIN"].nunique(),
        "avg_price": core["Price"].mean(),
        "median_price": core["Price"].median(),
        "avg_rating": core["Reviews Rating"].mean(),
        "median_rating": core["Reviews Rating"].median(),
    }

    innova = core[core["is_innova"]].copy()
    innova_row = {
        "metric": "INNOVA",
        "total_revenue": round(float(innova["ASIN Revenue"].sum()), 2),
        "total_sales": innova["ASIN Sales"].sum(),
        "asin_count": innova["ASIN"].nunique(),
        "avg_price": innova["Price"].mean(),
        "median_price": innova["Price"].median(),
        "avg_rating": innova["Reviews Rating"].mean(),
        "median_rating": innova["Reviews Rating"].median(),
    }

    return pd.DataFrame([overall, innova_row])


def write_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    if title in workbook.sheetnames:
        sheet = workbook[title]
        workbook.remove(sheet)
    worksheet = workbook.create_sheet(title)

    worksheet.append(list(frame.columns))
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    rows = frame.astype(object).where(pd.notnull(frame), None).values.tolist()
    for row in rows:
        worksheet.append(row)


def build_workbook(core: pd.DataFrame, output_path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(workbook, "auto_filtered_core", core)
    write_sheet(workbook, "brand_summary", build_brand_summary(core))
    write_sheet(workbook, "top_asins", build_top_asins(core))
    write_sheet(workbook, "innova_asins", build_innova_asins(core))
    write_sheet(workbook, "product_type", build_product_type_summary(core))
    write_sheet(workbook, "price_tiers", build_price_tier_summary(core))
    write_sheet(workbook, "features_all", build_features_all(core))
    write_sheet(workbook, "features_auto_only", build_features_auto_only(core))
    write_sheet(workbook, "kpi_overview", build_kpi_overview(core))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    raw = load_raw_csvs(args.input_dir)
    core = classify_core_market(raw)
    build_workbook(core, args.output)
    print(f"Wrote workbook: {args.output}")
    print(f"Core rows: {len(core)}")
    print(f"Revenue: {core['ASIN Revenue'].sum():.2f}")
    print(f"Units: {int(core['ASIN Sales'].sum())}")


if __name__ == "__main__":
    main()
