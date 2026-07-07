#!/usr/bin/env python3
from __future__ import annotations

import argparse
import importlib.util
import sys
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
BASE_FORMATTER = REPO_ROOT / "NewProductCategory" / "Borescope" / "apply_borescope_excel_formatting.py"


def load_base_formatter():
    spec = importlib.util.spec_from_file_location("base_amazon_product_formatter", BASE_FORMATTER)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Unable to load formatter from {BASE_FORMATTER}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def apply_formatting(path: Path) -> None:
    fmt = load_base_formatter()
    wb = load_workbook(path)

    if wb.calculation is not None:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True

    for ws in wb.worksheets:
        fmt.apply_title_row(ws, fmt.title_text_for_sheet(path, ws.title))

        if ws.title == "Summary":
            fmt.format_summary_sheet(ws)
        elif ws.title in {"Top 50 Revenue", "Top 50 Units"}:
            fmt.format_flat_sheet(ws, large=False)
        elif ws.title == "All ASINs":
            fmt.format_flat_sheet(ws, large=True)
        elif ws.title == "Top 50 Summary":
            fmt.format_stacked_summary(ws)
        elif ws.title == "Price Tiers":
            fmt.format_price_tiers(ws)
        elif ws.title == "Metadata":
            fmt.format_metadata(ws)
        else:
            fmt.format_flat_sheet(ws, large=ws.max_row > 500)

        fmt.finalize_sheet(ws)

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply Amazon product formatting to Thermal Imager workbooks.")
    parser.add_argument("files", nargs="+", type=Path, help="Workbook files to format in place.")
    args = parser.parse_args()

    for path in args.files:
        apply_formatting(path)
        print(f"Formatted {path}")


if __name__ == "__main__":
    main()
